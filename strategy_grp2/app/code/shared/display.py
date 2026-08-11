"""How a number LOOKS in every Excel artifact this project writes.

One module owns the presentation conventions, because they are a project-wide design
decision rather than a property of any single sheet (SM, 2026-08-11: *"I would prefer fixed
decimals, default 2. Except for integer attributes like rank and any counting variables, of
course. This should be a general design spec for data display across all of strategy_grp2"*).
Writers keep deciding WHAT to write; this decides how it reads.

Two conventions live here:

* **Fixed decimals, 2 by default.** Values are still stored at full precision — only the
  displayed form is fixed, so sorting and further arithmetic in Excel are unaffected. A
  group of numbers that is entirely whole (`rank`, `n_elites`, `chain_n`, a daynum, a
  1-based ordinal) is shown as an integer instead; that test is made on the data, not on a
  hand-maintained name list, so a factor added to the board later needs no edit here.
* **Alternating block tint.** A colour, not a rule under the last member, because a fill
  travels with its row when the user sorts the sheet and a border does not.
"""
from __future__ import annotations

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

DECIMALS = 2
DEC_FMT = "0." + "0" * DECIMALS
INT_FMT = "0"

#: alternate strategy blocks on the flat per-label sheets. Full-strength yellow, SM's own
#: pick (2026-08-11) over the pale tint first shipped — the band has to be readable at a
#: glance down a 460-row sheet, not merely present.
BAND_FILL = PatternFill("solid", fgColor="FFFF00")


def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def format_for(values) -> str:
    """`INT_FMT` when every number in the group is whole, `DEC_FMT` otherwise.

    Decided per group rather than per cell on purpose: judging each cell alone would print
    a genuine 0.0 in a decimal column as a bare `0`, which reads as a different kind of
    quantity than the 1.23 above it.
    """
    nums = [v for v in values if _is_number(v)]
    if not nums:
        return DEC_FMT
    return INT_FMT if all(float(v).is_integer() for v in nums) else DEC_FMT


def _apply(run: list) -> None:
    if not run:
        return
    fmt = format_for(c.value for c in run)
    for cell in run:
        cell.number_format = fmt


def harmonize(ws: Worksheet, by: str = "column") -> None:
    """Give every numeric cell a fixed-decimal format, decided per contiguous run.

    A run is a maximal stretch of consecutive numeric cells down a column — or across a row
    on a transposed sheet (`by="row"`, e.g. `Step3_compare`, where a metric is a row and a
    strategy is a column). Anything non-numeric — a header, a note, a blank — ends a run,
    which is what lets one pass serve the block-structured sheets (`Step3a_stopout`,
    `Step4_walkforward`) as well as the flat ones without being told where the blocks are:
    a block's own header row already separates its numbers from the block above.
    """
    lines = ws.iter_cols() if by == "column" else ws.iter_rows()
    for line in lines:
        run: list = []
        for cell in line:
            if _is_number(cell.value):
                run.append(cell)
            else:
                _apply(run)
                run = []
        _apply(run)


def harmonize_workbook(wb: Workbook, row_oriented: set[str] = frozenset()) -> None:
    """`harmonize` every sheet, transposed ones named in `row_oriented`."""
    for ws in wb.worksheets:
        harmonize(ws, by="row" if ws.title in row_oriented else "column")


def band(ws: Worksheet, blocks: list[tuple[int, int]], last_col: int,
         fill: PatternFill = BAND_FILL) -> None:
    """Tint every other `(top, bottom)` block, so one strategy's rows read as one thing.

    Cells that already carry a fill are left alone — a semantic colour (an elevated group,
    a flagged stop level) outranks the banding, which is only a reading aid.
    """
    for i, (top, bottom) in enumerate(blocks):
        if i % 2 == 0:
            continue
        for r in range(top, bottom + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(r, c)
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = fill
