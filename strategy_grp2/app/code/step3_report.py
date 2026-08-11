"""
Per-row operational detail: `report/backtesting/run<N>_<date>.xlsx`, one file per active
D-purpose row — the day-by-day hop history underneath Step3_compare's single-column summary.
Adapted from strategy_grp v1's shared/report.py Operational sheet to v2's fused hop timeline
(step3_backtest.Hop) and board-driven informational_attributes. Was on DesignVersion2.md's
deferred list; built 2026-08-03 per SM's spec.
"""

from __future__ import annotations

import shutil
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import step3_backtest as bt
from shared import display
from shared import market
from shared.config import REPORT_ROOT
from shared.data_loader import daynum_to_date, load_longi
from shared.datacheck import DataUnavailable

_BOLD = Font(bold=True)
_SMALL = Font(size=9)
_HDR_FILL = PatternFill("solid", fgColor="BDD7EE")
_GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_GRY_FILL = PatternFill("solid", fgColor="EEEEEE")
_CAND_FILL = PatternFill("solid", fgColor="E2EFDA")
_BENCH_FILL = PatternFill("solid", fgColor="F2F2F2")
_CTR = Alignment(horizontal="center")
_PCT_FMT = '+0.00;-0.00;"-"'

# Open (still-realizing) hops get their own header/gain colouring — the same palette
# strategy_grp v1's separate extension.py sheet used — so it stays visually obvious that the
# newest ~period columns are a partial snapshot, not a closed result like the rest of the
# timeline. Fused into one sheet in v2 (see step3_backtest.py), so colour is what now carries
# the distinction v1 got for free by using two different files.
_OPEN_HDR_FILL = PatternFill("solid", fgColor="DDEBF7")   # lighter blue — open/extension day
_EXT_POS_FILL = PatternFill("solid", fgColor="FFFF99")    # light yellow — open, positive gain
_EXT_NEG_FILL = PatternFill("solid", fgColor="FFC000")    # orange — open, negative gain
_STOP_FILL = PatternFill("solid", fgColor="C00000")       # dark red — Step 3a exited this ticker


def _gain_fill(val: float, realized: bool = True) -> PatternFill:
    if pd.isna(val):
        return _GRY_FILL
    if realized:
        return _GRN_FILL if val >= 0 else _RED_FILL
    return _EXT_POS_FILL if val >= 0 else _EXT_NEG_FILL


def _gspc(h: "bt.Hop") -> float:
    return h.ref_values.get("^GSPC_rsi", float("nan"))


def _write_operational(ws, hops: list["bt.Hop"], label: str, params: dict) -> None:
    """
    Row 1: label                | daynum1 | daynum2 | ...  (newest-left)
    Row 2: (blank)              | date1   | date2   | ...
    Row 3: n_candidates         | pool size before picking, per hop
    Row 4: dominance_cutoff     | that day's Step-1 decile cutoff
    Row 5: n_stopped            | tickers Step 3a exited this hop (0 when stop_loss is off)
    Rows 6..6+n-1: ticker_rank1..N  (n = focusset_size) — a Step-3a-exited ticker gets a
                   dark-red fill (h.stopped), so the timeline shows WHERE a stop bit, not
                   just the count
    +1: avg_gain
    +2: mkt_gain, alpha  (+beta when longi_beta3m.csv is available)
    +len(attrs)*n: informational_attributes per-stock rows, from the row's own board column
    +len(ref_values): market context (^GSPC_rsi, ^STOXX_rsi, ...)

    The newest columns (h.realized=False — still-realizing, partial-return hops, see
    step3_backtest.py's fused timeline) get a different header fill and yellow/orange
    gain colouring instead of green/red on avg_gain/mkt_gain/alpha, matching v1's separate
    extension.py sheet — a visual cue that these are a snapshot-so-far, not a closed result.
    """
    n = int(params["focusset_size"])
    threshold = params.get("no_go_gspc_rsi")

    label_row, date_row, cand_row, cutoff_row, stopped_row = 1, 2, 3, 4, 5
    ticker_top = 6
    avg_row = ticker_top + n
    mkt_row, alpha_row = avg_row + 1, avg_row + 2
    has_beta = any(pd.notna(h.beta) for h in hops)
    beta_row = alpha_row + 1 if has_beta else None
    bench_bottom = (beta_row or alpha_row) + 1

    attrs = [a for a in (params.get("informational_attributes") or ())]
    info_top = bench_bottom
    ref_top = info_top + n * len(attrs)

    # ---- header rows (open/extension days get a distinct fill) ----
    ws.cell(label_row, 1, label).font = _BOLD
    for j, h in enumerate(hops, start=2):
        hdr_fill = _HDR_FILL if h.realized else _OPEN_HDR_FILL
        c = ws.cell(label_row, j, h.daynum); c.font, c.fill, c.alignment = _BOLD, hdr_fill, _CTR
        c = ws.cell(date_row, j, daynum_to_date(h.daynum)); c.font, c.fill, c.alignment = _SMALL, hdr_fill, _CTR

    # ---- n_candidates row (NEW) ----
    lbl = ws.cell(cand_row, 1, "n_candidates"); lbl.font, lbl.fill = _BOLD, _CAND_FILL
    for j, h in enumerate(hops, start=2):
        cell = ws.cell(cand_row, j, h.n_candidates or None)
        cell.fill, cell.alignment = _CAND_FILL, _CTR

    # ---- dominance_cutoff row ----
    lbl = ws.cell(cutoff_row, 1, "dominance_cutoff"); lbl.font, lbl.fill = _BOLD, _CAND_FILL
    for j, h in enumerate(hops, start=2):
        val = h.dom_cutoff
        cell = ws.cell(cutoff_row, j, round(val, 2) if pd.notna(val) else None)
        cell.fill, cell.alignment = _CAND_FILL, _CTR

    # ---- n_stopped row (NEW) ----
    lbl = ws.cell(stopped_row, 1, "n_stopped"); lbl.font, lbl.fill = _BOLD, _CAND_FILL
    for j, h in enumerate(hops, start=2):
        cell = ws.cell(stopped_row, j, len(h.stopped) or None)
        cell.fill, cell.alignment = _CAND_FILL, _CTR

    # ---- ticker rows (a Step-3a-exited ticker gets a dark-red fill) ----
    for i in range(n):
        row = ticker_top + i
        for j, h in enumerate(hops, start=2):
            if i >= len(h.tickers):
                continue
            ticker = h.tickers[i]
            cell = ws.cell(row, j, ticker)
            if ticker in h.stopped:
                cell.fill = _STOP_FILL
                cell.font = Font(color="FFFFFF")

    # ---- avg_gain row ----
    lbl = ws.cell(avg_row, 1, "avg_gain"); lbl.font = _BOLD
    for j, h in enumerate(hops, start=2):
        no_go = threshold is not None and pd.notna(_gspc(h)) and _gspc(h) < threshold
        val = market.hop_avg(h.gains)
        cell = ws.cell(avg_row, j)
        cell.font, cell.number_format, cell.alignment = _BOLD, _PCT_FMT, _CTR
        if no_go or pd.isna(val):
            cell.value, cell.fill = None, _GRY_FILL
        else:
            cell.value, cell.fill = round(val, 4), _gain_fill(val, h.realized)

    # ---- benchmark rows: mkt_gain, alpha, (beta) ----
    ws.cell(mkt_row, 1, "mkt_gain").font = _BOLD; ws.cell(mkt_row, 1).fill = _BENCH_FILL
    ws.cell(alpha_row, 1, "alpha").font = _BOLD; ws.cell(alpha_row, 1).fill = _BENCH_FILL
    if beta_row:
        ws.cell(beta_row, 1, "beta").font = _BOLD; ws.cell(beta_row, 1).fill = _BENCH_FILL
    for j, h in enumerate(hops, start=2):
        no_go = threshold is not None and pd.notna(_gspc(h)) and _gspc(h) < threshold
        g = market.hop_avg(h.gains)
        a = g - h.mkt_gain if pd.notna(g) and pd.notna(h.mkt_gain) else float("nan")
        rows = [(mkt_row, h.mkt_gain, False), (alpha_row, a, False)]
        if beta_row:
            rows.append((beta_row, h.beta, True))
        for row, val, is_beta in rows:
            cell = ws.cell(row, j)
            cell.alignment = _CTR
            if pd.isna(val):
                cell.value, cell.fill = None, _GRY_FILL
                continue
            if is_beta:
                cell.number_format = "0.00"
                cell.value, cell.fill = round(val, 2), _BENCH_FILL
                continue
            cell.number_format = _PCT_FMT
            cell.value = None if no_go else round(val, 4)
            cell.fill = _GRY_FILL if no_go else _gain_fill(val, h.realized)

    # ---- informational_attributes per-stock rows (board-driven, not hardcoded) ----
    for k, attr in enumerate(attrs):
        try:
            info_df = load_longi(f"longi_{attr}.csv")
        except DataUnavailable as exc:
            print(f"[step3_report] {label}: informational_attributes entry "
                  f"{attr!r} not found, skipped -- {exc}", flush=True)
            continue
        for i in range(n):
            row = info_top + n * k + i
            ws.cell(row, 1, f"{attr}_{i + 1}").font = _BOLD
            for j, h in enumerate(hops, start=2):
                cell = ws.cell(row, j); cell.alignment = _CTR
                if i >= len(h.tickers):
                    continue
                t, col = h.tickers[i], str(h.daynum)
                if t in info_df.index and col in info_df.columns and pd.notna(info_df.at[t, col]):
                    cell.value = round(float(info_df.at[t, col]), 2)

    # ---- market-context ref rows ----
    ref_keys = list(hops[0].ref_values.keys()) if hops else []
    for idx, key in enumerate(ref_keys):
        row = ref_top + idx
        ws.cell(row, 1, key).font = _BOLD
        for j, h in enumerate(hops, start=2):
            val = h.ref_values.get(key, float("nan"))
            cell = ws.cell(row, j); cell.alignment = _CTR
            if pd.notna(val):
                cell.value = round(val, 2)

    ws.column_dimensions["A"].width = 22
    for j in range(2, len(hops) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.row_dimensions[1].height = 16
    ws.freeze_panes = "B5"
    # transposed sheet: a metric is a ROW, a hop is a column (shared/display.py)
    display.harmonize(ws, by="row")


def _clear_stale(folder: Path) -> None:
    """Archive whatever is already in report/backtesting/ before writing this tick's files —
    matches outputboard's own _archive_prior rule. Also sweeps up any pre-v2 leftovers."""
    for path in folder.glob("*.xlsx"):
        dest = REPORT_ROOT / "_archive" / time.strftime("%Y%m%d_%H%M%S")
        dest.mkdir(parents=True, exist_ok=True)
        shutil.move(str(path), str(dest / path.name))


def write_run_reports(backtests: dict[str, "bt.BacktestResult"]) -> list[Path]:
    """One report/backtesting/run<N>_<date>.xlsx per active D-purpose row, N = tick order."""
    folder = REPORT_ROOT / "backtesting"
    folder.mkdir(parents=True, exist_ok=True)
    _clear_stale(folder)

    today = time.strftime("%Y%m%d")
    written: list[Path] = []
    for i, (label, result) in enumerate(backtests.items(), start=1):
        wb = Workbook()
        ws = wb.active
        ws.title = "Operational"
        _write_operational(ws, result.hops, label, result.params)
        path = folder / f"run{i}_{today}.xlsx"
        wb.save(path)
        written.append(path)
    return written
