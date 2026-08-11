"""
The pipeline driver — longi.py's role for this project. Reads the control board, and for
every active row runs steps 0-2 (purpose=P) or 0-4 (purpose=D), then hands the results to
outputboard.py. See DesignVersion2.md for the full step write-up.

    python conductor.py --make-board   # write/refresh the board from param_spec.py
    python conductor.py --dry-run      # parse + validate every row; touches no data
    python conductor.py --check        # step 0 only: universe/group counts, data guard
    python conductor.py                # development tick: steps 0-4 for active rows -> compare_strategies_<date>.xlsx
    python conductor.py --production   # steps 0-2 for EVERY active row (P and D) -> StrategicStocks.xlsx
"""

from __future__ import annotations

import shutil
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import control_board
import outputboard
import preflight
import step0_data
import step2_focusset
from shared import display
from shared import expression as expr
from shared.config import REPORT_ROOT
from shared.data_loader import daynum_to_date, load_stamdata

_HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")


def _report_board_freshness() -> None:
    """Print when the board on disk was last saved. Cheap, and it is the one line that
    makes 'I forgot to save' visible after the fact as well as before."""
    state = control_board.board_file_state()
    open_note = " — STILL OPEN IN EXCEL" if state.is_open else ""
    print(f"control_board.xlsx last saved: {state.saved_ago}{open_note}\n")


def cmd_make_board() -> int:
    path = control_board.write_board()
    print(f"Wrote {path}")
    return 0


def cmd_dry_run() -> int:
    result = control_board.read_board()

    if result.board_errors:
        print("Board-level errors:")
        for msg in result.board_errors:
            print(f"  ! {msg}")

    for row in result.runs:
        tag = "active" if row.active else "  idle"
        label = row.resolved.get("label", "?")
        if row.ok:
            print(f"[{tag}] row {row.row_num:>3}  {label:<28} OK  "
                  f"purpose={row.resolved.get('purpose')} "
                  f"level={row.resolved.get('level')} "
                  f"group={row.resolved.get('group_expression')!r}")
        else:
            print(f"[{tag}] row {row.row_num:>3}  {label:<28} FAILED")
            for msg in row.errors:
                print(f"           ! {msg}")

    n_active = len(result.active_runs)
    n_active_bad = sum(1 for r in result.active_runs if not r.ok)
    print(f"\n{len(result.runs)} row(s) total, {n_active} active, "
          f"{n_active_bad} active row(s) with errors.")

    if not result.ok:
        print("\n--dry-run: at least one active row has errors, or the board itself does "
              "(see above). Fix the board and re-run.")
        return 1
    print("\n--dry-run: every active row parses cleanly. Nothing else was resolved "
          "(no data read, no files written).")
    return 0


# ---------------------------------------------------------------------------
# Step 0 diagnostics — `--check`
# ---------------------------------------------------------------------------

def cmd_check() -> int:
    board = control_board.read_board()
    active = [r for r in board.runs if r.active and r.ok]
    if not active:
        print("No active, cleanly-parsing rows to check.")
        return 0

    preflight.ensure_data(board.runs, settings=board.settings)

    bad = 0
    for row in active:
        label = row.resolved.get("label")
        try:
            s0 = step0_data.resolve_step0(row.resolved)
        except expr.ExpressionError as exc:
            print(f"[{label}] STEP0 FAILED: {exc}")
            bad += 1
            continue
        n_groups = len(s0.group_sizes)
        print(f"[{label}] universe={len(s0.universe)} tickers, {n_groups} group(s), "
              f"dominance_attribute={s0.dominance_attribute!r}, "
              f"priority_attribute={s0.priority_attribute!r}")
        if n_groups <= 12:
            sizes = ", ".join(f"{k}={v}" for k, v in sorted(s0.group_sizes.items()))
            print(f"           group sizes: {sizes}")

    print(f"\n{len(active)} row(s) checked, {bad} failed.")
    return 1 if bad else 0


# ---------------------------------------------------------------------------
# Production tick — `--production`, steps 0-2 -> StrategicStocks.xlsx
# ---------------------------------------------------------------------------

def _production_pick(row: "control_board.RunRow"):
    """(label, daynum, tickers, params, s0) for one active row's current gross list."""
    daynum, tickers, _elevated, params, s0 = step2_focusset.current_pick(row.resolved)
    return row.resolved.get("label"), daynum, tickers, params, s0


def _archive_prior_strategic_stocks() -> None:
    path = REPORT_ROOT / "StrategicStocks.xlsx"
    if not path.exists():
        return
    dest = REPORT_ROOT / "_archive" / time.strftime("%Y%m%d_%H%M%S")
    dest.mkdir(parents=True, exist_ok=True)
    shutil.move(str(path), str(dest / path.name))


def _write_strategic_stocks(picks: list[tuple]) -> Path:
    stamdata = load_stamdata()
    wb = Workbook()
    wb.remove(wb.active)
    for label, daynum, tickers, params, _s0 in picks:
        ws = wb.create_sheet(label[:31])   # Excel's 31-char sheet-name limit
        ws.cell(row=1, column=1, value=f"As of daynum {daynum} ({daynum_to_date(daynum)})")
        headers = ["Rank", "Ticker", "Name", params["priority_attribute"]]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
        prio_df = None
        try:
            from shared.data_loader import load_longi
            prio_df = load_longi(f"longi_{params['priority_attribute']}.csv")
        except Exception:                                          # noqa: BLE001
            prio_df = None
        for i, ticker in enumerate(tickers, start=1):
            r = 2 + i
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=ticker)
            name = stamdata.at[ticker, "Name"] if ticker in stamdata.index and "Name" in stamdata.columns else ""
            ws.cell(row=r, column=3, value=name)
            val = None
            if prio_df is not None and ticker in prio_df.index and str(daynum) in prio_df.columns:
                val = prio_df.at[ticker, str(daynum)]
            ws.cell(row=r, column=4, value=val)
        for col, width in zip("ABCD", (6, 10, 32, 14)):
            ws.column_dimensions[col].width = width

    display.harmonize_workbook(wb)
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    path = REPORT_ROOT / "StrategicStocks.xlsx"
    wb.save(path)
    return path


def cmd_production() -> int:
    board = control_board.read_board()
    active = [r for r in board.runs if r.active and r.ok]
    if not active:
        print("No active rows.")
        return 0

    preflight.ensure_data(board.runs, settings=board.settings)

    picks = []
    bad = 0
    for row in active:
        label = row.resolved.get("label")
        try:
            result = _production_pick(row)
        except (expr.ExpressionError, ValueError) as exc:
            print(f"[{label}] FAILED: {exc}")
            bad += 1
            continue
        _, daynum, tickers, params, _s0 = result
        if not tickers:
            print(f"[{label}] no pick at daynum {daynum} (cash hop)")
        else:
            print(f"[{label}] daynum {daynum}: {tickers}")
        picks.append(result)

    if not picks:
        print("\nNo row produced a result — StrategicStocks.xlsx not written.")
        return 1

    _archive_prior_strategic_stocks()
    path = _write_strategic_stocks(picks)
    print(f"\nWrote {path}  ({len(picks)} row(s), {bad} failed)")
    return 1 if bad else 0


def cmd_develop() -> int:
    board = control_board.read_board()
    active = [r for r in board.runs if r.active and r.ok]
    if not active:
        rejected = [r for r in board.runs if r.active and not r.ok]
        print(f"No active, cleanly-parsing rows ({len(rejected)} active row(s) rejected).")
        for row in rejected:
            print(f"  ! row {row.row_num} '{row.resolved.get('label')}': "
                  f"{'; '.join(row.errors)}")
        for msg in board.board_errors:
            print(f"  ! BOARD: {msg}")
        return 1 if rejected or board.board_errors else 0

    preflight.ensure_data(board.runs, settings=board.settings)
    path = outputboard.assemble(board, board.settings)
    print(f"Wrote {path}")
    return 0


def main() -> int:
    args = set(sys.argv[1:])

    commands = {
        "--make-board": (cmd_make_board, "--make-board (it would overwrite your open copy)"),
        "--dry-run": (cmd_dry_run, "--dry-run"),
        "--check": (cmd_check, "--check"),
        "--production": (cmd_production, "the production tick"),
    }
    fn, action = cmd_develop, "the development tick"
    for flag, (candidate_fn, candidate_action) in commands.items():
        if flag in args:
            fn, action = candidate_fn, candidate_action
            break

    # Every entry point reads the board, so every entry point checks it was saved first.
    try:
        control_board.require_saved_board(action, allow_open="--board-open-ok" in args)
    except control_board.BoardOpenError as exc:
        print(f"\n*** {exc}\n")
        return 2
    _report_board_freshness()
    return fn()


if __name__ == "__main__":
    sys.exit(main())
