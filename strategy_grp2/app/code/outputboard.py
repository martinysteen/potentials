"""
The output board — one workbook per development tick,
`report/compare_strategies_<date>.xlsx`. Separates results by process step, per
DesignVersion2.md's input/output separation principle:

    Runs               every active row, verbatim, beside its status
    Step1_groups       elevated groups + member tickers, current daynum
    Step2_picks        the gross list, current daynum (what production also ships)
    Step3_compare      transposed comparison: metric rows x one column per active row
    Step4_walkforward  per distinct candidate set: pooled summary, per-candidate
                       comparison, fold table
    Charts             cumulative chain / IS-vs-OOS / avg-vs-median gain, per row

Prior dated workbooks move to `_archive/` only once new output exists (matches
strategy_grp v1's rule for its own combined report).

Development ticks also write `report/picks_<daynum>.csv` (2026-08-13, long-format since
2026-08-14, see write_picks_csv) — one row per (label, daynum, ticker) pick across every
active `D` row's ENTIRE dominance history, not just today. Separate file, not a sheet
here, since it is a different grain (full history, not just the current daynum) from the
shared Step1/Step2 tables above. SM, 2026-08-14, switching it from the original wide xlsx
pivot to long-format CSV: "I will then import this in as needed Excel and pivot the long
format there."
"""

from __future__ import annotations

import shutil
import time
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import control_board as cb
import param_spec as spec
import step1_dominance
import step2_focusset
import step3_backtest as bt
import step3_report
import step3a_stopout as stopout
import step4_walkforward as wf
from shared import display
from shared import expression as expr
from shared import market
from shared.config import MINAGGR_PERIODS, REPORT_ROOT
from shared.data_loader import load_longi, load_stamdata
from shared.datacheck import DataUnavailable

_BOLD = Font(bold=True)
_HEAD = PatternFill("solid", fgColor="BDD7EE")
_ERR_FILL = PatternFill("solid", fgColor="FFC7CE")
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")
_THIN_FILL = PatternFill("solid", fgColor="F8CBAD")
_NOTE = Font(italic=True, size=9, color="595959")


# ---------------------------------------------------------------------------
# Runs / Step1_groups / Step2_picks — current-daynum snapshot, every active row
# ---------------------------------------------------------------------------

def current_picks(active_rows: list["cb.RunRow"]) -> dict[str, dict]:
    """label -> {daynum, tickers, elevated, params, s0, error} for every active row.

    Public: shared with conductor.cmd_production, which needs the same steps 0-2
    resolution (and error handling) as the development tick's Runs/Step1_groups/
    Step2_picks sheets, just skipping steps 3/4 and writing a different output file
    (StrategicStocks.xlsx, see step2_table/write_strategic_stocks below)."""
    out: dict[str, dict] = {}
    n = len(active_rows)
    for i, row in enumerate(active_rows, start=1):
        label = row.resolved.get("label")
        print(f"[{i}/{n}] {label}: steps 0-2 ...", flush=True)
        try:
            daynum, tickers, elevated, params, s0 = step2_focusset.current_pick(row.resolved)
            out[label] = {"daynum": daynum, "tickers": tickers, "elevated": elevated,
                         "params": params, "s0": s0, "error": None}
            print(f"    universe={len(s0.universe)} tickers, {len(s0.group_sizes)} group(s), "
                  f"{len(tickers)} pick(s) at daynum {daynum}", flush=True)
        except (expr.ExpressionError, ValueError) as exc:
            out[label] = {"error": str(exc)}
            print(f"    FAILED: {exc}", flush=True)
    return out


def _write_runs_sheet(wb: Workbook, active_rows: list["cb.RunRow"],
                     picks: dict[str, dict], rejected_rows: list["cb.RunRow"],
                     board_errors: list[str]) -> None:
    """Every row the board marked `D` — INCLUDING the ones a tick could not run (2026-08-13:
    was every `active` row; `D` and `P` are now independent, and this sheet belongs to the
    development tick, so it's `D` rows only — see assemble()).

    A rejected row used to be filtered out in conductor.cmd_develop and never mentioned
    again, so the board said 4 active and the workbook showed 3 with nothing to say where
    the fourth went (SM, 2026-08-04). It now appears with status=REJECTED and the parse
    error that rejected it, and board-level errors (duplicate labels, bad Settings) get a
    banner above the table.
    """
    ws = wb.create_sheet("Runs")
    r = 1
    for msg in board_errors:
        cell = ws.cell(r, 1, f"BOARD ERROR: {msg}")
        cell.font, cell.fill = _BOLD, _ERR_FILL
        r += 1
    if board_errors:
        r += 1

    header_row = r
    headers = list(spec.RUNS_COLUMNS) + ["status", "universe_size", "n_groups", "daynum", "error"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, h)
        cell.font, cell.fill = _BOLD, _HEAD
    r += 1

    for row in list(active_rows) + list(rejected_rows):
        label = row.resolved.get("label")
        for c, name in enumerate(spec.RUNS_COLUMNS, start=1):
            ws.cell(r, c, spec.format_value(name, row.resolved.get(name)))
        info = picks.get(label, {})
        if not row.ok:
            status, err = "REJECTED", "; ".join(row.errors)
        elif info.get("error"):
            status, err = "FAILED", info["error"]
        else:
            status, err = "OK", None
        status_cell = ws.cell(r, len(spec.RUNS_COLUMNS) + 1, status)
        status_cell.fill = _OK_FILL if status == "OK" else _ERR_FILL
        if status == "OK":
            ws.cell(r, len(spec.RUNS_COLUMNS) + 2, len(info["s0"].universe))
            ws.cell(r, len(spec.RUNS_COLUMNS) + 3, len(info["s0"].group_sizes))
            ws.cell(r, len(spec.RUNS_COLUMNS) + 4, info["daynum"])
        else:
            ws.cell(r, len(spec.RUNS_COLUMNS) + 5, err)
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16


_ELITE_LIST_CAP = 60


def _write_step1_sheet(wb: Workbook, picks: dict[str, dict]) -> None:
    """One row per group -- EVERY group in the row's grouping, with the elevated ones marked
    (SM, 2026-08-10). Elevated-only hid the half of the picture that explains the other half:
    a group missing by one elite and a group with none at all both showed as simply absent.
    Concretely, an `FKgrade` row elevated 4 of 6 grades and nothing on the sheet said grade
    `1` had 8 elites against a threshold of 10, nor that grade `4` cleared that same bar at a
    LOWER elite density (9.6% of 270 members vs 11.1% of 72) -- both are properties of the
    absolute-count rule, and both are now readable straight off the table.

    Columns: `dom_group` (renamed from `group`, SM 2026-08-05 -- "dominance" is what elevates
    it), `n_members` (gross Stamdata group size, unchanged), then `n_elites` /
    `dom_threshold` / `dom_today` from step1_dominance.group_status() -- the arithmetic Step 1
    actually ran, not a re-derivation (see its docstring for why the threshold is computed
    from attribute-file coverage rather than gross membership); `elevated`, the
    level-resolved verdict, identical to `dom_today` for a level-A row and deliberately not
    for level B/C where persistence over a trailing window decides; and `elite_members`, the
    group's qualifying-ticker roster sorted best-first (no numbering prefix -- SM, 2026-08-05:
    "my eyes get tired of all the extra numbers"). Not the gross membership: "I am absolutely
    not interested in gross members but a list of elite members" (SM).

    Elevated groups sort first, each block by descending `n_elites`, so near-misses sit
    directly under the line they failed to clear. `n_elites` summed down one label's block is
    now the universe's whole elite population (~10% of it, per `dominance_decile`), which it
    was not while only elevated groups had rows."""
    ws = wb.create_sheet("Step1_groups")
    headers = ["label", "daynum", "dom_group", "n_members", "n_elites", "dom_threshold",
               "dom_today", "elevated", "elite_members"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h); cell.font, cell.fill = _BOLD, _HEAD
    r = 2
    blocks: list[tuple[int, int]] = []
    for label, info in picks.items():
        block_top = r
        if info.get("error"):
            ws.cell(r, 1, label)
            ws.cell(r, 3, f"(step 0-2 failed: {info['error']})").font = _NOTE
            r += 1
            blocks.append((block_top, r - 1))
            continue
        s0 = info["s0"]
        params = info["params"]
        elevated = set(info["elevated"])
        if not elevated:
            # An absent label used to be indistinguishable from a label that ran fine and
            # elevated nothing today — the second is a normal, informative outcome for a
            # strict level-B/C row, not a missing result. Kept as its own line even now that
            # every group is listed below it: an all-blank `elevated` column is a weaker
            # statement than one that says so.
            ws.cell(r, 1, label)
            ws.cell(r, 2, info["daynum"])
            ws.cell(r, 3, "(no group elevated at this daynum)").font = _NOTE
            r += 1
        status = step1_dominance.group_status(
            s0.groups, params["dominance_attribute"], params["dominance_direction"],
            params["dominance_decile"], params["dom_count_min"], info["daynum"])
        elite_by_group = step1_dominance.elite_members(
            s0.groups, params["dominance_attribute"], params["dominance_direction"],
            params["dominance_decile"], info["daynum"])
        groups = sorted(s0.group_sizes,
                        key=lambda g: (g not in elevated, -len(elite_by_group.get(g, [])), g))
        for group in groups:
            elites = elite_by_group.get(group, [])
            elite_list_str = ", ".join(elites[:_ELITE_LIST_CAP])
            if len(elites) > _ELITE_LIST_CAP:
                elite_list_str += f" ... (+{len(elites) - _ELITE_LIST_CAP} more)"
            ws.cell(r, 1, label)
            ws.cell(r, 2, info["daynum"])
            ws.cell(r, 3, group)
            ws.cell(r, 4, int(s0.group_sizes.get(group, 0)))
            ws.cell(r, 5, len(elites))
            if group in status.index:
                ws.cell(r, 6, float(status.at[group, "dom_threshold"]))
                ws.cell(r, 7, "x" if bool(status.at[group, "dom_today"]) else "")
            ws.cell(r, 8, "x" if group in elevated else "")
            if group in elevated:
                ws.cell(r, 8).fill = _OK_FILL
                ws.cell(r, 3).font = _BOLD
            ws.cell(r, 9, elite_list_str)
            r += 1
        blocks.append((block_top, r - 1))
    ws.freeze_panes = "A2"
    display.band(ws, blocks, len(headers))
    for c, w in zip("ABCDEFGHI", (24, 10, 16, 11, 9, 13, 10, 9, 90)):
        ws.column_dimensions[c].width = w


def _fingerprint_attributes(picks: dict[str, dict]) -> list[str]:
    """Every Longi factor named anywhere on the active board, deduplicated and sorted.

    These four channels — dominance, priority, informational, post_filter — are exactly the
    ones `preflight.required_files_for_rows()` snapshots, so a name collected here is
    guaranteed readable. Names come off `Step0Result`, i.e. already twin-resolved
    (`conf` -> `conf_GICS`), and only from rows that actually resolved: a row that could not
    bind a twin is an `error` entry and contributes nothing.
    """
    names: set[str] = set()
    for info in picks.values():
        if info.get("error"):
            continue
        s0 = info["s0"]
        names.add(s0.dominance_attribute)
        names.add(s0.priority_attribute)
        names.update(s0.informational_attributes or ())
        if s0.post_filter is not None:
            names.update(t.column for t in s0.post_filter.terms)
    return sorted(n for n in names if n)


# `priority` (C) is the 1-based position in the gross list — it was called `rank` until
# 2026-08-11, which collided head-on with the `rank` ATTRIBUTE (the schema default for
# priority_attribute), putting an ordinal and a longi_rank.csv value on one line under one
# name. `name` (E, 2026-08-12) is Stamdata's company name, next to the ticker it names.
_STEP2_HEADERS_BASE: list[str] = [
    "label", "daynum", "priority", "ticker", "name", "dom_group", "dom/prio",
]


def step2_table(picks: dict[str, dict]) -> tuple[list[str], dict[str, list[list]]]:
    """(headers, {label: [row, row, ...]}) — the Step2_picks columns and values, one row
    per pick, computed once. Shared by the Step2_picks sheet and StrategicStocks.xlsx
    (2026-08-12, SM: "export all fields from Step2_picks" into production's own workbook)
    so both read off one computation rather than two that can drift apart. A label with no
    result (error, or an empty/cash pick) maps to an empty row list — callers decide how
    that reads on their own sheet (Step2_picks prints a note line; StrategicStocks.xlsx
    just leaves the tab's data area empty).

    Columns F/G onward are the pick's FINGERPRINT (SM 2026-08-11): one column per attribute
    NAME, so a column always means exactly one thing. They are the board-wide union across
    every label in `picks`, filled for every pick regardless of which row named the
    attribute — that is what lets an rsi-priority row and a rank-priority row be read side
    by side. Which attribute played which role is stated once per line in `dom/prio`; the
    other columns carry no role marking, deliberately.
    """
    stamdata = load_stamdata()
    attrs = _fingerprint_attributes(picks)
    headers = list(_STEP2_HEADERS_BASE) + attrs

    frames: dict[str, pd.DataFrame] = {}
    for attr in attrs:
        try:
            frames[attr] = load_longi(f"longi_{attr}.csv")
        except DataUnavailable as exc:
            print(f"[Step2_picks] attribute {attr!r} not readable, column left blank -- {exc}",
                  flush=True)

    out: dict[str, list[list]] = {}
    for label, info in picks.items():
        if info.get("error") or not info.get("tickers"):
            out[label] = []
            continue
        s0 = info["s0"]
        tickers = info["tickers"]
        col = str(info["daynum"])
        roles = f"{s0.dominance_attribute}/{s0.priority_attribute}"

        # one vectorized take per attribute per label, not one .at per cell
        values: dict[str, pd.Series] = {}
        for attr, df in frames.items():
            if col not in df.columns:
                continue
            values[attr] = df.loc[df.index.intersection(tickers), col].dropna()

        rows: list[list] = []
        for i, ticker in enumerate(tickers, start=1):
            name = (stamdata.at[ticker, "Name"]
                   if ticker in stamdata.index and "Name" in stamdata.columns else "")
            row = [label, info["daynum"], i, ticker, name, s0.groups.get(ticker), roles]
            for attr in attrs:
                series = values.get(attr)
                row.append(float(series.at[ticker])
                          if series is not None and ticker in series.index else None)
            rows.append(row)
        out[label] = rows
    return headers, out


def size_step2_columns(ws, headers: list[str]) -> None:
    """Column widths for a Step2_table sheet — shared by Step2_picks and every
    StrategicStocks.xlsx tab so the two read the same regardless of which wrote them."""
    for c, w in zip("ABCDEFG", (24, 10, 8, 12, 24, 12, 18)):
        ws.column_dimensions[c].width = w
    first_attr_col = len(_STEP2_HEADERS_BASE) + 1
    for k in range(len(headers) - len(_STEP2_HEADERS_BASE)):
        ws.column_dimensions[get_column_letter(first_attr_col + k)].width = 11


def _write_step2_sheet(wb: Workbook, picks: dict[str, dict]) -> None:
    """One row per pick, via step2_table() above. `dom_group` (SM 2026-08-05 — "just to
    make sure" each ticker's own dominant group is visible next to it, not just inferable
    from Step1_groups) is the elevated group `production_pick()` actually drew that ticker
    from."""
    ws = wb.create_sheet("Step2_picks")
    headers, rows_by_label = step2_table(picks)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h); cell.font, cell.fill = _BOLD, _HEAD

    r = 2
    blocks: list[tuple[int, int]] = []
    for label, info in picks.items():
        block_top = r
        if info.get("error"):
            ws.cell(r, 1, label)
            ws.cell(r, 4, f"(step 0-2 failed: {info['error']})").font = _NOTE
            r += 1
            blocks.append((block_top, r - 1))
            continue
        if not info["tickers"]:
            ws.cell(r, 1, label)
            ws.cell(r, 2, info["daynum"])
            ws.cell(r, 4, "(no pick — cash hop)").font = _NOTE
            r += 1
            blocks.append((block_top, r - 1))
            continue
        for row_vals in rows_by_label[label]:
            for c, val in enumerate(row_vals, start=1):
                ws.cell(r, c, val)
            r += 1
        blocks.append((block_top, r - 1))
    ws.freeze_panes = "A2"
    display.band(ws, blocks, len(headers))
    size_step2_columns(ws, headers)


# ---------------------------------------------------------------------------
# Step3_compare — transposed comparison, one column per active row
# ---------------------------------------------------------------------------

# The parameter block leads with the three settings that decide WHICH END of a ranking a
# row picks from — dominance_direction, priority_direction and from_rank are independent
# and a mismatch between them silently inverts the strategy (SM, 2026-08-04: "small_wins,
# but from_rank -1"), so they belong side by side where a comparison can catch it.
_STEP3_ROWS: list[str] = [
    "group_expression", "level", "period",
    "dominance_attribute", "dominance_direction",
    "priority_attribute", "priority_direction",
    "from_rank", "focusset_size", "tickers_per_group", "post_filter",
    "StartDaynum", "EndDaynum", "N_hops", "N_hops_active",
    "avg_gain", "median_gain", "hit_rate%", "avg_alpha", "avg_beta",
    "chain_ret", "chain_annual", "chain_n", "origin_sens%", "N_loss", "Worst",
    "avg_partial_gain", "avg_partial_alpha", "n_open_hops",
    "stop_loss", "n_stopped", "stop_net_per_position",
]


def _write_step3_sheet(wb: Workbook, backtests: dict[str, "bt.BacktestResult"],
                       run_paths: dict[str, Path] | None = None) -> None:
    """Row 1 is the header (one column per strategy); row 2 is ALWAYS the run_report
    filename (SM, 2026-08-11 — too much clicking to find the runX file for a strategy of
    interest), pinned directly under its own column's label so the pairing survives a
    manual column reorder in Excel — a row keyed by position would not. Metric rows start
    at row 3."""
    run_paths = run_paths or {}
    ws = wb.create_sheet("Step3_compare")
    ws.cell(1, 1, "metric").font = _BOLD
    order = sorted(backtests, key=lambda lbl: (
        -(backtests[lbl].metrics["chain_annual"]
          if pd.notna(backtests[lbl].metrics["chain_annual"]) else float("-inf"))
    ))
    for c, label in enumerate(order, start=2):
        cell = ws.cell(1, c, label); cell.font, cell.fill = _BOLD, _HEAD
        if backtests[label].metrics.get("thin"):
            cell.fill = _THIN_FILL
    ws.cell(2, 1, "run_report").font = _BOLD
    for c, label in enumerate(order, start=2):
        path = run_paths.get(label)
        ws.cell(2, c, path.name if path else None).font = _NOTE
    for r, key in enumerate(_STEP3_ROWS, start=3):
        ws.cell(r, 1, key).font = _BOLD
        for c, label in enumerate(order, start=2):
            result = backtests[label]
            if key in result.params:
                val = spec.format_value(key, result.params[key])
            else:
                val = result.metrics.get(key)
            if isinstance(val, float) and pd.notna(val):
                val = round(val, 4)
            elif isinstance(val, tuple):
                val = str(val)
            ws.cell(r, c, val)
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(order) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 16


# ---------------------------------------------------------------------------
# Step3a_stopout — cost/benefit sweep across stop levels, one block per active row
# ---------------------------------------------------------------------------

_STOPOUT_COLS: list[str] = [
    "stop", "n_positions", "n_stopped", "benefit", "cost", "net", "net_per_position",
    "chain_annual", "chain_n", "Worst", "N_loss", "avg_gain", "median_gain", "hit_rate%",
]
_STOPOUT_FOLD_COLS: list[str] = [
    "fold", "test_dates", "stop", "avg_gain", "median_gain", "hit_rate%",
    "chain_annual", "chain_n", "Worst", "N_loss",
]


def _compute_stopout(backtests: dict[str, "bt.BacktestResult"], settings: dict,
                     walk_results: list["wf.GroupResult"]) -> dict[str, dict]:
    """Per D-row Step3a sweep, computed ONCE and shared by the sheet and the Charts
    section below — step3a_stopout.levels_hops() applies each stop level to a row's
    hops_raw exactly once, reused by both the full-span sweep table and the per-fold
    stability table, and never re-simulates (it works off the already-built hops).

    A row is only eligible when its period is one longi_future_minaggr*.csv covers
    (20/50 — see shared.config.MINAGGR_PERIODS) AND there is something to show it at:
    the board's Settings.stop_sweep ladder, the row's own stop_loss, or both.

    The fold-stability table reuses each row's OWN GroupResult.folds from Step 4 (SM,
    2026-08-05: "the only thing on step 4 I really understand is the folds ... that
    could be done on stop-corrected lots as well as on intact lots") — same boundaries
    already on Step4_walkforward, no new fold geometry, no training-window selection."""
    tolerance = float(settings.get("stop_annual_tolerance", 5.0))
    sweep_levels = stopout.parse_sweep_levels(settings.get("stop_sweep", ""))

    folds_by_label: dict[str, list[dict]] = {}
    for result in walk_results:
        for owner in result.owners:
            folds_by_label[owner] = result.folds

    out: dict[str, dict] = {}
    for label, result in backtests.items():
        period = int(result.params["period"])
        if period not in MINAGGR_PERIODS:
            out[label] = {"period": period, "eligible": False, "rows": [], "flagged": None,
                         "fold_rows": []}
            continue
        own_stop = result.params.get("stop_loss")
        levels = sorted(set(sweep_levels) | ({own_stop} if own_stop else set()))
        if not levels:
            out[label] = {"period": period, "eligible": False, "rows": [], "flagged": None,
                         "fold_rows": []}
            continue

        by_level = stopout.levels_hops(result.hops_raw, period, levels)
        rows_ = stopout.metrics_rows(by_level, result.params, settings)
        folds = folds_by_label.get(label, [])
        fold_rows = stopout.fold_metrics(by_level, result.params, settings, folds) if folds else []

        out[label] = {"period": period, "eligible": True, "rows": rows_,
                      "flagged": stopout.best_level(rows_, tolerance), "fold_rows": fold_rows}
    return out


def _write_step3a_sheet(wb: Workbook, stopout_data: dict[str, dict], settings: dict) -> None:
    """The sweep table (step3a_stopout.sweep) plus the risk-first flagged level
    (step3a_stopout.best_level, highlighted) per eligible row. A row outside {20, 50} or
    with nothing to sweep gets a one-line note instead of an empty block, matching the
    "no group elevated" / "no pick" convention on Step1_groups/Step2_picks — an absent
    block must not read as a missing result."""
    ws = wb.create_sheet("Step3a_stopout")
    tolerance = float(settings.get("stop_annual_tolerance", 5.0))
    r = 1
    for label, info in stopout_data.items():
        cell = ws.cell(r, 1, label); cell.font = Font(bold=True, size=12)
        r += 1
        if not info["eligible"]:
            note = (f"(no longi_future_minaggr*.csv for period={info['period']} — "
                    f"only {list(MINAGGR_PERIODS)} covered)"
                    if info["period"] not in MINAGGR_PERIODS else
                    "(no stop_loss set and Settings.stop_sweep is blank — nothing to sweep)")
            ws.cell(r, 1, note).font = _NOTE
            r += 2
            continue

        for c, h in enumerate(_STOPOUT_COLS, start=1):
            cell = ws.cell(r, c, h); cell.font, cell.fill = _BOLD, _HEAD
        r += 1
        for row_ in info["rows"]:
            is_flagged = info["flagged"] is not None and row_["stop"] == info["flagged"]
            for c, key in enumerate(_STOPOUT_COLS, start=1):
                val = row_.get(key)
                if key == "stop":
                    val = "off" if val is None else val
                elif isinstance(val, float) and pd.notna(val):
                    val = round(val, 4)
                cell = ws.cell(r, c, val)
                if is_flagged:
                    cell.fill = _OK_FILL
            r += 1
        flagged_txt = (info["flagged"] if info["flagged"] is not None
                      else f"(none within {tolerance:.0f}% chain_annual tolerance)")
        ws.cell(r, 1, f"flagged (risk-first: best Worst/N_loss within "
                      f"{tolerance:.0f}% chain_annual giveback): {flagged_txt}").font = _NOTE
        r += 2

        # ---- fold stability: the SAME folds already on Step4_walkforward, every stop
        # level scored (not selected) on each one -- no training window, just "how does
        # this level's own out-of-sample number move across time slots" (SM, 2026-08-05).
        ws.cell(r, 1, "Fold stability — Step 4's own folds, stop-corrected vs intact "
                      "(no selection, every level scored on every fold)").font = _NOTE
        r += 1
        if not info["fold_rows"]:
            ws.cell(r, 1, "(no Step 4 folds for this row — walk-forward produced none, "
                          "e.g. too little history for even one fold)").font = _NOTE
            r += 2
        else:
            for c, h in enumerate(_STOPOUT_FOLD_COLS, start=1):
                cell = ws.cell(r, c, h); cell.font, cell.fill = _BOLD, _HEAD
            r += 1
            for fr in info["fold_rows"]:
                for c, key in enumerate(_STOPOUT_FOLD_COLS, start=1):
                    val = fr.get(key)
                    if key == "stop":
                        val = "off" if val is None else val
                    elif isinstance(val, float) and pd.notna(val):
                        val = round(val, 4)
                    ws.cell(r, c, val)
                r += 1
            r += 2

    ws.freeze_panes = "A1"
    ws.column_dimensions["A"].width = 18
    for c in range(2, len(_STOPOUT_COLS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14


# ---------------------------------------------------------------------------
# Step4_walkforward — one summary block + fold table per active row
# ---------------------------------------------------------------------------

_FOLD_COLS: list[str] = [
    "fold", "test_dates", "selected", "is_n", "oos_n", "is_annual", "oos_annual",
    "oos_avg_gain", "oos_median_gain", "oos_hit_rate%", "zeroskill_avg_gain",
    "oos_oracle", "oos_alpha",
]
_SUMMARY_COLS: list[str] = [
    "candidates", "folds", "oos_lots", "is_avg_gain", "oos_avg_gain", "gain_gap",
    "oos_median_gain", "oos_hit_rate%", "is_alpha", "oos_alpha", "alpha_gap",
    "zeroskill_avg_gain", "selection_skill_gain", "is_annual", "oos_annual",
]
_CANDIDATE_COLS: list[str] = [
    "candidate", "folds_selected", "oos_lots", "is_avg_gain", "oos_avg_gain", "gain_gap",
    "oos_median_gain", "oos_hit_rate%", "is_alpha", "oos_alpha", "alpha_gap",
    "is_annual", "oos_annual",
]


def _write_step4_sheet(wb: Workbook, walk_results: list["wf.GroupResult"]) -> None:
    """One block per distinct candidate set: a Summary row (the SELECTED candidate, i.e.
    does selection travel out of sample), a per-candidate table (each strategy's own OOS
    numbers — the per-strategy comparison), then the fold table."""
    ws = wb.create_sheet("Step4_walkforward")
    r = 1
    for result in walk_results:
        summary = wf.summarize(result)
        cell = ws.cell(r, 1, f"{' + '.join(result.owners)}  "
                             f"(vs {len(result.candidate_labels)} candidate(s), "
                             f"period={result.period}d)")
        cell.font = Font(bold=True, size=12)
        r += 1
        ws.cell(r, 1, "Summary — the candidate SELECTED each fold, not any one row's own "
                      "result; see the per-candidate table below").font = _NOTE
        r += 1
        for c, key in enumerate(_SUMMARY_COLS, start=1):
            ws.cell(r, c, key).font = _BOLD
        r += 1
        for c, key in enumerate(_SUMMARY_COLS, start=1):
            val = summary.get(key)
            ws.cell(r, c, round(val, 4) if isinstance(val, float) and pd.notna(val) else val)
        r += 2

        ws.cell(r, 1, "Per candidate — each strategy's own numbers over the same folds, "
                      "best oos_avg_gain first").font = _NOTE
        r += 1
        for c, h in enumerate(_CANDIDATE_COLS, start=1):
            cell = ws.cell(r, c, h); cell.font, cell.fill = _BOLD, _HEAD
        r += 1
        for cand in wf.summarize_candidates(result):
            for c, key in enumerate(_CANDIDATE_COLS, start=1):
                val = cand.get(key)
                ws.cell(r, c, round(val, 4) if isinstance(val, float) and pd.notna(val) else val)
            r += 1
        r += 2

        for c, h in enumerate(_FOLD_COLS, start=1):
            cell = ws.cell(r, c, h); cell.font, cell.fill = _BOLD, _HEAD
        r += 1
        for f in result.folds:
            for c, key in enumerate(_FOLD_COLS, start=1):
                val = f.get(key)
                ws.cell(r, c, round(val, 4) if isinstance(val, float) and pd.notna(val) else val)
            r += 1
        r += 2
    for c in range(1, max(len(_FOLD_COLS), len(_CANDIDATE_COLS), len(_SUMMARY_COLS)) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions["A"].width = 24


# ---------------------------------------------------------------------------
# Charts — cumulative chain / IS-vs-OOS / avg-vs-median, per active row
# ---------------------------------------------------------------------------

def _greedy_chain_lots(hops: list["bt.Hop"], period: int, threshold) -> list[tuple[int, float]]:
    """Single-origin non-overlapping chain lots (daynum, gain), ascending — a plain equity
    curve for the chart. (Step3_compare's chain_annual is phase-averaged for stability;
    this single origin is only for a visual cumulative-gain shape.)"""
    usable = sorted(
        ((h.daynum, g) for h in hops if h.realized and bt.gate_ok(h, threshold)
         for g in [market.hop_avg(h.gains)] if pd.notna(g)),
        key=lambda t: t[0],
    )
    lots: list[tuple[int, float]] = []
    next_allowed: int | None = None
    for dn, g in usable:
        if next_allowed is None or dn >= next_allowed:
            lots.append((dn, g))
            next_allowed = dn + period
    return lots


def _write_charts_sheet(wb: Workbook, backtests: dict[str, "bt.BacktestResult"],
                       walk_results: list["wf.GroupResult"],
                       stopout_data: dict[str, dict]) -> None:
    ws = wb.create_sheet("Charts")
    labels = list(backtests)

    # --- Table 1: cumulative chain, by lot index (not calendar daynum -- see docstring) ---
    cum_by_label: dict[str, list[float]] = {}
    for label, result in backtests.items():
        threshold = result.params.get("no_go_gspc_rsi")
        period = int(result.params["period"])
        lots = _greedy_chain_lots(result.hops, period, threshold)
        running = 0.0
        cum: list[float] = []
        for _dn, g in lots:
            running += g
            cum.append(running)
        cum_by_label[label] = cum

    max_lots = max((len(v) for v in cum_by_label.values()), default=0)
    row0 = 1
    ws.cell(row0, 1, "Cumulative chain (by lot index)").font = Font(bold=True, size=12)
    ws.cell(row0 + 1, 1, "lot #")
    for c, label in enumerate(labels, start=2):
        ws.cell(row0 + 1, c, label).font = _BOLD
    for i in range(max_lots):
        r = row0 + 2 + i
        ws.cell(r, 1, i + 1)
        for c, label in enumerate(labels, start=2):
            vals = cum_by_label[label]
            if i < len(vals):
                ws.cell(r, c, round(vals[i], 4))
    if max_lots:
        chart1 = LineChart()
        chart1.title = "Cumulative chain gain (%) by lot index"
        chart1.y_axis.title, chart1.x_axis.title = "cumulative %", "lot #"
        data = Reference(ws, min_col=2, max_col=1 + len(labels),
                         min_row=row0 + 1, max_row=row0 + 1 + max_lots)
        cats = Reference(ws, min_col=1, min_row=row0 + 2, max_row=row0 + 1 + max_lots)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        ws.add_chart(chart1, f"A{row0 + max_lots + 4}")

    # --- Table 2: IS vs OOS avg_gain per CANDIDATE (each strategy's own walk-forward
    # numbers). Charting the per-set Summary instead would draw one identical bar pair per
    # owner whenever rows share a candidate set — see _write_step4_sheet. ---
    row1 = row0 + max_lots + 24
    ws.cell(row1, 1, "IS vs OOS avg_gain, per strategy").font = Font(bold=True, size=12)
    ws.cell(row1 + 1, 1, "label"); ws.cell(row1 + 1, 2, "is_avg_gain"); ws.cell(row1 + 1, 3, "oos_avg_gain")
    for c in (1, 2, 3):
        ws.cell(row1 + 1, c).font = _BOLD
    per_candidate: dict[str, dict] = {}
    for result in walk_results:
        for cand in wf.summarize_candidates(result):
            per_candidate.setdefault(cand["candidate"], cand)
    wf_labels = list(per_candidate)
    for i, label in enumerate(wf_labels):
        cand = per_candidate[label]
        r = row1 + 2 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(cand["is_avg_gain"], 4) if pd.notna(cand["is_avg_gain"]) else None)
        ws.cell(r, 3, round(cand["oos_avg_gain"], 4) if pd.notna(cand["oos_avg_gain"]) else None)
    if wf_labels:
        chart2 = BarChart()
        chart2.type, chart2.title = "col", "In-sample vs out-of-sample avg_gain (%)"
        data = Reference(ws, min_col=2, max_col=3, min_row=row1 + 1, max_row=row1 + 1 + len(wf_labels))
        cats = Reference(ws, min_col=1, min_row=row1 + 2, max_row=row1 + 1 + len(wf_labels))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, f"E{row1}")

    # --- Table 3: avg_gain vs median_gain per row -- the shape behind the mean ---
    row2 = row1 + max(len(wf_labels), 6) + 20
    ws.cell(row2, 1, "avg_gain vs median_gain, per row").font = Font(bold=True, size=12)
    ws.cell(row2 + 1, 1, "label"); ws.cell(row2 + 1, 2, "avg_gain"); ws.cell(row2 + 1, 3, "median_gain")
    for c in (1, 2, 3):
        ws.cell(row2 + 1, c).font = _BOLD
    for i, label in enumerate(labels):
        m = backtests[label].metrics
        r = row2 + 2 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(m["avg_gain"], 4) if pd.notna(m["avg_gain"]) else None)
        ws.cell(r, 3, round(m["median_gain"], 4) if pd.notna(m["median_gain"]) else None)
    if labels:
        chart3 = BarChart()
        chart3.type, chart3.title = "col", "avg_gain vs median_gain (%) -- the shape behind the mean"
        data = Reference(ws, min_col=2, max_col=3, min_row=row2 + 1, max_row=row2 + 1 + len(labels))
        cats = Reference(ws, min_col=1, min_row=row2 + 2, max_row=row2 + 1 + len(labels))
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        ws.add_chart(chart3, f"E{row2}")

    # --- Table 4: chain_annual & Worst across stop levels, per eligible row (Step 3a) ---
    row3 = row2 + max(len(labels), 6) + 20
    eligible = {lbl: info for lbl, info in stopout_data.items() if info["eligible"]}
    ws.cell(row3, 1, "Step 3a: chain_annual & Worst across stop levels").font = Font(bold=True, size=12)
    r = row3 + 1
    for label, info in eligible.items():
        ws.cell(r, 1, label).font = _BOLD
        r += 1
        cats_row = r
        ws.cell(r, 1, "stop")
        for c, row_ in enumerate(info["rows"], start=2):
            ws.cell(r, c, "off" if row_["stop"] is None else row_["stop"])
        r += 1
        data_top = r
        for key in ("chain_annual", "Worst"):
            ws.cell(r, 1, key).font = _BOLD
            for c, row_ in enumerate(info["rows"], start=2):
                val = row_.get(key)
                ws.cell(r, c, round(val, 4) if isinstance(val, float) and pd.notna(val) else None)
            r += 1
        n_levels = len(info["rows"])
        if n_levels:
            chart4 = LineChart()
            chart4.title = f"{label}: chain_annual & Worst vs stop level"
            data = Reference(ws, min_col=1, max_col=1 + n_levels, min_row=data_top, max_row=r - 1)
            cats = Reference(ws, min_col=2, max_col=1 + n_levels, min_row=cats_row, max_row=cats_row)
            chart4.add_data(data, titles_from_data=True, from_rows=True)
            chart4.set_categories(cats)
            ws.add_chart(chart4, f"{get_column_letter(3 + n_levels)}{data_top - 1}")
        r += 2

    ws.column_dimensions["A"].width = 24


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------

def _archive_prior() -> None:
    for path in REPORT_ROOT.glob("compare_strategies_*.xlsx"):
        dest = REPORT_ROOT / "_archive" / time.strftime("%Y%m%d_%H%M%S")
        dest.mkdir(parents=True, exist_ok=True)
        shutil.move(str(path), str(dest / path.name))


def archive_prior_strategic_stocks() -> None:
    """Moves every existing StrategicStocks_*.xlsx/.csv (and the pre-2026-08-12 unsuffixed
    StrategicStocks.xlsx, same glob) to one dated _archive/ folder. Wrapped per-file in
    try/except: a file Excel currently has open can't be moved over SMB, and that should
    warn, not crash a tick that has otherwise-good output ready to write."""
    paths = list(REPORT_ROOT.glob("StrategicStocks*.xlsx")) + list(REPORT_ROOT.glob("StrategicStocks*.csv"))
    if not paths:
        return
    dest = REPORT_ROOT / "_archive" / time.strftime("%Y%m%d_%H%M%S")
    dest.mkdir(parents=True, exist_ok=True)
    for path in paths:
        try:
            shutil.move(str(path), str(dest / path.name))
        except OSError as exc:
            print(f"[outputboard] WARNING: could not archive {path.name} ({exc}) -- "
                  f"probably open in Excel; left in place", flush=True)


def _picks_daynum(picks: dict[str, dict]) -> "int | str":
    """The daynum StrategicStocks_<daynum> is named for -- every successful pick's own
    daynum, which is the same across active rows in practice (one frozen snapshot per
    tick). Falls back to 'unknown' only when every row failed (archive_prior_strategic_
    stocks/write_strategic_stocks still run so nothing crashes, just an unhelpful name)."""
    daynums = {info["daynum"] for info in picks.values() if not info.get("error")}
    if not daynums:
        return "unknown"
    if len(daynums) > 1:
        print(f"[outputboard] NOTE: active rows disagree on daynum ({sorted(daynums)}), "
              f"filename uses the latest", flush=True)
    return max(daynums)


def _publish_strategic_csv() -> None:
    """rclone-syncs StrategicStocks_<daynum>.csv to GoogleDrive:PotSystem/repositoryRTBI/
    Strategy immediately after it's written (SM, 2026-08-12: "asap after creation"), via
    the shared repositoryRTBI publish contract (shared/app/code/repository.py) -- the same
    ownership-scoped rclone sync every other producer (longi, group_conformity) uses,
    rather than a one-off upload path. `sync` (not `copy`) also means an OLDER dated CSV
    that archive_prior_strategic_stocks() already moved out of app/report/ is cleaned up
    on Drive too, on the next publish -- Drive always mirrors the current file, full
    history stays recoverable locally under _archive/.

    Never fails the tick: the local .xlsx/.csv are the primary deliverable, a network
    hiccup or a temporarily-unreachable Drive is a warning here, not a crash."""
    import sys as _sys
    shared_code = Path(__file__).resolve().parents[3] / "shared" / "app" / "code"
    if str(shared_code) not in _sys.path:
        _sys.path.insert(0, str(shared_code))
    try:
        import repository
        rc = repository.publish(repository.OWNERS["strategy_grp2"])
        if rc != 0:
            print(f"[outputboard] WARNING: Drive publish exited {rc} -- local files are "
                  f"fine, the Drive copy may be stale until the next successful publish",
                  flush=True)
    except Exception as exc:                                    # noqa: BLE001
        print(f"[outputboard] WARNING: Drive publish failed ({exc}) -- local files are "
              f"fine, the Drive copy may be stale until the next successful publish",
              flush=True)


def write_strategic_stocks(picks: dict[str, dict]) -> tuple[Path, Path]:
    """StrategicStocks_<daynum>.xlsx + StrategicStocks_<daynum>.csv (2026-08-12, SM: "a csv
    file exported in addition to StrategicStocks.xlsx ... containing what is in the All
    tab ... European csv format", "the xlsx shall add _<daynum> to its filename"). The
    daynum suffix means a same-day rerun overwrites in place while a new day's run never
    collides with a prior day's file -- see archive_prior_strategic_stocks(), called by
    both entry points before this, for what happens to the previous one.

    xlsx: one tab per active row, same fields as Step2_picks ("export all fields from
    Step2_picks"), computed by the exact same step2_table() so the two never drift apart.
    A first tab, 'All', holds every row from every tab concatenated, tinted per strategy
    like Step2_picks (SM: "make a tab called All ... place this as the first tab").

    csv: exactly the All tab's rows -- one flat table, same headers, European format
    (`sep=';', decimal=','`, this project's hard rule) -- then published to Drive, see
    _publish_strategic_csv() above.

    Written ONLY by conductor.cmd_production (2026-08-13 correction -- previously also
    written by the bare development tick's assemble(), which meant a wild development trial
    could reach real users via the Drive publish below; see DesignVersion2.md's 2026-08-13
    correction). Every active row still ships, unconditionally, within that one path -- no
    purpose column to filter on."""
    headers, rows_by_label = step2_table(picks)
    daynum = _picks_daynum(picks)
    all_rows: list[list] = []

    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("All")
    for c, h in enumerate(headers, start=1):
        cell = ws_all.cell(1, c, h); cell.font, cell.fill = _BOLD, _HEAD
    r = 2
    blocks: list[tuple[int, int]] = []
    for label in picks:
        block_top = r
        for row_vals in rows_by_label.get(label, []):
            for c, val in enumerate(row_vals, start=1):
                ws_all.cell(r, c, val)
            r += 1
            all_rows.append(row_vals)
        if r > block_top:
            blocks.append((block_top, r - 1))
    ws_all.freeze_panes = "A2"
    display.band(ws_all, blocks, len(headers))
    size_step2_columns(ws_all, headers)

    for label in picks:
        ws = wb.create_sheet(label[:31])   # Excel's 31-char sheet-name limit
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(1, c, h); cell.font, cell.fill = _BOLD, _HEAD
        for r2, row_vals in enumerate(rows_by_label.get(label, []), start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(r2, c, val)
        ws.freeze_panes = "A2"
        size_step2_columns(ws, headers)

    display.harmonize_workbook(wb)
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    xlsx_path = REPORT_ROOT / f"StrategicStocks_{daynum}.xlsx"
    wb.save(xlsx_path)

    csv_path = REPORT_ROOT / f"StrategicStocks_{daynum}.csv"
    pd.DataFrame(all_rows, columns=headers).to_csv(csv_path, sep=";", decimal=",", index=False)
    print(f"    wrote {csv_path}", flush=True)

    _publish_strategic_csv()

    return xlsx_path, csv_path


def _picks_long_rows(label: str, history: dict[int, list[str]]) -> list[tuple[str, int, str, int]]:
    """(label, daynum, ticker, priority) — one row per pick, priority is production_pick's
    own 1-based order (same number Step2_picks' `priority` column shows). daynum descending
    (newest first, Longi's own convention), priority ascending within a daynum."""
    rows = [
        (label, daynum, ticker, i)
        for daynum, tickers in history.items()
        for i, ticker in enumerate(tickers, start=1)
    ]
    rows.sort(key=lambda row: (-row[1], row[3]))
    return rows


def write_picks_csv(active_rows: list["cb.RunRow"]) -> Path | None:
    """picks_<daynum>.csv — long format, one row per (label, daynum, ticker) pick across
    every active `D` row's ENTIRE dominance history (~650 daynums), not just today (SM,
    2026-08-13: "a similar list must be pulled on any other of the time line"). Originally
    a wide ticker x daynum matrix per row (one Excel tab each); switched to long format
    2026-08-14 (SM: "cancel Excel output and deliver instead a long format csv containing
    label-daynum-ticker-priority. I will then import this in as needed Excel and pivot the
    long format there") — one flat table across all rows instead of one tab per row. No
    attribute columns, by request — purely the label-daynum-ticker-priority list.

    Development-tick only. Never written by --production and never Drive-published — this
    is a research/backtesting artifact, not the day's advice list. European format
    (`sep=';', decimal=','`), this project's hard rule. None if every row failed step 0/1
    (nothing to write)."""
    all_rows: list[tuple[str, int, str, int]] = []
    daynums_seen: set[int] = set()
    for row in active_rows:
        label = row.resolved.get("label")
        print(f"[picks] {label}: pulling full history ...", flush=True)
        try:
            _s0, history = step2_focusset.pick_history(row.resolved)
        except (expr.ExpressionError, ValueError) as exc:
            print(f"    FAILED: {exc}", flush=True)
            continue
        all_rows.extend(_picks_long_rows(label, history))
        daynums_seen.update(history)
        n_tickers = len({t for tickers in history.values() for t in tickers})
        print(f"    {n_tickers} distinct ticker(s) across {len(history)} daynum(s)", flush=True)

    if not all_rows:
        return None

    for path in list(REPORT_ROOT.glob("picks_*.xlsx")) + list(REPORT_ROOT.glob("picks_*.csv")):
        dest = REPORT_ROOT / "_archive" / time.strftime("%Y%m%d_%H%M%S")
        dest.mkdir(parents=True, exist_ok=True)
        shutil.move(str(path), str(dest / path.name))

    daynum = max(daynums_seen) if daynums_seen else "unknown"
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    path = REPORT_ROOT / f"picks_{daynum}.csv"
    df = pd.DataFrame(all_rows, columns=["label", "daynum", "ticker", "priority"])
    df.to_csv(path, sep=";", decimal=",", index=False)
    return path


def assemble(board: "cb.BoardResult", settings: dict) -> Path:
    """Build the development-tick workbook (compare_strategies_<date>.xlsx) only.

    2026-08-13 correction (SM): a bare tick no longer writes or publishes StrategicStocks --
    that file is what real users read as the day's advice, and a development tick is where
    wild trial rows live. StrategicStocks.xlsx/.csv is now written by ONE path only,
    conductor.cmd_production, fired by an explicit `--production`. See DesignVersion2.md's
    2026-08-13 correction for the incident that prompted this (superseding the 2026-08-12
    "every active row ships in one invocation" decision).

    Same-day follow-up (SM): row selection itself is now `D`-only too, independent of `P` —
    a row marked `D` for development work never appears in a `--production` run (cron-fired
    or otherwise) unless `P` is marked on it as well. See the 2026-08-13 correction's second
    half in DesignVersion2.md."""
    active_rows = [r for r in board.runs if r.d_active and r.ok]
    rejected_rows = [r for r in board.runs if r.d_active and not r.ok]
    if rejected_rows or board.board_errors:
        print(f"\n!!! {len(rejected_rows)} D row(s) REJECTED by board validation — "
              f"they are on the Runs sheet with their error, but produce no results:", flush=True)
        for row in rejected_rows:
            print(f"    row {row.row_num} '{row.resolved.get('label')}': "
                  f"{'; '.join(row.errors)}", flush=True)
        for msg in board.board_errors:
            print(f"    BOARD: {msg}", flush=True)
    print(f"\n=== Steps 0-2: {len(active_rows)} active row(s) ===", flush=True)
    picks = current_picks(active_rows)

    print(f"\n=== Step 2 (dev-only): picks_<daynum>.csv, full history ===", flush=True)
    picks_path = write_picks_csv(active_rows)
    if picks_path:
        print(f"    wrote {picks_path}", flush=True)

    active_by_label = {r.resolved["label"]: r.resolved for r in active_rows}

    print(f"\n=== Step 3: backtest ({len(active_by_label)} row(s)) ===", flush=True)
    backtests: dict[str, "bt.BacktestResult"] = {}
    for i, (label, row_resolved) in enumerate(active_by_label.items(), start=1):
        print(f"[{i}/{len(active_by_label)}] {label}: building hops ...", flush=True)
        backtests[label] = bt.run_backtest(row_resolved, settings, progress_label=label)
        m = backtests[label].metrics
        print(f"    chain_annual={m['chain_annual']:.2f}  chain_n={m['chain_n']}  "
              f"N_hops={m['N_hops']}", flush=True)

    run_paths: dict[str, Path] = {}
    if backtests:
        run_paths = step3_report.write_run_reports(backtests)
        for p in run_paths.values():
            print(f"    wrote {p}", flush=True)

    # A walk-forward test belongs to its CANDIDATE SET, not to the row that declared it:
    # two rows naming the same candidates run the identical experiment and would otherwise
    # print the identical fold table twice under two headings. Group by the resolved set,
    # run each distinct one once, and list every owner on the block.
    print(f"\n=== Step 4: walk-forward ({len(active_by_label)} row(s)) ===", flush=True)
    by_candidate_set: dict[frozenset[str], list[str]] = {}
    candidates_for: dict[frozenset[str], dict[str, dict]] = {}
    for label, row_resolved in active_by_label.items():
        candidates = wf.resolve_candidates(label, row_resolved, active_by_label)
        key = frozenset(candidates)
        by_candidate_set.setdefault(key, []).append(label)
        candidates_for[key] = candidates
    if len(by_candidate_set) < len(active_by_label):
        print(f"    {len(active_by_label)} row(s) -> {len(by_candidate_set)} distinct "
              f"candidate set(s); rows sharing a set share one test", flush=True)

    hops_cache = {label: (r.hops, r.params) for label, r in backtests.items()}
    walk_results: list["wf.GroupResult"] = []
    for i, (key, owners) in enumerate(by_candidate_set.items(), start=1):
        candidates = candidates_for[key]
        print(f"[{i}/{len(by_candidate_set)}] {', '.join(owners)}: walk-forward vs "
              f"{len(candidates)} candidate(s) ...", flush=True)
        try:
            walk_results.append(wf.walk_group(
                owners, candidates, settings,
                min_train=int(settings.get("wf_min_train", wf.DEFAULT_MIN_TRAIN)),
                test_len=int(settings.get("wf_test_len", wf.DEFAULT_TEST_LEN)),
                hops_cache=hops_cache))
        except ValueError as exc:
            print(f"[outputboard] walk-forward skipped for '{owners[0]}': {exc}")

    stopout_data: dict[str, dict] = {}
    if backtests:
        print(f"\n=== Step 3a: stop-out sweep ===", flush=True)
        stopout_data = _compute_stopout(backtests, settings, walk_results)
        for label, info in stopout_data.items():
            if info["eligible"]:
                n_folds = len(set(fr["fold"] for fr in info["fold_rows"]))
                print(f"    {label}: {len(info['rows']) - 1} level(s), flagged={info['flagged']}, "
                      f"{n_folds} fold(s) scored", flush=True)

    print("\n=== Writing workbook ===", flush=True)
    wb = Workbook()
    wb.remove(wb.active)
    _write_runs_sheet(wb, active_rows, picks, rejected_rows, board.board_errors)
    _write_step1_sheet(wb, picks)
    _write_step2_sheet(wb, picks)
    if backtests:
        _write_step3_sheet(wb, backtests, run_paths)
        _write_step3a_sheet(wb, stopout_data, settings)
    if walk_results:
        _write_step4_sheet(wb, walk_results)
    if backtests:
        _write_charts_sheet(wb, backtests, walk_results, stopout_data)

    # Fixed decimals last, once, over every sheet — a display convention, not a per-writer
    # decision (shared/display.py). Step3_compare is transposed, so its runs go across.
    display.harmonize_workbook(wb, row_oriented={"Step3_compare"})

    _archive_prior()
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    path = REPORT_ROOT / f"compare_strategies_{date.today():%Y%m%d}.xlsx"
    wb.save(path)

    return path
