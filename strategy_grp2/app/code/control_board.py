"""
The control board — read, validate, and (re)generate `app/control/control_board.xlsx`.

The processor never writes to this file while a tick is running (a file open in Excel
can't be written anyway, and input/output separation is the point — see
DesignVersion2.md). `--make-board` is the one path that touches it, and only to add
columns/legend rows the schema has grown since the file was last written; existing rows'
values are preserved.

Three sheets:
    Runs     one row per run; `D`/`P` are the two independent ticks (2026-08-13: replaced
             a single `active` column — see param_spec.PARAMS for columns, and
             DesignVersion2.md's 2026-08-13 correction for why)
    Settings global knobs, not per-run (see param_spec.SETTINGS)
    Legend   generated from param_spec — never hand-edited
"""

from __future__ import annotations

import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))

import param_spec as spec
from shared.config import CONTROL_BOARD_PATH, CONTROL_ROOT, MINAGGR_PERIODS

_HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
_META_FILL = PatternFill("solid", fgColor="D6DCE4")
_STEP_FILLS = {
    0: PatternFill("solid", fgColor="FFE599"),
    1: PatternFill("solid", fgColor="C6EFCE"),
    2: PatternFill("solid", fgColor="FFF2CC"),
    3: PatternFill("solid", fgColor="F8CBAD"),
    4: PatternFill("solid", fgColor="D9D2E9"),
}
_ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")


# ---------------------------------------------------------------------------
# "Did you save the board?" guard
# ---------------------------------------------------------------------------
#
# VBA can ask a workbook `ActiveWorkbook.Saved`; we cannot. That flag lives inside the
# Excel process on the Windows host, and this code runs on the Ubuntu server over SSH —
# there is no COM to ask, and the unsaved edits themselves exist only in Excel's memory,
# so nothing on disk can reveal them. What IS on disk is Excel's owner file, `~$<name>`,
# written next to a workbook while it is open and deleted on close. It travels over the
# Samba share, so the server sees it.
#
# So this guard is deliberately one notch stricter than VBA's: it stops when the board is
# OPEN, not when it is provably dirty. An open-but-saved board is stopped too. That is the
# right trade for the failure it exists to prevent — a tick silently running the previous
# board because an edit was still sitting in Excel — since the stop costs one keystroke and
# the silent run costs a whole misread comparison. `--board-open-ok` overrides it.

_LOCK_PREFIX = "~$"


class BoardOpenError(RuntimeError):
    """The control board is open in Excel, so a tick would risk reading unsaved edits."""


@dataclass(frozen=True)
class BoardFileState:
    path: Path
    lock_path: Path
    is_open: bool
    owner: str | None
    mtime: float | None

    @property
    def saved_ago(self) -> str:
        if self.mtime is None:
            return "never (file missing)"
        age = max(0, int(time.time() - self.mtime))
        stamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(self.mtime))
        if age < 90:
            return f"{stamp} ({age}s ago)"
        if age < 5400:
            return f"{stamp} ({age // 60} min ago)"
        return f"{stamp} ({age // 3600}h {age % 3600 // 60}m ago)"


def _lock_owner(path: Path) -> str | None:
    """The user named inside an Excel owner file: a length byte, then the name in ANSI.
    Best-effort — a corrupt or empty lock file is still a lock file."""
    try:
        raw = path.read_bytes()
        if not raw:
            return None
        name = raw[1:1 + raw[0]].decode("latin-1", "replace").strip()
        return name or None
    except OSError:
        return None


def board_file_state(path: Path = CONTROL_BOARD_PATH) -> BoardFileState:
    lock_path = path.with_name(_LOCK_PREFIX + path.name)
    is_open = lock_path.exists()
    return BoardFileState(
        path=path,
        lock_path=lock_path,
        is_open=is_open,
        owner=_lock_owner(lock_path) if is_open else None,
        mtime=path.stat().st_mtime if path.exists() else None,
    )


def require_saved_board(action: str = "run", *, allow_open: bool = False,
                       path: Path = CONTROL_BOARD_PATH) -> BoardFileState:
    """Raise BoardOpenError if the board is open in Excel. `action` names what was about to
    happen, so the message says what did not run."""
    state = board_file_state(path)
    if not state.is_open or allow_open:
        return state
    who = f" by '{state.owner}'" if state.owner else ""
    raise BoardOpenError(
        f"control_board.xlsx is OPEN in Excel{who} — {action} stopped.\n"
        f"    Board on disk was last saved: {state.saved_ago}\n"
        f"    An edit you have not saved is invisible from here, so this would have run\n"
        f"    the PREVIOUS board without saying so. Save (Ctrl+S) and close it, then retry.\n"
        f"    If Excel is NOT open, it crashed and left a stray lock file — delete it:\n"
        f"        rm '{state.lock_path}'\n"
        f"    To proceed anyway, using the board as it currently sits on disk:\n"
        f"        add --board-open-ok"
    )


# ---------------------------------------------------------------------------
# Value coercion — one function per column, driven by param_spec.ParamDef
# ---------------------------------------------------------------------------

_TRUTHY = {"x", "X", "yes", "y", "true", "1"}


def _coerce_bool(raw) -> bool:
    if isinstance(raw, bool):
        return raw
    return str(raw).strip() in _TRUTHY


def _is_blank(raw) -> bool:
    return raw is None or (isinstance(raw, str) and raw.strip() == "")


def coerce(pdef: "spec.ParamDef", raw, row_num: int):
    """Blank cell -> schema default (error if required). Otherwise dtype/parser-driven
    coercion, raising ValueError with the row number so a bad cell is easy to find."""
    blank = _is_blank(raw)
    if blank:
        if pdef.required:
            raise ValueError(f"row {row_num}: '{pdef.name}' is required and blank — {pdef.help}")
        return pdef.default

    try:
        if pdef.parser is not None:
            value = pdef.parser(raw)
        elif pdef.dtype == "bool":
            value = _coerce_bool(raw)
        elif pdef.dtype == "int":
            value = int(float(raw))
        elif pdef.dtype == "float":
            value = float(raw)
        elif pdef.dtype in ("str", "expr"):
            value = str(raw).strip()
        elif pdef.dtype == "choice":
            s = str(raw).strip()
            match = next((c for c in (pdef.choices or ()) if c.lower() == s.lower()), None)
            if match is None:
                raise ValueError(f"'{raw}' is not one of {list(pdef.choices or ())}")
            value = match
        elif pdef.dtype == "list_str":
            value = tuple(x.strip() for x in str(raw).split(",") if x.strip())
        else:
            raise ValueError(f"unhandled dtype '{pdef.dtype}'")
    except ValueError as exc:
        raise ValueError(f"row {row_num}: '{pdef.name}'={raw!r} invalid — {exc}") from None

    if pdef.dtype == "int" and pdef.choices and str(value) not in pdef.choices:
        raise ValueError(
            f"row {row_num}: '{pdef.name}'={value} invalid — must be one of {list(pdef.choices)}"
        )
    return value


# ---------------------------------------------------------------------------
# Runs sheet: read + resolve
# ---------------------------------------------------------------------------


@dataclass
class RunRow:
    row_num: int                       # 1-based Excel row (for error messages)
    raw: dict[str, object]             # exactly what the cells held
    resolved: dict[str, object] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    @property
    def d_active(self) -> bool:
        return bool(self.resolved.get("D", False))

    @property
    def p_active(self) -> bool:
        return bool(self.resolved.get("P", False))

    @property
    def active(self) -> bool:
        """In play at ALL, in either mode -- for validation-only concerns that don't care
        which entry point runs (label-dedupe, preflight's data snapshot, --dry-run/--check
        diagnostics). Never use this to decide which rows a TICK actually runs -- that's
        d_active for the bare tick, p_active for --production, and they are deliberately
        independent (2026-08-13 correction, DesignVersion2.md)."""
        return self.d_active or self.p_active


_GROUP_TOKEN_RE = re.compile(r"Stamdata\.(\w+)")


def _derive_label(resolved: dict) -> str:
    """A short, readable default when the row's `label` cell is blank. This is a display
    heuristic only — the real parse of group_expression lives in step0_data.py's
    expression engine (phase 2); here it just extracts column names for a readable name.

    Includes from_rank whenever it differs from the default best-n (1): two rows that
    differ ONLY in from_rank — e.g. the 1 / "mid" / -1 comparison the design calls for
    (see DesignVersion2.md) — must not collide on label, since the board-level duplicate
    check would otherwise reject exactly the rows meant to sit side by side."""
    expr = str(resolved.get("group_expression", "")).strip()
    if expr == "#ALL" or not expr:
        group_token = "ALL"
    else:
        tokens = _GROUP_TOKEN_RE.findall(expr)
        group_token = "_".join(dict.fromkeys(tokens)) if tokens else "expr"
    level = resolved.get("level", "A")
    prio = resolved.get("priority_attribute", "")
    label = f"{level}_{group_token}_{prio}"
    parsed = resolved.get("from_rank")
    if parsed and parsed != ("edge", 1):
        import param_spec as spec
        label += f"_{spec.format_from_rank(parsed)}"
    return label


def resolve_row(raw: dict[str, object], row_num: int) -> RunRow:
    row = RunRow(row_num=row_num, raw=dict(raw))
    for pdef in spec.PARAMS:
        try:
            row.resolved[pdef.name] = coerce(pdef, raw.get(pdef.name), row_num)
        except ValueError as exc:
            row.errors.append(str(exc))

    # stop_loss (Step 3a) leans on longi_future_minaggr<period>d.csv, which exists for
    # only two of the seven-pack horizons -- a row asking for a stop at, say, period=100
    # would silently have nothing to apply it to (see shared.config.minaggr_file).
    stop_loss = row.resolved.get("stop_loss")
    period = row.resolved.get("period")
    if stop_loss and period is not None and int(period) not in MINAGGR_PERIODS:
        row.errors.append(
            f"row {row_num}: stop_loss={stop_loss} requires period in {list(MINAGGR_PERIODS)} "
            f"(longi_future_minaggr*.csv only exists for those horizons), got period={period}"
        )

    if not row.resolved.get("label"):
        try:
            row.resolved["label"] = _derive_label(row.resolved)
        except Exception:                            # noqa: BLE001 — never let a label crash a row
            row.resolved["label"] = f"row{row_num}"
    return row


# ---------------------------------------------------------------------------
# Board-level read/validate
# ---------------------------------------------------------------------------


@dataclass
class BoardResult:
    runs: list[RunRow]
    settings: dict[str, object]
    board_errors: list[str]            # errors not tied to one row (e.g. duplicate labels)

    @property
    def active_runs(self) -> list[RunRow]:
        return [r for r in self.runs if r.active]

    @property
    def ok(self) -> bool:
        return not self.board_errors and all(r.ok for r in self.active_runs)


def _read_runs_sheet(ws: Worksheet) -> list[dict[str, object]]:
    headers = [c.value for c in ws[1]]
    rows: list[dict[str, object]] = []
    for excel_row in ws.iter_rows(min_row=2):
        values = [c.value for c in excel_row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append({h: v for h, v in zip(headers, values) if h})
    return rows


def _read_settings_sheet(ws: Worksheet) -> dict[str, object]:
    out: dict[str, object] = {}
    for excel_row in ws.iter_rows(min_row=2):
        name = excel_row[0].value
        if not name:
            continue
        out[str(name).strip()] = excel_row[1].value if len(excel_row) > 1 else None
    return out


def _resolve_settings(raw: dict[str, object]) -> tuple[dict[str, object], list[str]]:
    resolved: dict[str, object] = {}
    errors: list[str] = []
    for sdef in spec.SETTINGS:
        cell = raw.get(sdef.name)
        if cell is None or (isinstance(cell, str) and cell.strip() == ""):
            resolved[sdef.name] = sdef.default
            continue
        try:
            if sdef.dtype == "bool":
                resolved[sdef.name] = _coerce_bool(cell)
            elif sdef.dtype == "int":
                resolved[sdef.name] = int(float(cell))
            elif sdef.dtype == "float":
                resolved[sdef.name] = float(cell)
            else:
                resolved[sdef.name] = str(cell).strip()
        except ValueError as exc:
            errors.append(f"Settings '{sdef.name}'={cell!r} invalid — {exc}")
            resolved[sdef.name] = sdef.default
    return resolved, errors


def _dedupe_active_labels(runs: list["RunRow"]) -> list[str]:
    """Append ' (2)', ' (3)', ... to any active row's label that collides with an earlier
    active row's -- label is used as a dict key everywhere downstream (current_picks,
    Step2_picks, walk-forward candidate sets), so an unresolved collision silently drops
    one row's results instead of erroring (SM, 2026-08-12, after discovering exactly that:
    "non-unique labels can cause overwriting but it is easy to forget manually creating
    unique labels"). In-memory only -- `row.resolved["label"]` changes for THIS run's
    dict keys and report sheets; control_board.xlsx itself is never touched here (reading
    the board never writes to it). First active occurrence keeps its plain label; later
    ones get numbered, in row order."""
    seen: dict[str, int] = {}
    renamed: list[str] = []
    for row in runs:
        if not row.active:
            continue
        label = row.resolved.get("label")
        seen[label] = seen.get(label, 0) + 1
        if seen[label] > 1:
            new_label = f"{label} ({seen[label]})"
            row.resolved["label"] = new_label
            renamed.append(f"row {row.row_num}: '{label}' -> '{new_label}'")
    return renamed


def read_board(path: Path = CONTROL_BOARD_PATH) -> BoardResult:
    if not path.exists():
        raise FileNotFoundError(
            f"No control board at {path}. Run `python conductor.py --make-board` first."
        )
    wb = load_workbook(path, data_only=True)
    board_errors: list[str] = []

    if "Runs" not in wb.sheetnames:
        raise ValueError(f"{path}: no 'Runs' sheet found")
    raw_rows = _read_runs_sheet(wb["Runs"])
    runs = [resolve_row(raw, row_num=i + 2) for i, raw in enumerate(raw_rows)]

    settings: dict[str, object] = {}
    if "Settings" in wb.sheetnames:
        raw_settings = _read_settings_sheet(wb["Settings"])
        settings, setting_errors = _resolve_settings(raw_settings)
        board_errors.extend(setting_errors)
    else:
        settings = {s.name: s.default for s in spec.SETTINGS}

    renamed = _dedupe_active_labels(runs)
    if renamed:
        board_errors.append(
            "duplicate active labels auto-renamed to guarantee uniqueness -- label is a "
            "dict key everywhere downstream (current_picks, Step2_picks, walk-forward "
            "candidate sets), so an unresolved collision would silently drop one row's "
            "results rather than error: " + "; ".join(renamed)
        )

    return BoardResult(runs=runs, settings=settings, board_errors=board_errors)


# ---------------------------------------------------------------------------
# Board generation — `--make-board`
# ---------------------------------------------------------------------------


def _existing_rows(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if "Runs" not in wb.sheetnames:
        return []
    return _read_runs_sheet(wb["Runs"])


def _existing_settings(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    if "Settings" not in wb.sheetnames:
        return {}
    return _read_settings_sheet(wb["Settings"])


_UNKNOWN_FILL = PatternFill("solid", fgColor="F2F2F2")


def _write_runs_sheet(wb: Workbook, preserved_rows: list[dict[str, object]]) -> None:
    """Regenerate the Runs header + preserved row data. A column outside the current
    schema (SM's own scratch notes, e.g. `avgGain`) is carried through UNCHANGED as a
    trailing column, never discarded -- the docstring on write_board() has always promised
    this, but until 2026-08-12 the implementation only reported the name and threw the
    values away, which cost SM two such columns' data on a live board. Fixed now: unknown
    columns get appended after the schema's own, in first-seen order, tinted grey to mark
    them as not board-validated."""
    ws = wb.create_sheet("Runs")

    # One-time 2026-08-13 migration: `active` split into independent `D`/`P` columns (see
    # DesignVersion2.md's 2026-08-13 correction). A legacy `active` mark becomes `D` only,
    # NEVER `P` -- the whole point of the split is that a row mid-development must not
    # silently start shipping to production the next time --make-board regenerates the
    # board. The original `active` cell is left in place too (picked up as an "extra"
    # column below), so the migration is visible, not hidden.
    migrated = 0
    for old_row in preserved_rows:
        if "D" not in old_row and old_row.get("active"):
            old_row["D"] = old_row["active"]
            migrated += 1
    if migrated:
        print(f"[control_board] {migrated} row(s): legacy `active` mark migrated to `D` "
              f"(2026-08-13 P/D split) -- `P` left BLANK; mark it explicitly on any row "
              f"that should also ship to production.")

    extra: list[str] = []
    for old_row in preserved_rows:
        for name in old_row:
            if name and name not in spec.PARAMS_BY_NAME and name not in extra:
                extra.append(name)
    headers = list(spec.RUNS_COLUMNS) + extra

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        pdef = spec.PARAMS_BY_NAME.get(name)
        if pdef is None:
            cell.fill = _UNKNOWN_FILL
        else:
            cell.fill = _META_FILL if pdef.step is None else _STEP_FILLS.get(pdef.step, _HEADER_FILL)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r, old_row in enumerate(preserved_rows, start=2):
        for c, name in enumerate(headers, start=1):
            if name in old_row:
                ws.cell(row=r, column=c, value=old_row[name])
    if extra:
        print(f"[control_board] columns not in the schema, carried through unchanged: {sorted(extra)}")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_settings_sheet(wb: Workbook, preserved: dict[str, object]) -> None:
    ws = wb.create_sheet("Settings")
    ws.cell(row=1, column=1, value="setting").font = Font(bold=True)
    ws.cell(row=1, column=2, value="value").font = Font(bold=True)
    ws.cell(row=1, column=3, value="help").font = Font(bold=True)
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = _HEADER_FILL
    for r, sdef in enumerate(spec.SETTINGS, start=2):
        ws.cell(row=r, column=1, value=sdef.name)
        ws.cell(row=r, column=2, value=preserved.get(sdef.name, sdef.default))
        ws.cell(row=r, column=3, value=sdef.help).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 90


def _write_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Legend")
    headers = ["column", "step", "type", "default", "choices", "help"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for r, pdef in enumerate(spec.PARAMS, start=2):
        ws.cell(row=r, column=1, value=pdef.name)
        ws.cell(row=r, column=2, value="meta" if pdef.step is None else pdef.step)
        ws.cell(row=r, column=3, value=pdef.dtype)
        ws.cell(row=r, column=4, value=str(pdef.default) if pdef.default is not None else "")
        ws.cell(row=r, column=5, value=", ".join(pdef.choices) if pdef.choices else "")
        ws.cell(row=r, column=6, value=pdef.help).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 90


def write_board(path: Path = CONTROL_BOARD_PATH) -> Path:
    """(Re)generate the board from param_spec, preserving any existing Runs rows and
    Settings overrides. Safe to call repeatedly — it only ever adds columns the schema
    has grown, never drops user data (a column the schema no longer has is reported,
    not deleted, so nothing vanishes silently)."""
    preserved_rows = _existing_rows(path)
    preserved_settings = _existing_settings(path)

    CONTROL_ROOT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)   # drop the default blank sheet
    _write_runs_sheet(wb, preserved_rows)
    _write_settings_sheet(wb, preserved_settings)
    _write_legend_sheet(wb)
    wb.save(path)
    return path


if __name__ == "__main__":
    # Quick manual check: `python control_board.py` regenerates then prints a summary.
    p = write_board()
    print(f"Wrote {p}")
    result = read_board(p)
    print(f"{len(result.runs)} row(s), {len(result.active_runs)} active, "
          f"ok={result.ok}")
