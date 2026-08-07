"""
Verdict: do extreme gains come from low-conformity group members?

Reads the conformity grades from analyze_conformity.py and longi_future_per{20,50}d.csv,
builds a full (ticker, daynum) panel, and buckets it by conformity decile (and,
separately, by sector-beta decile) to see whether low-conformity members produce
wider dispersion / fatter tails in forward gain — not just a different mean.
Every stat is repeated on the first vs. second half of history: a pattern that
only holds in one half is not a finding (see docs/1_group_conformity.md).

A secondary, explicitly low-power check reads the existing DomGICS_* run*.xlsx
reports (../../../_archive/strategy_grp/app/report/) to see whether a hop's realized
gain correlates with its focusset's mean/min conformity. With ~30 independent lots
this can only corroborate the panel result, never decide it on its own. strategy_grp
was retired on 2026-08-07, so those reports are frozen — this check no longer moves
from run to run, and it degrades quietly (skipped, warning only) if they go away.

Usage:
    python analyze_conformity_gains.py
"""
import argparse
import glob
import os

import numpy as np
import pandas as pd

ATTRS = ["GICS", "Sector2"]
# Forward horizons, named by the longi_future_per* suffix — the "seven-pack" literal
# trading-day ladder (1d/5d/10d/20d/50d/100d/200d) produced by
# longi/app/code/longi_future_performance.py. These replaced future_gain{20,50}d.csv on
# 2026-07-31 — first via an intermediate 1m/3m (22/66-day) naming, then the same day
# corrected to this literal 20d/50d ladder. The entry convention also changed (entry is now
# the day AFTER the signal day), so results are NOT comparable to conformity_vs_gain.csv rows
# written before 2026-07-31, even though "20d"/"50d" happen to match the old future_gain20d/
# 50d day counts numerically.
HORIZONS = ["20d", "50d"]
N_DECILES = 10


def _read_matrix(path):
    df = pd.read_csv(path, sep=";", decimal=",", index_col=0)
    df.columns = df.columns.astype(int)
    return df


def _decile_stats(panel, bucket_col, value_col="gain"):
    """Per-decile-of-bucket_col: count, central tendency, dispersion, tails."""
    ranks, bin_edges = pd.qcut(panel[bucket_col], N_DECILES, labels=False, duplicates="drop", retbins=True)
    panel = panel.assign(_decile=ranks + 1)  # 1..N, 1 = lowest conformity/lowest beta
    rows = []
    for decile, grp in panel.groupby("_decile"):
        g = grp[value_col]
        rows.append({
            "decile": int(decile),
            "n": len(g),
            "mean_gain": g.mean(),
            "median_gain": g.median(),
            "std_gain": g.std(),
            "p5_gain": g.quantile(0.05),
            "p95_gain": g.quantile(0.95),
            "pct_abs_gain_gt10": (g.abs() > 10).mean() * 100,
        })
    return pd.DataFrame(rows).sort_values("decile")


def _build_panel(attr, horizon, conf_dir, gain_dir):
    conf = _read_matrix(os.path.join(conf_dir, f"longi_conf_{attr}.csv"))
    beta = _read_matrix(os.path.join(conf_dir, f"longi_sectorbeta_{attr}.csv"))
    gain = _read_matrix(os.path.join(gain_dir, f"longi_future_per{horizon}.csv"))

    tickers = conf.index.intersection(beta.index).intersection(gain.index)
    daynums = conf.columns.intersection(beta.columns).intersection(gain.columns)
    conf = conf.loc[tickers, daynums]
    beta = beta.loc[tickers, daynums]
    gain = gain.loc[tickers, daynums]

    panel = pd.concat(
        {"conf": conf.stack(), "beta": beta.stack(), "gain": gain.stack()}, axis=1
    )
    panel.index.names = ["ticker", "daynum"]
    panel = panel.dropna(subset=["conf", "gain"])
    return panel.reset_index()


def _verdict_for(attr, horizon, panel, out_rows):
    splits = {"all": panel}
    if len(panel) > 200:
        mid = panel["daynum"].median()
        splits["first_half"] = panel[panel["daynum"] <= mid]
        splits["second_half"] = panel[panel["daynum"] > mid]

    for split_name, sub in splits.items():
        if len(sub) < N_DECILES * 5:
            continue
        for bucket_type, col in [("conformity", "conf"), ("beta", "beta")]:
            sub_valid = sub.dropna(subset=[col])
            if len(sub_valid) < N_DECILES * 5:
                continue
            stats = _decile_stats(sub_valid, col)
            stats.insert(0, "split", split_name)
            stats.insert(0, "bucket_type", bucket_type)
            stats.insert(0, "horizon", horizon)
            stats.insert(0, "attribute", attr)
            out_rows.append(stats)


def _monotonicity_note(stats_df):
    """Spearman-style check: does std_gain move monotonically with decile?"""
    if stats_df is None or len(stats_df) < 3:
        return "insufficient buckets"
    corr = stats_df["decile"].corr(stats_df["std_gain"], method="spearman")
    if pd.isna(corr):
        return "undefined"
    if corr > 0.6:
        return f"std_gain RISES with decile (rho={corr:.2f}) -> higher conformity = MORE dispersion (unexpected)"
    if corr < -0.6:
        return f"std_gain FALLS with decile (rho={corr:.2f}) -> low conformity = more dispersion, as hypothesized"
    return f"no clear monotone pattern (rho={corr:.2f})"


def _secondary_hop_check(conf_dir, strategy_grp_report_dir):
    """
    Low-power corroboration: for each DomGICS_* strategy's latest run*.xlsx,
    pull the focusset per hop from the Operational sheet's ticker grid and the
    hop's realized gain from HopData (both raw values, no formulas involved),
    then correlate hop gain against the focusset's mean/min conformity.
    """
    try:
        import openpyxl
    except ImportError:
        print("\n[secondary check] openpyxl not available, skipping.")
        return None

    conf_mats = {}
    for attr in ATTRS:
        p = os.path.join(conf_dir, f"longi_conf_{attr}.csv")
        if os.path.exists(p):
            conf_mats[attr] = _read_matrix(p)

    if not conf_mats:
        print("\n[secondary check] no conformity matrices found, skipping.")
        return None

    results = []
    strategy_dirs = sorted(glob.glob(os.path.join(strategy_grp_report_dir, "DomGICS_*")))
    for sdir in strategy_dirs:
        strategy = os.path.basename(sdir)
        runs = sorted(glob.glob(os.path.join(sdir, "run*.xlsx")))
        if not runs:
            continue
        run_path = runs[-1]
        try:
            hopdata = pd.read_excel(run_path, sheet_name="HopData")
        except Exception as e:
            print(f"[secondary check] {strategy}: could not read HopData ({e}), skipping.")
            continue
        if "daynum" not in hopdata.columns or "gain" not in hopdata.columns:
            print(f"[secondary check] {strategy}: HopData missing daynum/gain columns, skipping.")
            continue

        try:
            wb = openpyxl.load_workbook(run_path, data_only=True)
            ws = wb["Operational"]
        except Exception as e:
            print(f"[secondary check] {strategy}: could not read Operational sheet ({e}), skipping.")
            continue

        # daynum header row is row 1, columns B..
        max_col = ws.max_column
        col_daynum = {}
        for c in range(2, max_col + 1):
            v = ws.cell(1, c).value
            if v is not None:
                try:
                    col_daynum[c] = int(v)
                except (TypeError, ValueError):
                    pass

        # locate ticker block: starts after optional dominance_cutoff row (row 3),
        # ends at N_survivors or avg_gain label, whichever comes first.
        ticker_top = 4 if ws.cell(3, 1).value == "dominance_cutoff" else 3
        ticker_bottom = None
        for r in range(ticker_top, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v in ("N_survivors", "avg_gain"):
                ticker_bottom = r  # exclusive
                break
        if ticker_bottom is None or ticker_bottom <= ticker_top:
            print(f"[secondary check] {strategy}: could not locate ticker block, skipping.")
            continue

        per_hop = []
        for c, dn in col_daynum.items():
            picks = [ws.cell(r, c).value for r in range(ticker_top, ticker_bottom)]
            picks = [p for p in picks if p]
            if not picks:
                continue
            for attr, mat in conf_mats.items():
                if dn not in mat.columns:
                    continue
                vals = mat.loc[mat.index.intersection(picks), dn].dropna()
                if vals.empty:
                    continue
                per_hop.append({"daynum": dn, "attribute": attr,
                                 "mean_conf": vals.mean(), "min_conf": vals.min()})
        if not per_hop:
            print(f"[secondary check] {strategy}: no matched hops, skipping.")
            continue

        per_hop_df = pd.DataFrame(per_hop)
        merged = per_hop_df.merge(hopdata[["daynum", "gain"]], on="daynum", how="inner")
        for attr, grp in merged.groupby("attribute"):
            if len(grp) < 8:
                continue
            corr_mean = grp["mean_conf"].corr(grp["gain"])
            corr_min = grp["min_conf"].corr(grp["gain"])
            results.append({
                "strategy": strategy, "attribute": attr, "n_hops": len(grp),
                "corr(hop_gain, focusset_mean_conf)": corr_mean,
                "corr(hop_gain, focusset_min_conf)": corr_min,
            })

    if not results:
        print("\n[secondary check] no usable hop/ticker data found across DomGICS_* reports.")
        return None
    return pd.DataFrame(results)


def run(input_dir, conf_dir, output_dir, strategy_grp_report_dir):
    input_dir = os.path.expanduser(input_dir)
    conf_dir = os.path.expanduser(conf_dir)
    output_dir = os.path.expanduser(output_dir)
    strategy_grp_report_dir = os.path.expanduser(strategy_grp_report_dir)
    os.makedirs(output_dir, exist_ok=True)

    all_stats = []
    monotonicity_notes = []
    for attr in ATTRS:
        for horizon in HORIZONS:
            print(f"\n=== {attr} / {horizon} ===")
            panel = _build_panel(attr, horizon, conf_dir, input_dir)
            print(f" - panel rows: {len(panel)}")
            if panel.empty:
                continue
            _verdict_for(attr, horizon, panel, all_stats)

    if not all_stats:
        print("No panels built — check that analyze_conformity.py has been run first.")
        return

    out_df = pd.concat(all_stats, ignore_index=True)
    out_path = os.path.join(output_dir, "conformity_vs_gain.csv")
    out_df.to_csv(out_path, sep=";", decimal=",", index=False)
    print(f"\nWrote {out_path} ({len(out_df)} rows)")

    print("\n--- Monotonicity check (dispersion vs conformity decile, full history) ---")
    for attr in ATTRS:
        for horizon in HORIZONS:
            sub = out_df[
                (out_df["attribute"] == attr) & (out_df["horizon"] == horizon)
                & (out_df["split"] == "all") & (out_df["bucket_type"] == "conformity")
            ]
            note = _monotonicity_note(sub)
            print(f" {attr} / {horizon}: {note}")
            monotonicity_notes.append((attr, horizon, "all", note))

    print("\n--- Half-history consistency check (does the pattern survive a split?) ---")
    for attr in ATTRS:
        for horizon in HORIZONS:
            for split in ("first_half", "second_half"):
                sub = out_df[
                    (out_df["attribute"] == attr) & (out_df["horizon"] == horizon)
                    & (out_df["split"] == split) & (out_df["bucket_type"] == "conformity")
                ]
                note = _monotonicity_note(sub)
                print(f" {attr} / {horizon} / {split}: {note}")

    print("\n=== Secondary check: DomGICS_* hop-level corroboration (low power, ~dozens of hops) ===")
    hop_df = _secondary_hop_check(conf_dir, strategy_grp_report_dir)
    if hop_df is not None:
        hop_path = os.path.join(output_dir, "conformity_vs_gain_hop_secondary.csv")
        hop_df.to_csv(hop_path, sep=";", decimal=",", index=False)
        print(hop_df.to_string(index=False))
        print(f"Wrote {hop_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verdict: low-conformity group members vs. forward gain dispersion.")
    parser.add_argument("--input_dir", type=str, default="~/potentials/group_conformity/app/input",
                        help="Directory with longi_future_per{20,50}d.csv")
    parser.add_argument("--conf_dir", type=str, default="~/potentials/group_conformity/app/output",
                        help="Directory with longi_conf_*.csv / longi_sectorbeta_*.csv (analyze_conformity.py output)")
    parser.add_argument("--output_dir", type=str, default="~/potentials/group_conformity/app/output")
    parser.add_argument("--strategy_grp_report_dir", type=str,
                        default="~/potentials/_archive/strategy_grp/app/report",
                        help="Where DomGICS_*/run*.xlsx live (frozen since strategy_grp "
                             "was retired 2026-08-07), for the secondary check")
    args = parser.parse_args()
    run(args.input_dir, args.conf_dir, args.output_dir, args.strategy_grp_report_dir)
