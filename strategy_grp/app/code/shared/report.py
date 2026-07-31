"""
Write reports for strategy runs.

Per run:   app/report/<strategy>/run<N>_<date>.xlsx
           Sheet "Operational" — ticker rows + avg rows + ref rows + attribute count rows
           Sheet "Summary"     — metadata + avg gains

Master:    app/report/summary.csv  (one appended row per run, all strategies)
"""

import csv
import statistics
import time
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.chain import realizable_chain, chain_lot_stats, chain_origin_sensitivity
from shared.config import REPORT_ROOT, SUMMARY_CSV, future_gain_file
from shared.data_loader import daynum_to_date, load_longi, load_stamdata


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hop_avg_topn(hop: dict, gain_key: str, n: int) -> float:
    """Average gain for the top-n tickers (rank order) in one hop."""
    tickers = hop.get("tickers", [])[:n]
    gains   = hop.get(gain_key, {})
    vals    = [gains[t] for t in tickers if pd.notna(gains.get(t, float("nan")))]
    return sum(vals) / len(vals) if vals else float("nan")


def _grand_avg_topn(hop_results: list[dict], gain_key: str, n: int,
                    params: dict | None = None) -> float:
    """Average gain for the top-n tickers across active hops (respects No_go_GSPC_rsi)."""
    threshold = params.get("No_go_GSPC_rsi") if params else None
    vals: list[float] = []
    for h in hop_results:
        if threshold is not None:
            gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            if not pd.isna(gspc) and gspc < threshold:
                continue
        tickers = h.get("tickers", [])[:n]
        gains   = h.get(gain_key, {})
        vals.extend(gains[t] for t in tickers if pd.notna(gains.get(t, float("nan"))))
    return sum(vals) / len(vals) if vals else float("nan")


def _avg_rows(n: int) -> list[tuple[str, str, int]]:
    """The single avg-gain row for the active focusset size — (label, gain_key, top_n).

    Horizon-agnostic: the forward horizon is the `period` param; gains live under the
    single key "gains" in each hop. Names carry no 20d/50d suffix so one run = one column.
    """
    return [("avg_gain", "gains", n)]


# Gains for the chosen horizon live under this single hop key.
_GAIN_KEY = "gains"


# ---------------------------------------------------------------------------
# Benchmark rows — mkt_gain / alpha / beta
#
# `alpha` here is ACTIVE RETURN (excess over benchmark): focusset gain minus the
# benchmark's gain for the same daynum. It is NOT Jensen's alpha — it assumes beta = 1
# and a zero risk-free rate, and makes no regression. The distinction is not academic
# for this strategy: the focusset runs at ~1.7x market beta (mean beta3m 1.68, realized
# 1.87 by regression on 20d hops), so a chunk of a positive alpha is levered market
# exposure rather than selection. That is exactly why `beta` is reported beside it —
# read the two together, and discount alpha when beta is high.
#
# Jensen's alpha was considered and rejected: it needs a fitted beta, and with ~31
# INDEPENDENT 20d lots in this history (124 hops overlapping 4x at step=5) and R^2 ~ 0.05
# the intercept carries more estimation noise than the correction is worth. Active return
# is definitional — nothing to estimate, nothing to drift.
#
# The benchmark is the equal-weighted cross-sectional mean of EVERY ticker on that daynum,
# deliberately not a cap-weighted index: "the average stock you could have picked that
# day" is the right counterfactual for a stock-picking strategy.
# ---------------------------------------------------------------------------

def _market_gain(period: int) -> pd.Series:
    """Benchmark: equal-weighted cross-sectional mean forward gain per daynum."""
    return load_longi(future_gain_file(period)).mean(axis=0)


def _beta_frame() -> pd.DataFrame | None:
    """longi_beta3m.csv, or None when absent — the beta row is optional, and a missing
    factor file must not take the whole report down with it."""
    try:
        return load_longi("longi_beta3m.csv")
    except (FileNotFoundError, OSError):
        return None


def _hop_market(hop: dict, mkt: pd.Series) -> float:
    """The benchmark's gain for this hop's daynum."""
    v = mkt.get(str(hop["daynum"]), float("nan"))
    return float(v) if pd.notna(v) else float("nan")


def _hop_beta(hop: dict, beta_df: pd.DataFrame | None, n: int) -> float:
    """Mean beta3m across this hop's focusset — context for the alpha row."""
    if beta_df is None:
        return float("nan")
    col = str(hop["daynum"])
    if col not in beta_df.columns:
        return float("nan")
    vals = [beta_df.at[t, col] for t in hop.get("tickers", [])[:n]
            if t in beta_df.index and pd.notna(beta_df.at[t, col])]
    return sum(vals) / len(vals) if vals else float("nan")


def _grand_avg_alpha(hop_results: list[dict], n: int, params: dict) -> tuple[float, float]:
    """(mean alpha, mean beta) across active hops — same No_go gating as _grand_avg_topn,
    so the headline alpha is measured over exactly the hops the strategy would trade."""
    threshold = params.get("No_go_GSPC_rsi")
    mkt       = _market_gain(params.get("period", 22))
    beta_df   = _beta_frame()
    alphas: list[float] = []
    betas:  list[float] = []
    for h in hop_results:
        if threshold is not None:
            gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            if not pd.isna(gspc) and gspc < threshold:
                continue
        gain = _hop_avg_topn(h, _GAIN_KEY, n)
        m    = _hop_market(h, mkt)
        if pd.notna(gain) and pd.notna(m):
            alphas.append(gain - m)
        b = _hop_beta(h, beta_df, n)
        if pd.notna(b):
            betas.append(b)
    return (sum(alphas) / len(alphas) if alphas else float("nan"),
            sum(betas) / len(betas) if betas else float("nan"))


def _chain_metric_labels() -> list[str]:
    """Realizable-chain summary labels, in write order."""
    return ["chain_ret", "chain_annual", "chain_n"]


def _chain_metrics(hop_results: list[dict], gain_key: str, n: int,
                   hold: int, params: dict) -> tuple[float, float, int]:
    """
    Realizable non-overlapping ADDITIVE chain over this run's full span.

    Thin wrapper over shared.chain.realizable_chain (the single source of the chain
    math). best_strategy.py re-runs the same function with a common floor/cap so
    cross-strategy chain returns are comparable; see shared/chain.py.

    Returns (total_return_pct, annual_pct, n_trades) — additive sum of lot gains and
    its simple per-year average (not a compound CAGR).
    """
    threshold = params.get("No_go_GSPC_rsi")
    rows = ((h["daynum"], _hop_avg_topn(h, gain_key, n),
             h.get("ref_values", {}).get("^GSPC_rsi", float("nan")))
            for h in hop_results)
    return realizable_chain(rows, hold, threshold, phase_average=True)


def _count_attr(hop_results: list[dict], stamdata: pd.DataFrame,
                attr_col: str) -> tuple[list[str], list[dict]]:
    """
    For each hop count occurrences of each unique value of attr_col in the focusset.
    Returns (sorted_unique_values, list_of_count_dicts) — values sorted by total desc.
    """
    all_vals: set[str] = set()
    hop_counts: list[dict] = []

    for h in hop_results:
        counts: dict[str, int] = {}
        for ticker in h.get("tickers", []):
            if ticker in stamdata.index:
                raw = stamdata.at[ticker, attr_col]
                if pd.notna(raw):
                    v = str(raw).strip()
                    if v:
                        counts[v] = counts.get(v, 0) + 1
                        all_vals.add(v)
        hop_counts.append(counts)

    totals = {v: sum(hc.get(v, 0) for hc in hop_counts) for v in all_vals}
    sorted_vals = sorted(all_vals, key=lambda v: -totals[v])
    return sorted_vals, hop_counts


def _informational_attr_names(params: dict) -> list[str]:
    """params['informational_attributes'] (Step 3 — display only) normalized to a list
    of longi factor short names — [] when absent (plain filter strategies). Deferred
    import: shared.dominance imports this module at load time, so importing it back at
    module level here would cycle; by the time this function actually runs, both
    modules are fully initialized.
    """
    from shared.dominance import informational_attr_list
    return informational_attr_list(params.get("informational_attributes"))


def _write_informational_attr_stock_rows(ws, hop_results: list[dict], start_row: int,
                                          attr: str, n_tickers: int) -> None:
    """N rows (attr_1 .. attr_N), one per focusset rank position — a direct per-stock draw
    from longi_{attr}.csv, no aggregation. Row i aligns with the ticker_rank(i+1) row above."""
    info_df = load_longi(f"longi_{attr}.csv")
    for i in range(n_tickers):
        row      = start_row + i
        lbl_cell = ws.cell(row, 1, f"{attr}_{i + 1}")
        lbl_cell.font = _BOLD
        for j, h in enumerate(hop_results, start=2):
            tickers = h.get("tickers", [])
            col     = str(h["daynum"])
            cell    = ws.cell(row, j)
            cell.alignment = _CTR
            if i < len(tickers):
                t = tickers[i]
                if t in info_df.index and col in info_df.columns and pd.notna(info_df.at[t, col]):
                    cell.value = round(float(info_df.at[t, col]), 2)
                else:
                    cell.value = None
            else:
                cell.value = None


def _param_cell(v):
    """Excel/CSV-safe representation of a param value.

    openpyxl rejects a raw list as a cell value, and a bare list also reads ugly in
    the CSV — so a list param (e.g. multiple informational_attributes names) is joined into one
    comma-separated string; everything else passes through unchanged.
    """
    return ",".join(v) if isinstance(v, list) else v


def _next_run_num(folder: Path) -> int:
    nums = []
    for p in folder.glob("run*.xlsx"):
        try:
            nums.append(int(p.name[3:].split("_")[0]))
        except ValueError:
            pass
    return max(nums, default=0) + 1


def _count_active_hops(hop_results: list[dict], params: dict) -> int:
    """Count hops actually invested: a non-empty focusset surviving No_go_GSPC_rsi.

    No-pick hops (empty tickers — days the strategy sat out, now recorded so they are
    visible in HopData/Operational) are NOT active, so N_hops_active is the number of
    days with a real investment (the chain-side mirror of the ladder's invested count).
    NaN RSI counts as passing the No_go gate.
    """
    threshold = params.get("No_go_GSPC_rsi")
    count = 0
    for h in hop_results:
        if not h.get("tickers"):
            continue
        gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
        if threshold is None or pd.isna(gspc) or gspc >= threshold:
            count += 1
    return count


def _usable_daynums(hop_results: list[dict], gain_key: str, params: dict) -> list[int]:
    """Daynums of hops that actually invest: a real (non-NaN) top-N gain surviving No_go.

    This is the strategy's *usable* span. The cross / win-prob strategies have no real
    gain before the ML warm-up (~daynum 1797), so their usable range starts later than
    the full evaluated range — even though the empty warm-up hops are still recorded.
    """
    threshold = params.get("No_go_GSPC_rsi")
    n = params.get("focusset_size", 10)
    dns: list[int] = []
    for h in hop_results:
        if threshold is not None:
            gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            if not pd.isna(gspc) and gspc < threshold:
                continue
        if pd.notna(_hop_avg_topn(h, gain_key, n)):
            dns.append(int(h["daynum"]))
    return dns


def _chain_hop_rows(hop_results: list[dict], n: int) -> list[tuple[int, float, float]]:
    """(daynum, top-N avg gain, gspc_rsi) per hop — the chain's raw inputs (one span)."""
    return [(h["daynum"], _hop_avg_topn(h, _GAIN_KEY, n),
             h.get("ref_values", {}).get("^GSPC_rsi", float("nan")))
            for h in hop_results]


def _chain_dispersion(hop_results: list[dict], n: int, hold: int,
                      params: dict) -> tuple[float, int, float]:
    """Chained WORST-CASE dispersion over this run's own span (no floor/cap).

    Returns (worst_lot, n_loss, origin_sens_pct) using the SAME rule best_strategy.py
    applies on the common span — worst = lowest lot over all start origins, n_loss = most
    losers in any one origin's chain (worst < 0 <=> n_loss >= 1), sens = annual spread
    across origins. So the run-file Summary and the comparison sheet reconcile (they differ
    only by span, exactly as chain_ret does).
    """
    thr  = params.get("No_go_GSPC_rsi")
    rows = _chain_hop_rows(hop_results, n)
    _avg, worst, n_loss = chain_lot_stats(rows, hold, thr, phase_average=True)
    sens = chain_origin_sensitivity(rows, hold, thr)
    return worst, n_loss, sens


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_BOLD     = Font(bold=True)
_SMALL    = Font(size=9)
_HDR_FILL = PatternFill("solid", fgColor="BDD7EE")  # light blue header
_GRN_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red
_GRY_FILL = PatternFill("solid", fgColor="EEEEEE")  # gray (n/a)
_SEP_FILL = PatternFill("solid", fgColor="D9E1F2")  # separator 20d→50d
_REF_FILL   = PatternFill("solid", fgColor="FFF2CC")  # yellow for ref rows
_INPUT_FILL = PatternFill("solid", fgColor="FFE599")  # amber for editable input cells
_SURV_FILL  = PatternFill("solid", fgColor="DDEBF7")  # pale blue for N_survivors row
_CUTOFF_FILL = PatternFill("solid", fgColor="E2EFDA")  # pale green for dominance_cutoff row
_BENCH_FILL  = PatternFill("solid", fgColor="F2F2F2")  # pale grey for mkt_gain/alpha/beta

_ATTR_FILLS: dict[str, PatternFill] = {
    "GICS":    PatternFill("solid", fgColor="D9D2E9"),  # light purple
    "Sector2": PatternFill("solid", fgColor="FCE5CD"),  # light peach
    "Zone":    PatternFill("solid", fgColor="D0E0E3"),  # light teal
}

_PCT_FMT = '+0.00;-0.00;"-"'
_CTR     = Alignment(horizontal="center")


def _gain_fill(val: float) -> PatternFill:
    if pd.isna(val):
        return _GRY_FILL
    return _GRN_FILL if val >= 0 else _RED_FILL


# ---------------------------------------------------------------------------
# Operational sheet
# ---------------------------------------------------------------------------

def _fill_operational(ws, hop_results: list[dict], params: dict) -> None:
    """
    Row 1   : (blank)  | daynum1 | daynum2 | ...
    Row 2   : (blank)  | date1   | date2   | ...
    Row 3   : dominance_cutoff (only for strategies that expose one, e.g. DomGICS) —
              that day's Step-1 decile cutoff value; absent otherwise, in which case
              everything below shifts up accordingly
    Rows …n: (blank) | ticker_rank1 … ticker_rankN
    +1 row  : avg_gain
    +2 rows : mkt_gain (benchmark for that daynum) and alpha (avg_gain - mkt_gain,
              ACTIVE RETURN, not Jensen's — see the _market_gain block), plus a third
              `beta` row when longi_beta3m.csv is available
    +N rows per informational_attributes entry (only for strategies that expose one,
              e.g. DomGICS), N = focusset_size: <attr>_1 .. <attr>_N, a direct per-stock
              draw from longi_{attr}.csv for that hop's rank-i ticker — no aggregation,
              row i aligned with the ticker_rank(i+1) row above; absent otherwise, in
              which case everything below shifts up accordingly
    +4 rows : market context (^GSPC_rsi, ^STOXX_rsi, ^HSI_rsi, ^VIX)
    +?? rows: GICS count rows (sorted by frequency)
    +?? rows: Sector2 count rows
    +?? rows: Zone count rows
    """
    daynums   = [h["daynum"] for h in hop_results]
    n         = params.get("focusset_size", max(len(h.get("tickers", [])) for h in hop_results))
    n_tickers = n
    threshold = params.get("No_go_GSPC_rsi")

    # Optional dominance_cutoff row, inserted at row 3 (just under the date row), above
    # the ticker rows — only for strategies whose hops carry a Step-1 day-by-day cutoff
    # (the DomGICS_* family). Ticker rows and everything below shift down accordingly.
    has_cutoff = any(h.get("dom_cutoff") is not None for h in hop_results)
    cutoff_off = 1 if has_cutoff else 0

    # Optional N_survivors row, inserted directly below the ticker rows.
    has_surv = any("n_survivors" in h for h in hop_results)
    surv_off = 1 if has_surv else 0

    # Optional informational_attributes per-stock rows (N rows per attribute, N =
    # focusset_size), inserted below the avg rows — row-dynamic so a bigger focusset_size
    # (more ticker rows) or more attributes never collide with what follows.
    info_attrs = _informational_attr_names(params)
    n_info     = n_tickers * len(info_attrs)

    # Benchmark rows below the avg rows: mkt_gain + alpha always, beta only when
    # longi_beta3m.csv is present. See the _market_gain/_hop_beta block above for what
    # `alpha` means here (active return, NOT Jensen's).
    mkt_series = _market_gain(params.get("period", 22))
    beta_df    = _beta_frame()
    has_beta   = beta_df is not None and any(
        pd.notna(_hop_beta(h, beta_df, n)) for h in hop_results)
    n_bench    = 2 + (1 if has_beta else 0)

    # ---- section top rows, derived ONCE and chained ----
    # Every section's first row is the previous section's top plus its height. This used to
    # be four independent copies of the same running sum, and inserting a section meant
    # updating all four: miss one and the ref rows silently overwrite the informational
    # rows while the No_go formula points at the wrong row. Derive, never re-add.
    rows_list   = _avg_rows(n)
    ticker_top  = 3 + cutoff_off
    avg_top     = ticker_top + n_tickers + surv_off
    bench_top   = avg_top + len(rows_list)
    info_top    = bench_top + n_bench
    ref_top     = info_top + n_info

    # The row where ^GSPC_rsi lands, for the No_go formulas in the avg/benchmark rows.
    ref_keys     = list(hop_results[0].get("ref_values", {}).keys()) if hop_results else []
    gspc_rsi_row = next((ref_top + i for i, k in enumerate(ref_keys) if "GSPC_rsi" in k), None)

    # ---- header rows ----
    # A1/A2 hold the No_go label and editable threshold (safe for any focusset size).
    if threshold is not None:
        c = ws.cell(1, 1, "No_go_GSPC_rsi"); c.font = _BOLD
        c = ws.cell(2, 1, threshold);        c.font, c.fill, c.number_format = _BOLD, _INPUT_FILL, "0"
    else:
        ws.cell(1, 1).fill = _HDR_FILL
        ws.cell(2, 1).fill = _HDR_FILL
    for j, dn in enumerate(daynums, start=2):
        c = ws.cell(1, j, dn);                 c.font, c.fill, c.alignment = _BOLD, _HDR_FILL, _CTR
        c = ws.cell(2, j, daynum_to_date(dn)); c.font, c.fill, c.alignment = _SMALL, _HDR_FILL, _CTR

    # ---- dominance_cutoff row (optional — only when hops carry a Step-1 day-by-day
    # cutoff, the DomGICS_* family) — row 3, just under the date row, above the tickers.
    if has_cutoff:
        crow = 3
        c = ws.cell(crow, 1, "dominance_cutoff"); c.font, c.fill = _BOLD, _CUTOFF_FILL
        for j, h in enumerate(hop_results, start=2):
            val  = h.get("dom_cutoff")
            cell = ws.cell(crow, j)
            cell.fill, cell.alignment = _CUTOFF_FILL, _CTR
            cell.value = round(val, 2) if val is not None else None

    # ---- ticker rows ----
    for i in range(n_tickers):
        row = ticker_top + i
        ws.cell(row, 1, "")
        for j, h in enumerate(hop_results, start=2):
            tickers = h.get("tickers", [])
            ws.cell(row, j, tickers[i] if i < len(tickers) else "")

    # ---- N_survivors row (optional — only when hops carry the count) ----
    if has_surv:
        srow = ticker_top + n_tickers
        c = ws.cell(srow, 1, "N_survivors"); c.font, c.fill = _BOLD, _SURV_FILL
        for j, h in enumerate(hop_results, start=2):
            val  = h.get("n_survivors")
            cell = ws.cell(srow, j)
            cell.font, cell.fill, cell.alignment = _BOLD, _SURV_FILL, _CTR
            cell.value = int(val) if val is not None else None

    # ---- avg rows ----
    for idx, (label, gain_key, top_n) in enumerate(rows_list):
        row      = avg_top + idx
        sep      = (gain_key == "gains_50d")
        lbl_cell = ws.cell(row, 1, label)
        lbl_cell.font = _BOLD
        if sep:
            lbl_cell.fill = _SEP_FILL
        for j, h in enumerate(hop_results, start=2):
            val    = _hop_avg_topn(h, gain_key, top_n)
            gspc_p = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            no_go  = threshold is not None and not pd.isna(gspc_p) and gspc_p < threshold
            cell   = ws.cell(row, j)
            cell.font, cell.number_format, cell.alignment = _BOLD, _PCT_FMT, _CTR

            if pd.notna(val) and gspc_rsi_row is not None:
                # Interactive formula: recalculates when user edits A4
                col_ltr    = get_column_letter(j)
                cell.value = f'=IF({col_ltr}{gspc_rsi_row}<$A$2,"",{round(val, 4)})'
                cell.fill  = _GRY_FILL if no_go else _gain_fill(val)
            elif pd.notna(val):
                cell.value = None if no_go else round(val, 4)
                cell.fill  = _GRY_FILL if no_go else _gain_fill(val)
            else:
                cell.value = None
                cell.fill  = PatternFill() if sep else _GRY_FILL

    # ---- benchmark rows: mkt_gain, alpha, (beta) ----
    # mkt_gain and alpha carry the same No_go formula as the avg rows so that editing the
    # threshold in A2 blanks all three together — a visible alpha next to a blanked
    # avg_gain would read as a trade the strategy never took.
    bench_rows: list[tuple[str, str]] = [("mkt_gain", "mkt"), ("alpha", "alpha")]
    if has_beta:
        bench_rows.append(("beta", "beta"))

    for idx, (label, kind) in enumerate(bench_rows):
        row = bench_top + idx
        lbl = ws.cell(row, 1, label)
        lbl.font, lbl.fill = _BOLD, _BENCH_FILL
        for j, h in enumerate(hop_results, start=2):
            if kind == "mkt":
                val = _hop_market(h, mkt_series)
            elif kind == "alpha":
                gain = _hop_avg_topn(h, _GAIN_KEY, n)
                m    = _hop_market(h, mkt_series)
                val  = gain - m if pd.notna(gain) and pd.notna(m) else float("nan")
            else:
                val = _hop_beta(h, beta_df, n)

            gspc_p = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            no_go  = threshold is not None and not pd.isna(gspc_p) and gspc_p < threshold
            cell   = ws.cell(row, j)
            cell.alignment = _CTR

            if pd.isna(val):
                cell.value, cell.fill = None, _GRY_FILL
                continue
            if kind == "beta":                       # a ratio, not a percentage
                cell.number_format = "0.00"
                cell.value = None if no_go else round(val, 2)
                cell.fill  = _GRY_FILL if no_go else _BENCH_FILL
                continue
            cell.number_format = _PCT_FMT
            if gspc_rsi_row is not None:
                col_ltr    = get_column_letter(j)
                cell.value = f'=IF({col_ltr}{gspc_rsi_row}<$A$2,"",{round(val, 4)})'
            else:
                cell.value = None if no_go else round(val, 4)
            cell.fill = _GRY_FILL if no_go else _gain_fill(val)

    # ---- informational_attributes per-stock rows (optional — only strategies that
    # expose one) ----
    for k, attr in enumerate(info_attrs):
        _write_informational_attr_stock_rows(ws, hop_results, info_top + n_tickers * k,
                                              attr, n_tickers)

    # ---- ref rows ----
    def _write_ref_rows(start_row: int, hop_key: str, label_suffix: str = "") -> int:
        keys = list(hop_results[0].get(hop_key, {}).keys()) if hop_results else []
        for idx, key in enumerate(keys):
            lbl = ws.cell(start_row + idx, 1, f"{key}{label_suffix}")
            lbl.font, lbl.fill = _BOLD, _REF_FILL
            for j, h in enumerate(hop_results, start=2):
                val  = h.get(hop_key, {}).get(key, float("nan"))
                cell = ws.cell(start_row + idx, j)
                cell.fill, cell.alignment = _REF_FILL, _CTR
                if pd.notna(val):
                    cell.value         = round(val, 2)
                    cell.number_format = "0.0" if key.endswith("_rsi") else "0.00"
                else:
                    cell.value = None
        return len(keys)

    base  = ref_top
    n_ref = _write_ref_rows(base, "ref_values")

    # ---- attribute frequency rows ----
    stamdata = load_stamdata()
    attr_row = base + n_ref

    for attr_col in ("GICS", "Sector2", "Zone"):
        fill = _ATTR_FILLS[attr_col]
        sorted_vals, hop_counts = _count_attr(hop_results, stamdata, attr_col)

        for i, val in enumerate(sorted_vals):
            row      = attr_row + i
            label    = f"{attr_col}_{val}"
            lbl_cell = ws.cell(row, 1, label)
            lbl_cell.font, lbl_cell.fill = _BOLD, fill
            for j, hc in enumerate(hop_counts, start=2):
                count = hc.get(val, 0)
                cell  = ws.cell(row, j)
                cell.fill, cell.alignment = fill, _CTR
                cell.value = count if count > 0 else None

        attr_row += len(sorted_vals)

    # ---- dimensions ----
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(daynums) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 14
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _fill_summary(ws, strategy_name: str, run_num: int, params: dict,
                  hop_results: list[dict], extra_summary: dict | None = None) -> None:
    daynums = [h["daynum"] for h in hop_results]
    n       = params.get("focusset_size", 10)

    # StartDaynum/EndDaynum are the strategy's USABLE span, chronological (Start = oldest).
    usable   = _usable_daynums(hop_results, _GAIN_KEY, params)
    start_dn = min(usable) if usable else (min(daynums) if daynums else "")
    end_dn   = max(usable) if usable else (max(daynums) if daynums else "")

    rows: list[tuple] = [
        ("StrategyName",  strategy_name),
        ("Run#",          run_num),
        ("StartDaynum",   start_dn),
        ("N_hops",        len(hop_results)),
        ("N_hops_active", _count_active_hops(hop_results, params)),
        ("EndDaynum",     end_dn),
    ]
    for k, v in params.items():
        rows.append((k, _param_cell(v)))
        if k == "dominance_attribute_direction":
            # Average of the per-daynum Step-1 dominance cutoff (DomGICS_* only — other
            # strategies never carry dom_cutoff, so this row never appears for them).
            cutoff_vals = [h["dom_cutoff"] for h in hop_results if h.get("dom_cutoff") is not None]
            if cutoff_vals:
                rows.append(("dominance_cutoff_avg", round(statistics.mean(cutoff_vals), 4)))

    for label, gain_key, top_n in _avg_rows(n):
        val = _grand_avg_topn(hop_results, gain_key, top_n, params)
        rows.append((label, round(val, 4) if pd.notna(val) else None))

    # avg_alpha sits directly under avg_gain: the same hops, measured against the
    # benchmark instead of against zero. avg_beta is its companion, not a metric in its
    # own right — a high avg_alpha at beta 1.7 is a different claim than the same alpha
    # at beta 1.0 (see the _market_gain block).
    avg_alpha, avg_beta = _grand_avg_alpha(hop_results, n, params)
    rows.append(("avg_alpha", round(avg_alpha, 4) if pd.notna(avg_alpha) else None))
    if pd.notna(avg_beta):
        rows.append(("avg_beta", round(avg_beta, 3)))

    # Realizable non-overlapping additive chain for the active horizon (= period).
    hold = int(params.get("period", 22))
    ret, annual, ntr = _chain_metrics(hop_results, _GAIN_KEY, n, hold, params)
    rows.append(("chain_ret",    round(ret, 4)    if pd.notna(ret)    else None))
    rows.append(("chain_annual", round(annual, 4) if pd.notna(annual) else None))
    rows.append(("chain_n",      ntr))

    worst, n_loss, sens = _chain_dispersion(hop_results, n, hold, params)
    rows.append(("origin_sens%", round(sens, 1) if pd.notna(sens) else None))
    rows.append(("N_loss", n_loss))
    rows.append(("Worst", round(worst, 4) if pd.notna(worst) else None))

    # Caller-supplied diagnostics that only the strategy pipeline can compute (the Dom_*
    # families' pick_turnover/group_turnover need the dominance table, which never reaches
    # this module). Written last so the block above keeps a fixed shape; aggregate_summary
    # turns any new key into a column on its own.
    for key, value in (extra_summary or {}).items():
        rows.append((key, round(value, 4) if isinstance(value, float) and pd.notna(value)
                     else (None if isinstance(value, float) else value)))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    # avg_beta is deliberately absent: it is a ratio, not a percentage, and green/red
    # gain shading on it would read as "beta 1.7 is good news".
    gain_labels = {r[0] for r in _avg_rows(n)} | {"chain_ret", "chain_annual", "avg_alpha"}
    for i, (k, v) in enumerate(rows, start=1):
        kc = ws.cell(i, 1, k);  kc.font = _BOLD
        vc = ws.cell(i, 2, v)
        if k in gain_labels:
            vc.number_format = _PCT_FMT
            if v is not None:
                vc.fill = _gain_fill(v)


# ---------------------------------------------------------------------------
# HopData sheet — machine-readable per-hop chain inputs
# ---------------------------------------------------------------------------

def _fill_hopdata(ws, hop_results: list[dict], params: dict) -> None:
    """
    Raw per-hop values needed to recompute the realizable chain later over an
    arbitrary daynum window (best_strategy.py clamps to a common floor/cap).

    Stored as plain numbers (not Excel formulas) so they read back reliably.
    Columns: daynum | gain | gspc_rsi | mkt_gain | beta
             (gain = top-N avg for the active period).

    mkt_gain and beta ride along so best_strategy.py can recompute avg_alpha/avg_beta over
    the COMMON span, exactly as it already recomputes avg_gain. Without them the alpha in
    the comparison sheet would be over each run's own span while the avg_gain beside it was
    over the shared one — the same class of mismatch that produced N_loss=42 next to
    chain_n=17. beta in particular cannot be recovered later: it needs the hop's tickers,
    which HopData does not carry.
    """
    n = params.get("focusset_size", 10)
    mkt     = _market_gain(params.get("period", 22))
    beta_df = _beta_frame()
    ws.append(["daynum", "gain", "gspc_rsi", "mkt_gain", "beta"])
    for h in hop_results:
        g    = _hop_avg_topn(h, _GAIN_KEY, n)
        gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
        m    = _hop_market(h, mkt)
        b    = _hop_beta(h, beta_df, n)
        ws.append([
            int(h["daynum"]),
            None if pd.isna(g)    else round(float(g), 6),
            None if pd.isna(gspc) else round(float(gspc), 4),
            None if pd.isna(m)    else round(float(m), 6),
            None if pd.isna(b)    else round(float(b), 4),
        ])


# ---------------------------------------------------------------------------
# Master summary CSV
# ---------------------------------------------------------------------------

def _summary_header() -> list[str]:
    """Current header row of summary.csv, or [] when unreadable."""
    try:
        with SUMMARY_CSV.open(newline="", encoding="utf-8") as f:
            return next(csv.reader(f, delimiter=";"), [])
    except (OSError, StopIteration):
        return []


def _append_summary_csv(strategy_name: str, run_num: int, params: dict,
                        hop_results: list[dict], extra_summary: dict | None = None) -> None:
    daynums = [h["daynum"] for h in hop_results]
    n       = params.get("focusset_size", 10)

    def _fmt(val: float) -> str:
        return "" if pd.isna(val) else f"{val:.4f}".replace(".", ",")

    extra_summary = extra_summary or {}
    param_cols   = list(params.keys())
    avg_labels   = [r[0] for r in _avg_rows(n)] + ["avg_alpha", "avg_beta"]
    chain_labels = _chain_metric_labels()
    extra_cols   = ["origin_sens%", "N_loss", "Worst"] + list(extra_summary)
    all_cols     = (["StrategyName", "Run#", "StartDaynum", "N_hops", "N_hops_active", "EndDaynum"]
                    + param_cols + avg_labels + chain_labels + extra_cols)

    SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)

    # The header is written once, at file creation, and every later run appends under it —
    # so the moment the column set changes (a strategy gains a PARAMS key, a new metric is
    # added) the file silently grows rows wider than their header and every column label
    # after the divergence point is wrong. It had already drifted to 22 header fields
    # against 29-field rows before avg_alpha/avg_beta existed.
    #
    # Archive rather than rewrite: the old rows belong to the old schema and cannot be
    # re-aligned to the new columns, but they are still a record worth keeping. Same
    # reversible convention run_sweep.archive_strategy uses for run*.xlsx.
    write_header = not SUMMARY_CSV.exists()
    if not write_header and _summary_header() != all_cols:
        stamp   = date.today().strftime("%Y%m%d")
        archive = SUMMARY_CSV.parent / "_archive" / f"summary_{stamp}_{int(time.time())}.csv"
        archive.parent.mkdir(parents=True, exist_ok=True)
        SUMMARY_CSV.rename(archive)
        print(f"  summary.csv columns changed — previous file archived to {archive.name}")
        write_header = True

    with SUMMARY_CSV.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        if write_header:
            w.writerow(all_cols)
        avg_vals   = [_fmt(_grand_avg_topn(hop_results, gk, tn, params)) for _, gk, tn in _avg_rows(n)]
        _alpha, _beta = _grand_avg_alpha(hop_results, n, params)
        avg_vals  += [_fmt(_alpha), _fmt(_beta)]
        hold = int(params.get("period", 22))
        ret, annual, ntr = _chain_metrics(hop_results, _GAIN_KEY, n, hold, params)
        chain_vals = [_fmt(ret), _fmt(annual), str(ntr)]
        worst, n_loss, sens = _chain_dispersion(hop_results, n, hold, params)
        extra_vals = [
            "" if pd.isna(sens) else f"{sens:.1f}".replace(".", ","),
            n_loss,
            _fmt(worst),
        ] + [_fmt(v) if isinstance(v, float) else v for v in extra_summary.values()]
        usable   = _usable_daynums(hop_results, _GAIN_KEY, params)
        start_dn = min(usable) if usable else (min(daynums) if daynums else "")
        end_dn   = max(usable) if usable else (max(daynums) if daynums else "")
        w.writerow(
            [strategy_name, run_num, start_dn, len(hop_results),
             _count_active_hops(hop_results, params), end_dn]
            + [_param_cell(params.get(k, "")) for k in param_cols]
            + avg_vals + chain_vals + extra_vals
        )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def save_report(strategy_name: str, params: dict, hop_results: list[dict],
                run_num: int | None = None, extra_summary: dict | None = None) -> None:
    """
    Write one Excel file and append one row to the master summary CSV.

    hop_results items contain:
        daynum, tickers (rank-ordered), gains_20d, gains_50d,
        ref_values (market context at daynum)

    extra_summary: extra {label: value} Summary rows / summary.csv columns the caller
    computed itself, for diagnostics this module cannot derive from hop_results alone
    (the Dom_* families' pick_turnover/group_turnover need the dominance table). Appended
    after the chain-dispersion block; aggregate_summary.py picks up new keys on its own.
    Keep the key set stable per strategy — summary.csv archives and restarts whenever its
    column set changes.
    """
    folder = REPORT_ROOT / strategy_name
    folder.mkdir(parents=True, exist_ok=True)

    if run_num is None:
        run_num = _next_run_num(folder)

    today     = date.today().strftime("%Y%m%d")
    xlsx_path = folder / f"run{run_num}_{today}.xlsx"

    wb    = Workbook()
    ws_op = wb.active
    ws_op.title = "Operational"
    _fill_operational(ws_op, hop_results, params)

    ws_sum = wb.create_sheet("Summary")
    _fill_summary(ws_sum, strategy_name, run_num, params, hop_results, extra_summary)

    ws_hop = wb.create_sheet("HopData")
    _fill_hopdata(ws_hop, hop_results, params)

    wb.save(xlsx_path)
    _append_summary_csv(strategy_name, run_num, params, hop_results, extra_summary)

    print(f"** Report written: {xlsx_path} **")
