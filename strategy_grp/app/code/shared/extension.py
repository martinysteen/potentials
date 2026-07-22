"""
Extension runner: shows the "known future" — for every trading day where the
strategy's forward horizon (future_gain{period}d) is not yet fully realized, it
lists that day's focusset and its partial realised gain so far.

The horizon is read from params["period"] (20 or 50), so the window is roughly the
last `period` trading days. It iterates EVERY day (step=1) from the gain cutoff to
today (newest first), computing gain as (exit_price - entry_price)/entry_price*100
from PotDat. The strategy's own step is NOT used to pick days — it only MARKS the
strategy's investment days (lighter blue in the daynum/date header rows), anchored
to today's daynum.

Output: app/report/<strategy>/extension_YYYYMMDD.xlsx (standalone), or — when a workbook is
passed to run_extension — one sheet appended to that shared workbook (the all-strategies run).

Operational sheet layout (focusset_size = N):
  Row 1      : A1=strategy name      | daynum headers  (blue; step days lighter blue)
  Row 2      : A2="Size/Step/No_RSI/from_rank" | date headers (blue; step days lighter blue)
  Row 3      : A3=param values string  | (blank)                          (blue)
  Rows 4…N+3 : ticker rows
  Row N+4    : per-column "avg_gainXd" labels                       (light green)
  Row N+5    : avg_partial_gain values                              (orange/light-yellow/grey)
  Row N+6/7  : "<info_attribute>_min"/"_max" of that day's picked tickers — only present
               when params["info_attribute"] is set (the DomGICS family); absent otherwise,
               in which case everything below shifts up by 2
  Row N+6/8  : (reserved — 50d labels)   [N+8 when the info rows above are present]
  Row N+7/9  : (reserved — 50d results)  [N+9 when the info rows above are present]
  Row N+8/10+: ref rows (^GSPC_rsi, ^STOXX_rsi, ^HSI_rsi, ^VIX)  (yellow)
  ...        : GICS / Sector2 / Zone occurrence counts

Entry points (from app/code/):
  python extension.py                        # extend ALL strategies (one sheet each)
  <strategy>.build_extension(workbook=None)  # extend one strategy directly (standalone file)
"""

from datetime import date
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.config import REPORT_ROOT
from shared.data_loader import load_longi, load_potdat, load_stamdata, daynum_to_date


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_BOLD         = Font(bold=True)
_SMALL        = Font(size=9)
_PLAIN        = Font()                                   # plain — default size, not bold
_HDR_FILL     = PatternFill("solid", fgColor="BDD7EE")   # blue — headers
_STEP_FILL    = PatternFill("solid", fgColor="DDEBF7")   # lighter blue — strategy-step days
_GRY_FILL     = PatternFill("solid", fgColor="EEEEEE")   # grey — suppressed / n/a
_REF_FILL     = PatternFill("solid", fgColor="FFF2CC")   # yellow — ref rows
_LBL_FILL     = PatternFill("solid", fgColor="E2EFDA")   # light green — label row
_EXT_POS_FILL = PatternFill("solid", fgColor="FFFF99")   # light yellow — positive gain
_EXT_NEG_FILL = PatternFill("solid", fgColor="FFC000")   # orange — negative gain

_ATTR_FILLS: dict[str, PatternFill] = {
    "GICS":    PatternFill("solid", fgColor="D9D2E9"),
    "Sector2": PatternFill("solid", fgColor="FCE5CD"),
    "Zone":    PatternFill("solid", fgColor="D0E0E3"),
}

_PCT_FMT = '+0.00;-0.00;"-"'
_CTR     = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def _find_gain_cutoff(gain_df: pd.DataFrame, period: int, min_valid: int = 10) -> int:
    """Return the newest daynum where future_gain{period}d has sufficient realized data."""
    for col in gain_df.columns:
        if gain_df[col].dropna().size >= min_valid:
            return int(col)
    raise ValueError(f"No valid daynum found in future_gain{period}d.csv")


def _compute_partial_gains(potdat: pd.DataFrame, tickers: list[str],
                            entry_daynum: int, exit_daynum: int) -> dict[str, float]:
    """(price[exit] - price[entry]) / price[entry] * 100. NaN if prices unavailable."""
    ec = str(entry_daynum)
    xc = str(exit_daynum)
    result: dict[str, float] = {}
    for t in tickers:
        try:
            pe = (float(potdat.at[t, ec])
                  if t in potdat.index and ec in potdat.columns and pd.notna(potdat.at[t, ec])
                  else float("nan"))
            px = (float(potdat.at[t, xc])
                  if t in potdat.index and xc in potdat.columns and pd.notna(potdat.at[t, xc])
                  else float("nan"))
            result[t] = ((px - pe) / pe * 100
                         if pd.notna(pe) and pd.notna(px) and pe != 0
                         else float("nan"))
        except (KeyError, ValueError):
            result[t] = float("nan")
    return result


def _hop_avg(gains: dict[str, float]) -> float:
    vals = [v for v in gains.values() if pd.notna(v)]
    return sum(vals) / len(vals) if vals else float("nan")


def _count_attr(hop_results: list[dict], stamdata: pd.DataFrame,
                attr_col: str) -> tuple[list[str], list[dict]]:
    all_vals: set[str] = set()
    hop_counts: list[dict] = []
    for h in hop_results:
        counts: dict[str, int] = {}
        for ticker in h.get("tickers", []):
            if ticker in stamdata.index:
                raw = stamdata.at[ticker, attr_col]
                if pd.notna(raw):
                    v = str(raw).strip()
                    if v:
                        counts[v] = counts.get(v, 0) + 1
                        all_vals.add(v)
        hop_counts.append(counts)
    totals = {v: sum(hc.get(v, 0) for hc in hop_counts) for v in all_vals}
    return sorted(all_vals, key=lambda v: -totals[v]), hop_counts


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _fill_operational(ws, hop_results: list[dict], params: dict,
                      strategy_name: str, mark_step: int = 1) -> None:
    daynums   = [h["daynum"] for h in hop_results]
    n         = params.get("focusset_size",
                            max(len(h.get("tickers", [])) for h in hop_results))
    threshold = params.get("No_go_GSPC_rsi")
    # Step grid is anchored to the newest price daynum (today), not the newest shown
    # entry, so the marking stays aligned even if the latest day had no pick.
    anchor    = (hop_results[0]["daynum"] + hop_results[0]["days_realized"]
                 if hop_results else 0)

    # ---- column-A header stack: A1 strategy, A2 param label, A3 param values ----
    def _pv(key: str) -> str:
        v = params.get(key)
        return "-" if v is None else str(v)

    c = ws.cell(1, 1, strategy_name);                     c.font, c.fill = _BOLD, _HDR_FILL
    c = ws.cell(2, 1, "Size/Step/No_RSI/from_rank");      c.font, c.fill = _PLAIN, _HDR_FILL
    c = ws.cell(3, 1, "/".join([_pv("focusset_size"), _pv("step"),
                                 _pv("No_go_GSPC_rsi"), _pv("from_rank")]))
    c.font, c.fill = _PLAIN, _HDR_FILL

    # ---- daynum (row 1) / date (row 2) headers; strategy-step days in lighter blue ----
    for j, dn in enumerate(daynums, start=2):
        is_step = mark_step > 0 and (anchor - dn) % mark_step == 0
        fill    = _STEP_FILL if is_step else _HDR_FILL
        c = ws.cell(1, j, dn);                 c.font, c.fill, c.alignment = _BOLD, fill, _CTR
        c = ws.cell(2, j, daynum_to_date(dn)); c.font, c.fill, c.alignment = _SMALL, fill, _CTR

    # ---- ticker rows (4 … N+3) ----
    for i in range(n):
        for j, h in enumerate(hop_results, start=2):
            tickers = h.get("tickers", [])
            ws.cell(i + 4, j, tickers[i] if i < len(tickers) else "")

    # ---- per-column "avg_gainXd" label row (N+4) ----
    lbl_row = n + 4
    c = ws.cell(lbl_row, 1, "realized"); c.font, c.fill = _BOLD, _LBL_FILL
    for j, h in enumerate(hop_results, start=2):
        cell = ws.cell(lbl_row, j, f"avg_gain{h['days_realized']}d")
        cell.fill, cell.alignment, cell.font = _LBL_FILL, _CTR, _SMALL

    # ---- avg_partial_gain values row (N+5) — orange/light-yellow/grey ----
    gain_row = n + 5
    c = ws.cell(gain_row, 1, "avg_partial_gain"); c.font = _BOLD
    for j, h in enumerate(hop_results, start=2):
        val    = _hop_avg(h["gains_partial"])
        gspc_p = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
        no_go  = threshold is not None and not pd.isna(gspc_p) and gspc_p < threshold
        cell   = ws.cell(gain_row, j)
        cell.font, cell.number_format, cell.alignment = _BOLD, _PCT_FMT, _CTR
        if pd.notna(val):
            cell.value = None if no_go else round(val, 4)
            cell.fill  = _GRY_FILL if no_go else (_EXT_NEG_FILL if val < 0 else _EXT_POS_FILL)
        else:
            cell.value = None
            cell.fill  = _GRY_FILL

    # ---- info_attribute min/max rows — only for strategies that expose one (DomGICS) ----
    next_row = gain_row + 1
    info_attr = params.get("info_attribute")
    if info_attr:
        info_df = load_longi(f"longi_{info_attr}.csv")
        min_row, max_row = next_row, next_row + 1
        c = ws.cell(min_row, 1, f"{info_attr}_min"); c.font = _BOLD
        c = ws.cell(max_row, 1, f"{info_attr}_max"); c.font = _BOLD
        for j, h in enumerate(hop_results, start=2):
            col = str(h["daynum"])
            vals = [float(info_df.at[t, col]) for t in h.get("tickers", [])
                    if t in info_df.index and col in info_df.columns
                    and pd.notna(info_df.at[t, col])]
            cmin, cmax = ws.cell(min_row, j), ws.cell(max_row, j)
            cmin.alignment = cmax.alignment = _CTR
            if vals:
                cmin.value, cmax.value = round(min(vals), 2), round(max(vals), 2)
        next_row = max_row + 1

    # ---- next 2 rows intentionally blank (reserved for 50d labels / results) ----

    # ---- ref rows ----
    gspc_row  = next_row + 2
    ref_keys = list(hop_results[0].get("ref_values", {}).keys()) if hop_results else []
    for idx, key in enumerate(ref_keys):
        row = gspc_row + idx
        c = ws.cell(row, 1, f"{key}"); c.font, c.fill = _BOLD, _REF_FILL
        for j, h in enumerate(hop_results, start=2):
            val  = h.get("ref_values", {}).get(key, float("nan"))
            cell = ws.cell(row, j)
            cell.fill, cell.alignment = _REF_FILL, _CTR
            if pd.notna(val):
                cell.value         = round(val, 2)
                cell.number_format = "0.0" if key.endswith("_rsi") else "0.00"

    # ---- attribute frequency rows ----
    stamdata = load_stamdata()
    attr_row = gspc_row + len(ref_keys)
    for attr_col in ("GICS", "Sector2", "Zone"):
        fill = _ATTR_FILLS[attr_col]
        sorted_vals, hop_counts = _count_attr(hop_results, stamdata, attr_col)
        for i, val in enumerate(sorted_vals):
            lbl_cell = ws.cell(attr_row + i, 1, f"{attr_col}_{val}")
            lbl_cell.font, lbl_cell.fill = _BOLD, fill
            for j, hc in enumerate(hop_counts, start=2):
                count = hc.get(val, 0)
                cell  = ws.cell(attr_row + i, j)
                cell.fill, cell.alignment = fill, _CTR
                cell.value = count if count > 0 else None
        attr_row += len(sorted_vals)

    # ---- dimensions ----
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(daynums) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = "B4"


def _write_xlsx(strategy_name: str, params: dict, hop_results: list[dict],
                mark_step: int = 1) -> Path:
    folder = REPORT_ROOT / strategy_name
    folder.mkdir(parents=True, exist_ok=True)
    xlsx_path = folder / f"extension_{date.today().strftime('%Y%m%d')}.xlsx"
    wb    = Workbook()
    ws_op = wb.active
    ws_op.title = "Extension"
    _fill_operational(ws_op, hop_results, params, strategy_name, mark_step)
    wb.save(xlsx_path)
    print(f"** Extension written: {xlsx_path} **")
    return xlsx_path


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_extension(
    strategy_name: str,
    params: dict,
    get_focusset_fn: Callable[[int], list[str]],
    get_ref_fn: Callable[[int], dict],
    workbook=None,
) -> Path | str | None:
    """
    Run the extension loop and emit the result.

    workbook is None (default): write a standalone report/<strategy>/extension_YYYYMMDD.xlsx
    and return its Path. workbook given: add the content as a NEW SHEET titled after the
    strategy in that workbook (caller saves it) and return the sheet title — this is how
    the all-strategies run packs one sheet per strategy into a single combined workbook.

    Returns None if the extension window was empty (data fully up to date) or no hop
    produced a focusset.

    get_focusset_fn(daynum) → list[str]  — strategy selector with pre-bound DataFrames
    get_ref_fn(daynum)      → dict       — market context (same as get_reference_values)
    """
    period: int  = int(params.get("period", 20))
    gain_df      = load_longi(f"future_gain{period}d.csv")
    potdat       = load_potdat()
    start_daynum = _find_gain_cutoff(gain_df, period)  # last fully-realized backtest daynum
    exit_daynum  = int(potdat.columns[0])            # most recent price available
    mark_step: int = params.get("step", 1)           # strategy cadence — used only to mark days
    pot_cols     = set(potdat.columns)

    print(f"--- {strategy_name} extension ({period}d horizon) ---")
    print(f"Backtest cutoff: {start_daynum} ({daynum_to_date(start_daynum)})")
    print(f"Exit daynum    : {exit_daynum}  ({daynum_to_date(exit_daynum)})")
    print(f"Entries        : every day (step=1); marking every {mark_step} days")

    # Every trading day in the not-yet-realized window (step=1), newest-first — the
    # extension just shows the "known future" of the recent picks; the strategy's step
    # only marks its investment days. Skip daynums absent from PotDat (guards gaps).
    ext_daynums: list[int] = []
    d = start_daynum + 1
    while d <= exit_daynum:
        if str(d) in pot_cols:
            ext_daynums.append(d)
        d += 1
    ext_daynums.reverse()

    if not ext_daynums:
        print("Extension period is empty — data fully up to date.\n")
        return None
    print(f"Days           : {len(ext_daynums)}"
          f" (daynums {ext_daynums[-1]}→{ext_daynums[0]})\n")

    hop_results: list[dict] = []
    for daynum in ext_daynums:
        tickers = get_focusset_fn(daynum)
        if not tickers:
            continue
        days          = exit_daynum - daynum
        gains_partial = _compute_partial_gains(potdat, tickers, daynum, exit_daynum)
        avg           = _hop_avg(gains_partial)
        print(f"  entry {daynum} ({daynum_to_date(daynum)})"
              f"  tickers={tickers}"
              f"  gain_{days}d avg={avg:+.2f}%")
        hop_results.append({
            "daynum":          daynum,
            "tickers":         tickers,
            "gains_partial":   gains_partial,
            "days_realized":   days,
            "ref_values": get_ref_fn(daynum),
        })

    print()
    if not hop_results:
        print(f"No valid hops — strategy selected nothing in the extension period.\n")
        return None

    if workbook is not None:
        ws = workbook.create_sheet(title=strategy_name[:31])   # one sheet per strategy
        _fill_operational(ws, hop_results, params, strategy_name, mark_step)
        print(f"Done: {len(hop_results)} extension days (sheet '{ws.title}').\n")
        return ws.title

    xlsx_path = _write_xlsx(strategy_name, params, hop_results, mark_step=mark_step)
    print(f"Done: {len(hop_results)} extension days.\n")
    return xlsx_path
