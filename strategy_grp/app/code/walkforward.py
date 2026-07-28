"""
Walk-forward evaluation — how much of the sweep's reported edge survives OUT OF SAMPLE.

Every number in best_strategy.xlsx is in-sample: the sweep scores each parameter-set
over the whole history and reports the winner's score over that same history. That
number is biased upward by however many parameter-sets were tried, and it cannot tell
you whether the winner would have been picked in advance. This harness answers that
question directly and changes no selection logic anywhere: it only re-scores.

The procedure (per strategy, repeated over rolling folds)
--------------------------------------------------------
    train = [oldest daynum .. T - period]     pick the best parameter-set here
    test  = [T + 1 .. T + test_len]           score THAT set here, never re-picking

`T - period` is an embargo, not an off-by-one: a hop entered at daynum d does not
realize until d + period, so training on hops up to T would let a hop that closes
inside the test window vote on the parameter choice. Without the embargo the whole
exercise leaks and reports a flattering number.

T then advances by test_len, so the test windows tile the recent history without
overlapping and every test lot is scored by parameters chosen strictly before it.

What it reports
---------------
Per fold and pooled across folds, for the parameter-set the training window selected:

  is_annual / oos_annual   chain_annual (shared.chain, the sweep's own ranking metric)
                           in and out of sample. The IS->OOS drop IS the overfit.
  oos_avg_gain             mean lot gain out of sample — steadier than chain_annual on
                           a short fold, where the annualization divides by a small span.
  oos_alpha                mean (lot gain - that daynum's cross-sectional market mean).
                           Absolute return cannot separate a good strategy from a good
                           market; this can.
  oos_mean_all             mean OOS score across ALL candidates in the grid — the
                           zero-skill baseline. If the selected set does not beat this,
                           the sweep's selection carries no information out of sample,
                           whatever its in-sample margin looked like.
  oos_oracle               best OOS score any candidate achieved — the ceiling, for scale.

`selection_skill` = oos(selected) - oos_mean_all is the headline: positive means
choosing parameters on history genuinely helps; ~0 means the sweep is fitting noise.

Usage (from app/code/)
----------------------
    python walkforward.py                 # grid = exactly what sweep_config.py runs
    python walkforward.py --wide          # add numeric axes (below) for a stronger test
    python walkforward.py --test-len 126  # 6-month folds instead of 3-month
    python walkforward.py --dry-run       # show the fold layout and grid size only

Only DomGICS_* strategies are covered: the harness rebuilds picks itself (so it can
re-score a window without re-running a whole report), which needs the dominance
pipeline. Other strategies are skipped with a note.

Reads only; the sole output is app/report/walkforward_<date>.xlsx. No run*.xlsx,
aggregated_summary.xlsx, summary.csv or best_strategy*.xlsx is touched.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sweep_config
from run_sweep import build_plan, discover_strategies
from shared.chain import realizable_chain
from shared.config import REPORT_ROOT
from shared.data_loader import daynum_to_date, load_longi
from shared.dominance import dominance_tables, select_focusset

# ---------------------------------------------------------------------------
# Harness configuration
# ---------------------------------------------------------------------------

# Fold geometry, in trading daynums (~252/year). Defaults are a compromise forced by a
# short history: the DomGICS data spans ~2.5 years, i.e. only ~30 INDEPENDENT 20d lots
# in total. 3-month folds after 15 months of training give ~5 folds of ~3 independent
# lots each — noisy per fold, which is exactly why the pooled figures are the ones to
# read. Widen test-len if the history ever grows.
DEFAULT_TEST_LEN: int = 63     # ~3 months
DEFAULT_MIN_TRAIN: int = 315   # ~15 months

# --wide only. Extra grid axes so the selection test has a real parameter space to chew
# on: with sweep_config's current 2 candidates, "did selection help?" has almost no
# statistical power. These are numeric axes only — priority_attribute stays whatever
# sweep_config sweeps, because its direction must come from
# run_config.PRIORITY_ATTRIBUTE_DICTIONARY and is not the harness's to invent.
WIDE_AXES: dict[str, list] = {
    "from_rank":                  [1, -1],
    "No_go_GSPC_rsi":             [0, 40, 45],
    "dominance_threshold_decile": [0.05, 0.10, 0.20],
    "tickers_per_gics":           [2, 3, 5],
}

# Params that change WHICH TICKERS a hop picks (or what their gain is). Two parameter-sets
# agreeing on all of these produce an identical hop series, so it is built once and shared.
# No_go_GSPC_rsi is deliberately absent: it gates hops at scoring time inside
# shared.chain, it does not change the picks — which makes it a free grid axis.
_PICK_KEYS: tuple[str, ...] = (
    "dominance_threshold_decile", "dom_count_threshold", "persistence_frac",
    "dominance_attribute", "dominance_attribute_direction",
    "tickers_per_gics", "focusset_size", "from_rank",
    "priority_attribute", "priority_attribute_direction", "step", "period",
)

# The grid axes worth naming in the report — the rest are constant across a run and
# would just be noise in a column.
_LABEL_KEYS: tuple[str, ...] = (
    "priority_attribute", "from_rank", "No_go_GSPC_rsi",
    "dominance_threshold_decile", "tickers_per_gics", "focusset_size",
)


# ---------------------------------------------------------------------------
# Hop series
# ---------------------------------------------------------------------------

_dom_cache: dict[tuple, tuple] = {}
_series_cache: dict[tuple, list[tuple[int, float, float]]] = {}


def _dom_table(params: dict, dom_col: str) -> pd.DataFrame:
    """One tier's GICS x daynum dominance table, memoized on the Step-1 params.

    dominance_tables() builds all three tiers in one pass and is the expensive call in
    the whole harness, so a grid that varies only Step-2 params pays for it once.
    """
    key = (params["dominance_threshold_decile"], params["dom_count_threshold"],
           params["persistence_frac"], params.get("dominance_attribute", "rank"),
           params.get("dominance_attribute_direction", True))
    if key not in _dom_cache:
        _dom_cache[key] = dominance_tables(*key)
    tables, _cutoffs = _dom_cache[key]
    return tables[dom_col]


def hop_series(params: dict, dom_col: str) -> list[tuple[int, float, float]]:
    """(daynum, hop gain %, GSPC RSI) for every hop over the full history.

    The hop gain is the mean realized future_gain{period}d across the focusset, NaN
    dropped — the same quantity shared.report feeds the chain, so a walk-forward score
    and a run*.xlsx score are measuring the same thing.
    """
    key = (dom_col,) + tuple(params[k] for k in _PICK_KEYS)
    if key in _series_cache:
        return _series_cache[key]

    period  = params["period"]
    step    = params["step"]
    gain_df = load_longi(f"future_gain{period}d.csv")
    rsi_df  = load_longi("longi_rsi.csv")
    dom     = _dom_table(params, dom_col)

    start = _first_realized_daynum(gain_df)
    stop  = int(gain_df.columns[-1])

    rows: list[tuple[int, float, float]] = []
    daynum = start
    while daynum >= stop:
        col = str(daynum)
        tickers = select_focusset(daynum, dom, params["tickers_per_gics"],
                                  params["focusset_size"], params.get("from_rank", 1),
                                  params.get("priority_attribute", "rank"),
                                  params.get("priority_attribute_direction", True))
        vals = [gain_df.at[t, col] for t in tickers
                if t in gain_df.index and col in gain_df.columns
                and pd.notna(gain_df.at[t, col])]
        gain = sum(vals) / len(vals) if vals else float("nan")
        gspc = (float(rsi_df.at["^GSPC", col])
                if "^GSPC" in rsi_df.index and col in rsi_df.columns
                and pd.notna(rsi_df.at["^GSPC", col]) else float("nan"))
        rows.append((daynum, gain, gspc))
        daynum -= step

    _series_cache[key] = rows
    return rows


def _first_realized_daynum(gain_df: pd.DataFrame, min_valid: int = 10) -> int:
    """Newest daynum whose forward gains have actually realized (mirrors
    shared.dominance._find_start_daynum — kept local so the harness never silently
    scores an unrealized tail)."""
    for col in gain_df.columns:
        if gain_df[col].dropna().size >= min_valid:
            return int(col)
    raise ValueError("No valid starting daynum found in future_gain data")


def market_series(period: int) -> pd.Series:
    """Cross-sectional mean future_gain{period}d per daynum — the passive benchmark
    every hop's alpha is measured against."""
    return load_longi(f"future_gain{period}d.csv").mean(axis=0)


# ---------------------------------------------------------------------------
# Scoring one window
# ---------------------------------------------------------------------------

def window_metrics(rows: list[tuple[int, float, float]], params: dict,
                   mkt: pd.Series, floor: int, cap: int) -> dict:
    """Score one parameter-set over one [floor, cap] daynum window (both inclusive).

    chain_annual comes from shared.chain.realizable_chain — the sweep's own ranking
    metric, phase-averaged, so training selection here reproduces what best_strategy.py
    would have chosen on that window. avg_gain/alpha are per-lot and far steadier on a
    short fold, where annualizing over a small span amplifies a single lot.
    """
    period = params["period"]
    nogo   = params.get("No_go_GSPC_rsi") or None   # 0 means "gate off", same as None
    _ret, annual, n = realizable_chain(rows, period, nogo, floor, cap, phase_average=True)

    lots: list[tuple[float, float]] = []
    for daynum, gain, gspc in rows:
        if daynum < floor or daynum > cap or pd.isna(gain):
            continue
        if nogo is not None and pd.notna(gspc) and gspc < nogo:
            continue
        m = mkt.get(str(daynum), float("nan"))
        lots.append((gain, float(m) if pd.notna(m) else float("nan")))

    gains  = [g for g, _m in lots]
    alphas = [g - m for g, m in lots if pd.notna(m)]
    return {
        "chain_annual": annual,
        "chain_n":      n,
        "avg_gain":     sum(gains) / len(gains) if gains else float("nan"),
        "alpha":        sum(alphas) / len(alphas) if alphas else float("nan"),
        "n_hops":       len(lots),
        "lots":         lots,
    }


def build_folds(dn_min: int, dn_max: int, min_train: int, test_len: int,
                period: int) -> list[tuple[int, int, int, int]]:
    """(train_floor, train_cap, test_floor, test_cap) per fold, oldest first.

    train_cap = T - period is the embargo described in the module docstring: no hop that
    is still open when the test window starts may influence the parameter choice.

    A trailing remainder shorter than `period` is dropped rather than reported: it cannot
    contain even one complete holding window, so its chain_annual would annualize a
    fraction of a lot into a headline number.
    """
    out: list[tuple[int, int, int, int]] = []
    t = dn_min + min_train
    while t < dn_max:
        test_hi = min(t + test_len, dn_max)
        if test_hi - t >= period:
            out.append((dn_min, t - period, t + 1, test_hi))
        t += test_len
    return out


def label(params: dict) -> str:
    """Compact one-line identity of a parameter-set, showing only the swept axes."""
    return " ".join(f"{k}={params[k]}" for k in _LABEL_KEYS if k in params)


# ---------------------------------------------------------------------------
# Walk-forward over one strategy
# ---------------------------------------------------------------------------

def walk_strategy(name: str, dom_col: str, grid: list[dict],
                  min_train: int, test_len: int, min_lots: int) -> dict:
    """Run every fold for one strategy. Returns a result bundle for the reporters.

    `min_lots` is a floor on the TRAINING chain's lot count, not a cosmetic filter.
    chain_annual divides an additive sum by the chain's own span, so a parameter-set
    sparse enough to realize a single lucky lot annualizes it over ~one holding window
    and posts a headline in the hundreds — it then wins the selection on noise. Nothing
    in best_strategy.py guards against this today (see the run notes); here it would
    make the walk-forward a test of that artifact rather than of the strategy.
    """
    period = grid[0]["period"]
    mkt    = market_series(period)

    series = {i: hop_series(p, dom_col) for i, p in enumerate(grid)}
    daynums = [dn for dn, _g, _r in series[0]]
    dn_min, dn_max = min(daynums), max(daynums)
    folds = build_folds(dn_min, dn_max, min_train, test_len, period)

    fold_rows: list[dict] = []
    cand_rows: list[dict] = []
    pooled_sel: list[tuple[float, float]] = []
    pooled_all: list[tuple[float, float]] = []

    for fi, (tr_lo, tr_hi, te_lo, te_hi) in enumerate(folds, start=1):
        scored = []
        for i, params in enumerate(grid):
            tr = window_metrics(series[i], params, mkt, tr_lo, tr_hi)
            te = window_metrics(series[i], params, mkt, te_lo, te_hi)
            scored.append((i, params, tr, te))
            cand_rows.append({
                "fold": fi, "params": label(params),
                "is_annual": tr["chain_annual"], "is_avg_gain": tr["avg_gain"],
                "oos_annual": te["chain_annual"], "oos_avg_gain": te["avg_gain"],
                "oos_alpha": te["alpha"], "oos_n": te["n_hops"],
                "selected": "",
            })
            pooled_all.extend(te["lots"])

        # Selection: the sweep's own metric, on the training window only.
        usable = [s for s in scored
                  if not pd.isna(s[2]["chain_annual"]) and s[2]["chain_n"] >= min_lots]
        if not usable:
            continue
        i_sel, p_sel, tr_sel, te_sel = max(usable, key=lambda s: s[2]["chain_annual"])
        for r in cand_rows[-len(grid):]:
            if r["params"] == label(p_sel):
                r["selected"] = "<<"
        pooled_sel.extend(te_sel["lots"])

        oos_annuals = [s[3]["chain_annual"] for s in scored
                       if not pd.isna(s[3]["chain_annual"])]
        oos_avgs    = [s[3]["avg_gain"] for s in scored if not pd.isna(s[3]["avg_gain"])]
        oos_alphas  = [s[3]["alpha"] for s in scored if not pd.isna(s[3]["alpha"])]

        fold_rows.append({
            "fold": fi,
            "train_from": tr_lo, "train_to": tr_hi,
            "test_from": te_lo, "test_to": te_hi,
            "test_dates": f"{daynum_to_date(te_lo)} .. {daynum_to_date(te_hi)}",
            "selected": label(p_sel),
            "is_annual": tr_sel["chain_annual"],
            "is_n": tr_sel["chain_n"],
            # The honest overfit measure. chain_annual divides by the window's own span,
            # so a 3-month test fold annualizes one good quarter into the hundreds and
            # its IS->OOS "gap" is mostly that artifact. Per-lot gain/alpha are span-free
            # and directly comparable between a 15-month train and a 3-month test.
            "is_avg_gain": tr_sel["avg_gain"],
            "is_alpha": tr_sel["alpha"],
            "oos_annual": te_sel["chain_annual"],
            "oos_avg_gain": te_sel["avg_gain"],
            "oos_alpha": te_sel["alpha"],
            "oos_n": te_sel["n_hops"],
            "oos_mean_all": sum(oos_avgs) / len(oos_avgs) if oos_avgs else float("nan"),
            "oos_alpha_mean_all": (sum(oos_alphas) / len(oos_alphas)
                                   if oos_alphas else float("nan")),
            "oos_oracle": max(oos_avgs) if oos_avgs else float("nan"),
            "oos_annual_mean_all": (sum(oos_annuals) / len(oos_annuals)
                                    if oos_annuals else float("nan")),
        })

    return {
        "strategy": name, "dom_col": dom_col, "period": period,
        "grid_size": len(grid), "folds": fold_rows, "candidates": cand_rows,
        "pooled_selected": pooled_sel, "pooled_all": pooled_all,
        "span": (dn_min, dn_max),
    }


def _pooled(lots: list[tuple[float, float]]) -> tuple[float, float, int]:
    """(mean gain, mean alpha, n) over pooled lots — NaN market values dropped for alpha."""
    if not lots:
        return float("nan"), float("nan"), 0
    gains  = [g for g, _m in lots]
    alphas = [g - m for g, m in lots if pd.notna(m)]
    return (sum(gains) / len(gains),
            sum(alphas) / len(alphas) if alphas else float("nan"),
            len(gains))


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

_BOLD = Font(bold=True)
_HEAD = PatternFill("solid", fgColor="DDEBF7")
_SEL  = PatternFill("solid", fgColor="FFF2CC")
_CTR  = Alignment(horizontal="center")


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font, c.fill, c.alignment = _BOLD, _HEAD, _CTR
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, None if isinstance(v, float) and pd.isna(v) else v)
            if j > 1:
                cell.alignment = _CTR
    for j, h in enumerate(headers, start=1):
        width = max([len(str(h))] + [len(str(r[j - 1])) for r in rows]) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(width, 44)
    ws.freeze_panes = "B2"


def write_report(results: list[dict], min_train: int, test_len: int,
                 suffix: str = "") -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    summary: list[list] = []
    for r in results:
        g_sel, a_sel, n_sel = _pooled(r["pooled_selected"])
        g_all, a_all, _n    = _pooled(r["pooled_all"])
        g_is  = _mean_of(r["folds"], "is_avg_gain")
        a_is  = _mean_of(r["folds"], "is_alpha")
        summary.append([
            r["strategy"], r["grid_size"], len(r["folds"]), n_sel,
            _r2(g_is), _r2(g_sel), _r2(g_sel - g_is),
            _r2(a_is), _r2(a_sel), _r2(a_sel - a_is),
            _r2(g_all), _r2(a_all),
            _r2(g_sel - g_all), _r2(a_sel - a_all),
            _r2(_mean_of(r["folds"], "is_annual")),
            _r2(_mean_of(r["folds"], "oos_annual")),
        ])
    _sheet(wb, "Summary",
           ["strategy", "grid", "folds", "oos_lots",
            "is_avg_gain", "oos_avg_gain", "gain_gap",
            "is_alpha", "oos_alpha", "alpha_gap",
            "zeroskill_avg_gain", "zeroskill_alpha",
            "selection_skill_gain", "selection_skill_alpha",
            "is_annual", "oos_annual"],
           summary)

    fold_rows: list[list] = []
    for r in results:
        for f in r["folds"]:
            fold_rows.append([
                r["strategy"], f["fold"], f["test_dates"],
                f"{f['train_from']}-{f['train_to']}", f"{f['test_from']}-{f['test_to']}",
                f["selected"], f["is_n"], f["oos_n"],
                _r2(f["is_annual"]), _r2(f["oos_annual"]), _r2(f["oos_annual_mean_all"]),
                _r2(f["oos_avg_gain"]), _r2(f["oos_mean_all"]), _r2(f["oos_oracle"]),
                _r2(f["oos_alpha"]), _r2(f["oos_alpha_mean_all"]),
            ])
    _sheet(wb, "Folds",
           ["strategy", "fold", "test period", "train dn", "test dn",
            "selected params", "is_n", "oos_n",
            "is_annual", "oos_annual", "oos_annual_mean_all",
            "oos_avg_gain", "oos_mean_all", "oos_oracle",
            "oos_alpha", "oos_alpha_mean_all"],
           fold_rows)

    cand_rows: list[list] = []
    for r in results:
        for c in r["candidates"]:
            cand_rows.append([
                r["strategy"], c["fold"], c["selected"], c["params"],
                _r2(c["is_annual"]), _r2(c["is_avg_gain"]),
                _r2(c["oos_annual"]), _r2(c["oos_avg_gain"]), _r2(c["oos_alpha"]),
                c["oos_n"],
            ])
    _sheet(wb, "Candidates",
           ["strategy", "fold", "sel", "params",
            "is_annual", "is_avg_gain",
            "oos_annual", "oos_avg_gain", "oos_alpha", "oos_n"],
           cand_rows)

    ws = wb.create_sheet("About", 0)
    for i, line in enumerate([
        ("Walk-forward evaluation", ""),
        ("", ""),
        ("min_train (daynums)", min_train),
        ("test_len (daynums)", test_len),
        ("min_lots", "training chains realizing fewer lots cannot be selected "
                     "(their chain_annual is one lot annualized)"),
        ("embargo", "train window ends period daynums before the test window opens"),
        ("selection metric", "chain_annual on the training window (the sweep's own metric)"),
        ("", ""),
        ("is_avg_gain / is_alpha", "the selected set scored on its own training window "
                                   "— i.e. what the sweep would have reported"),
        ("oos_avg_gain / oos_alpha", "the same set on the untouched test window"),
        ("gain_gap / alpha_gap", "oos - is. How much of the reported edge evaporates. "
                                 "THE overfit measure — read this, not annual_gap."),
        ("alpha", "lot gain - that daynum's cross-sectional market mean. Absolute "
                  "return cannot tell a good strategy from a good market."),
        ("is_annual / oos_annual", "chain_annual for continuity with best_strategy.xlsx. "
                                   "Unstable on a short fold: it divides by the window's "
                                   "own span, so one good quarter annualizes into the "
                                   "hundreds. Rank on the per-lot columns instead."),
        ("zeroskill_*", "same, averaged over EVERY candidate: what no selection skill gets"),
        ("selection_skill_*", "selected - zeroskill. ~0 means the sweep is fitting noise."),
        ("oos_oracle", "best any candidate did out of sample; the ceiling, for scale"),
    ], start=1):
        ws.cell(i, 1, line[0]).font = _BOLD
        ws.cell(i, 2, line[1])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78

    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    # The --wide grid gets its own file: it answers a different question (does selection
    # help at all?) than the sweep_config grid (how does what I actually run hold up?),
    # and one clobbering the other on the same day loses the comparison.
    path = REPORT_ROOT / f"walkforward{suffix}_{date.today():%Y%m%d}.xlsx"
    wb.save(path)
    return path


def _r2(v):
    return round(v, 2) if isinstance(v, (int, float)) and not pd.isna(v) else v


def _mean_of(folds: list[dict], key: str) -> float:
    """Mean of one per-fold metric, NaN folds skipped."""
    vals = [f[key] for f in folds if not pd.isna(f[key])]
    return sum(vals) / len(vals) if vals else float("nan")


def print_console(results: list[dict]) -> None:
    for r in results:
        g_sel, a_sel, n_sel = _pooled(r["pooled_selected"])
        g_all, a_all, _n    = _pooled(r["pooled_all"])
        print(f"\n=== {r['strategy']}  ({r['dom_col']}, {r['period']}d)  "
              f"grid={r['grid_size']}  folds={len(r['folds'])} ===")
        print(f"  {'fold':<5}{'test period':<26}{'is_ann':>8}{'is_n':>6}{'oos_ann':>9}"
              f"{'oos_gain':>10}{'zero':>8}{'oracle':>8}{'alpha':>8}  selected")
        for f in r["folds"]:
            print(f"  {f['fold']:<5}{f['test_dates']:<26}"
                  f"{_f(f['is_annual']):>8}{f['is_n']:>6}{_f(f['oos_annual']):>9}"
                  f"{_f(f['oos_avg_gain']):>10}{_f(f['oos_mean_all']):>8}"
                  f"{_f(f['oos_oracle']):>8}{_f(f['oos_alpha']):>8}  {f['selected']}")
        g_is = _mean_of(r["folds"], "is_avg_gain")
        a_is = _mean_of(r["folds"], "is_alpha")
        print(f"  in-sample (what the sweep would report)  "
              f"avg_gain {g_is:6.2f}  alpha {a_is:6.2f}")
        print(f"  out-of-sample ({n_sel} lots)              "
              f"avg_gain {g_sel:6.2f}  alpha {a_sel:6.2f}")
        print(f"  OVERFIT (oos - is)                       "
              f"avg_gain {g_sel - g_is:+6.2f}  alpha {a_sel - a_is:+6.2f}")
        print(f"  zero-skill baseline                      "
              f"avg_gain {g_all:6.2f}  alpha {a_all:6.2f}")
        print(f"  SELECTION SKILL (oos - zero-skill)       "
              f"avg_gain {g_sel - g_all:+6.2f}  alpha {a_sel - a_all:+6.2f}")


def _f(v):
    return "-" if v is None or (isinstance(v, float) and pd.isna(v)) else f"{v:.2f}"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--test-len", type=int, default=DEFAULT_TEST_LEN,
                    help=f"test-window length in daynums (default {DEFAULT_TEST_LEN})")
    ap.add_argument("--min-train", type=int, default=DEFAULT_MIN_TRAIN,
                    help=f"daynums of history before the first fold "
                         f"(default {DEFAULT_MIN_TRAIN})")
    ap.add_argument("--min-lots", type=int, default=4,
                    help="minimum lots a candidate's TRAINING chain must realize to be "
                         "selectable (default 4) — blocks sparse configs whose "
                         "chain_annual is one lucky lot annualized")
    ap.add_argument("--wide", action="store_true",
                    help="add WIDE_AXES to the grid for a stronger selection test")
    ap.add_argument("--dry-run", action="store_true",
                    help="print the fold layout and grid size, then stop")
    args = ap.parse_args()

    modules = discover_strategies()
    plan = build_plan(modules)

    results: list[dict] = []
    for name, grid in plan.items():
        dom_col = getattr(modules[name].main, "dom_col", None)
        if dom_col is None:
            print(f"  {name}: not a DomGICS_* strategy — skipped "
                  f"(walkforward rebuilds picks via the dominance pipeline)")
            continue

        if args.wide:
            from itertools import product
            axes = [[(k, v) for v in vals] for k, vals in WIDE_AXES.items()
                    if k in grid[0]]
            wide: list[dict] = []
            for base in grid:
                for combo in product(*axes):
                    p = dict(base)
                    p.update(dict(combo))
                    if p not in wide:
                        wide.append(p)
            grid = wide

        print(f"\n{name}: {len(grid)} candidate parameter-set(s)")
        if args.dry_run:
            period = grid[0]["period"]
            rows = hop_series(grid[0], dom_col)
            dns = [d for d, _g, _r in rows]
            folds = build_folds(min(dns), max(dns), args.min_train, args.test_len, period)
            print(f"  history {min(dns)}..{max(dns)} "
                  f"({daynum_to_date(min(dns))} .. {daynum_to_date(max(dns))}), "
                  f"{len(rows)} hops")
            for i, (a, b, c, d) in enumerate(folds, start=1):
                print(f"  fold {i}: train {a}..{b}  test {c}..{d}  "
                      f"({daynum_to_date(c)} .. {daynum_to_date(d)})")
            continue

        results.append(walk_strategy(name, dom_col, grid, args.min_train,
                                     args.test_len, args.min_lots))

    if args.dry_run or not results:
        return

    print_console(results)
    path = write_report(results, args.min_train, args.test_len,
                        "_wide" if args.wide else "")
    print(f"\nWrote {path}")


if __name__ == "__main__":
    main()
