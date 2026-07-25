"""
Compare strategies side by side — one column per strategy, metrics down the rows.

Each strategy's column is its best run by chain_annual (phase-averaged and re-clamped
to the span every strategy shares), with chain_ret as tiebreaker. Runs are grouped by
forward horizon (the `period` param): each horizon gets its OWN comparison (own common
span — chains of different hold lengths are never mixed in one table). The smallest
horizon is the primary sheet; further horizons (e.g. the 50d fallback) get one sheet each.

Returns are additive (sum of lot gains, no reinvestment); chain_annual is that sum
divided by the span in years — a simple average annual gain, not a compound CAGR.

Output:
  app/report/best_strategy.xlsx  — transposed, strategies as columns, in the fixed order
  given by STRATEGY_ORDER (sweep_config.py); unlisted strategies sort last

Usage (from app/code/):
  python best_strategy.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.chain import (realizable_chain, laddered_portfolio,
                          chain_lot_stats, laddered_lot_stats, chain_inv_pct,
                          chain_origin_range)
from shared.config import REPORT_ROOT
from sweep_config import STRATEGY_ORDER

# Decision metric is chain_annual — the additive average annual gain (phase-averaged,
# common-span); chain_ret breaks ties. Per-strategy columns are defined near main().
# Output is the combined workbook built by extension.run() — best_strategy_<date>.xlsx
# (comparison sheet first, then one extension sheet per strategy). This module no longer
# writes a file of its own.

# ===========================================================================
# Human-facing text — EDIT FREELY. Everything a reader sees as prose in the
# workbook is gathered here: the two title rows, the right-table header, and
# every row's plain-language Comment. Nothing here affects any calculation —
# only the words on the page.
#
# Comments are split by table: _COMMENTS_CHAIN for the LEFT (chained) column B,
# _COMMENTS_LADDER for the RIGHT (overlap) column I. A metric with no entry shows a
# blank comment, so add/remove keys freely — the key must match the row's name (the
# metric/param identifier in column A / H). Shared rows (period, focusset_size, …)
# live only in _COMMENTS_CHAIN and are reused for column I, so edit them once.
# ===========================================================================
_LADDER_TITLE = "Overlap investment"              # header of the right-hand table

_COMMENTS_CHAIN = {     # left table (column B) — plus the shared rows reused on the right
    "period":        "Trading days (per investment)",
    "StartDaynum":   "First usable daynum (oldest; later than series start if ML warm-up applies)",
    "EndDaynum":     "Last usable daynum (newest)",
    "chain_annual":  "Avg annual gain% (additive: sum of lot gains / years, no compounding)",
    "chain_ret":     "Additive return of full chain (sum of lot gains, no reinvestment)",
    "chain_n":       "Number of lots in full chain",
    "origin_sens%":  "Min and max of chain_annual across 4 start origins",
    "avg_gain":      "Average gain% per chain lot",
    "Worst":         "Worst chain lot (gain%)",
    "N_loss":        "Most negative lots in any one realized chain (of chain_n)",
    "chain_inv%":    "Share of active span invested (%) — idle when No_go gating / too few survivors block a reinvest",
    "focusset_size": "Number of stocks in each investment lot",
    "from_rank":     "Which end of the ranking to draw from: 1=best n, -1=worst n",
}
_COMMENTS_LADDER = {    # right table (column I) — the ladder_* rows only
    "ladder_annual": "Avg annual gain% (additive), overlap always-invested style (diagnostic)",
    "ladder_ret":    "Additive return of overlap portfolio (sum of lot gains)",
    "ladder_n":      "Total overlap investments (period/step sleeves, continuous)",
    "ladder_inv%":   "Share of tranches invested vs cash (%)",
    "ladder_avg_gain": "Average gain% per overlap investment",
    "ladder_worst":  "Worst overlap investment (gain%)",
    "ladder_n_loss": "Negative investments in overlap (of ladder_n)",
}

# The report is two side-by-side tables sharing the same strategy columns: a left
# "chained" table and a right "Overlap investment" table. _CHAINED_KEYS is the row
# order of the LEFT table (these first, then any remaining param cols appended). The
# right table mirrors it row-for-row, swapping chained metrics for their overlap
# twins via _CHAIN_TO_LADDER.
#
# StrategyName is absent: the strategy name is the column header, so a row would just
# repeat it. Floor/cap are absent: they are already stated in the title rows.
_CHAINED_KEYS = [
    "Run#", "period", "StartDaynum", "EndDaynum",
    "chain_annual", "origin_sens%", "chain_ret", "chain_n",
    "avg_gain", "Worst", "N_loss",
    "chain_inv%",
    "focusset_size", "step", "No_go_GSPC_rsi", "from_rank",
    "source_file",
]
_GAIN_COLS = {"avg_gain", "Worst", "ladder_avg_gain", "ladder_worst"}

# Never a row of their own: StrategyName (it's the header); floor/cap (they're in the
# title); N_hops/N_hops_active (redundant with chain_n on the left and ladder_n on the
# right — dropped from the comparison); the twinned ladder_* metrics (rendered in the
# right table by mapping their chained twin, not as standalone rows — ladder_inv% is the
# right twin of chain_inv%).
_DROP_ROWS = {"StrategyName", "chain_floor", "chain_cap",
              "N_hops", "N_hops_active",
              "ladder_annual", "ladder_ret", "ladder_n", "ladder_inv%",
              "ladder_avg_gain", "ladder_worst", "ladder_n_loss"}

# Chained metric -> its counterpart in the right-hand overlap table. A key not listed
# is shared verbatim (identical value, shown in both tables for every strategy). None
# drops the row from the overlap side (only source_file does). avg_gain/Worst/N_loss
# are per-investment dispersion and so are NOT shared — each side gets its own set,
# computed over its own (chain vs overlap) investment population.
_CHAIN_TO_LADDER = {
    "chain_annual":  "ladder_annual",
    "chain_ret":     "ladder_ret",
    "chain_n":       "ladder_n",
    "avg_gain":      "ladder_avg_gain",
    "Worst":         "ladder_worst",
    "N_loss":        "ladder_n_loss",
    "chain_inv%":    "ladder_inv%",
    # Chain-only, not because an overlap can't be origin-sensitive but because it diversifies it
    # away: the overlap holds all n=hold/step entry origins at once, so its blend is their exact
    # average (blend = sum(hops)/n, independent of the start) -> origin_sens -> 0 for step<<period,
    # returning to the full chain swing only at the n=1 edge (step=period). The serial chain,
    # picking one origin, is the one that actually feels it.
    "origin_sens%":  None,
    "source_file":   None,
}
def _is_gain_col(col: str) -> bool:
    return (col in _GAIN_COLS or "gain" in col
            or col.startswith(("chain_ret", "chain_annual", "ladder_ret", "ladder_annual")))

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_BOLD       = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=13)
_SMALL      = Font(size=9)
_HDR_FILL   = PatternFill("solid", fgColor="BDD7EE")   # blue header row
_SECT_FILL  = PatternFill("solid", fgColor="D6DCE4")   # grey section header
_GRN_FILL   = PatternFill("solid", fgColor="C6EFCE")   # green
_RED_FILL   = PatternFill("solid", fgColor="FFC7CE")   # red
_BEST_FILL  = PatternFill("solid", fgColor="FFE599")   # amber for "Best overall" row
_PARAM_FILL = PatternFill("solid", fgColor="FFFF99")   # yellow for simulation parameter headers
_PCT_FMT    = '+0.00;-0.00;"-"'
_CTR        = Alignment(horizontal="center")

_PARAM_COLS = {"focusset_size", "step", "period", "No_go_GSPC_rsi", "from_rank",
               "corner_bins", "vola_keep_frac", "q10_20_min", "q20_50_min",
               "dominance_threshold", "dom_count_threshold", "persistence_frac", "tickers_per_gics",
               "dominance_attribute", "dominance_attribute_direction",
               "priority_attribute", "priority_attribute_direction", "informational_attributes"}


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def load_all_runs() -> pd.DataFrame:
    """Read every aggregated_summary.xlsx under REPORT_ROOT and combine."""
    frames: list[pd.DataFrame] = []
    for folder in sorted(REPORT_ROOT.iterdir()):
        if not folder.is_dir():
            continue
        agg = folder / "aggregated_summary.xlsx"
        if not agg.exists():
            print(f"  skip {folder.name}: no aggregated_summary.xlsx")
            continue
        try:
            df = pd.read_excel(agg, sheet_name="Aggregated Summary")
            frames.append(df)
            print(f"  loaded {folder.name}: {len(df)} run(s)")
        except Exception as exc:
            print(f"  skip {folder.name}: {exc}")
    if not frames:
        raise ValueError("No aggregated_summary.xlsx files found under " + str(REPORT_ROOT))
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Common-span chain reclamp
# ---------------------------------------------------------------------------
# chain_ret/annual/n are written per run over that run's OWN span, so they are not
# comparable across strategies (a strategy that only covers recent daynums shows a
# bigger chain return than one spanning a longer, choppier history). Recompute them
# here over the span every compared strategy shares: [max(EndDaynum), min(StartDaynum)].

def _load_hopdata(strategy_name, source_file) -> list[tuple[int, float, float]] | None:
    """Read a run's HopData sheet -> [(daynum, gain, gspc_rsi), ...]."""
    if not strategy_name or not source_file or pd.isna(source_file):
        return None
    path = REPORT_ROOT / str(strategy_name) / str(source_file)
    if not path.exists():
        return None
    try:
        hd = pd.read_excel(path, sheet_name="HopData")
    except Exception:
        return None
    out = []
    for _, r in hd.iterrows():
        out.append((
            int(r["daynum"]),
            float(r["gain"]) if pd.notna(r.get("gain")) else float("nan"),
            float(r["gspc_rsi"]) if pd.notna(r.get("gspc_rsi")) else float("nan"),
        ))
    return out


def reclamp_chains(df: pd.DataFrame) -> tuple[pd.DataFrame, int | None, int | None]:
    """Recompute chain_ret/annual/n over the span shared by every compared run.

    The common span is read straight from each run's HopData — floor = the newest of the
    runs' oldest evaluated hops, cap = the oldest of the runs' newest hops — NOT from the
    Summary StartDaynum/EndDaynum, which now carry each strategy's *usable* span (and so
    differ per strategy; clamping on them would wrongly trim a longer-history strategy like
    Ranknow). Each run's hold horizon is its own `period`; the chain is phase-averaged.
    """
    needed = {"StrategyName", "source_file", "period"}
    if not needed.issubset(df.columns):
        print("  chain reclamp skipped: missing columns "
              f"{sorted(needed - set(df.columns))}")
        return df, None, None

    df = df.copy()
    if "origin_sens%" in df.columns:
        # Recomputed below as "min max" text, not a number -> needs an object dtype
        # column or the per-row string assignment upcasts and raises.
        df["origin_sens%"] = df["origin_sens%"].astype(object)

    # First pass: load each run's HopData once and find its evaluated daynum range.
    hop_cache: dict = {}
    run_min: list[int] = []
    run_max: list[int] = []
    for idx, row in df.iterrows():
        rows = _load_hopdata(row.get("StrategyName"), row.get("source_file"))
        hop_cache[idx] = rows
        if rows:
            dns = [r[0] for r in rows]
            run_min.append(min(dns))
            run_max.append(max(dns))
    if not run_min:
        return df, None, None

    floor, cap = max(run_min), min(run_max)   # common evaluated overlap
    df["chain_floor"], df["chain_cap"] = floor, cap

    n_done = n_missing = 0
    for idx, row in df.iterrows():
        rows = hop_cache.get(idx)
        if rows is None:
            n_missing += 1
            continue
        thr  = row.get("No_go_GSPC_rsi")
        thr  = None if pd.isna(thr) else float(thr)
        hold = int(row["period"])
        ret, annual, ntr = realizable_chain(
            ((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap,
            phase_average=True)
        df.at[idx, "chain_ret"]    = round(ret, 4)    if pd.notna(ret)    else None
        df.at[idx, "chain_annual"] = round(annual, 4) if pd.notna(annual) else None
        df.at[idx, "chain_n"]      = ntr

        # Per-lot dispersion of the CHAIN's ~chain_n non-overlapping lots, on the common
        # span — overwrites the run-Summary avg_gain/Worst/N_loss (those were over ALL of
        # the run's hops, over the run's own span; that is why Ranknow showed N_loss=42
        # next to chain_n=17). Now N_loss <= chain_n by construction.
        cavg, cworst, cnloss = chain_lot_stats(
            ((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap,
            phase_average=True)
        df.at[idx, "avg_gain"] = round(cavg, 4)   if pd.notna(cavg)   else None
        df.at[idx, "Worst"]    = round(cworst, 4) if pd.notna(cworst) else None
        df.at[idx, "N_loss"]   = cnloss

        # Chain's invested share of its active span (%) — idle when No_go gating or too few
        # survivors leave no usable hop at a reinvest point. The chain twin of ladder_inv%.
        cinv = chain_inv_pct(((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap)
        df.at[idx, "chain_inv%"] = round(cinv, 1) if pd.notna(cinv) else None

        # Min/max of chain_annual across start origins, rounded to 1 decimal. Diagnostic;
        # never feeds ranking. Stored as "min max" text, not a number — see the
        # origin_sens% comment for why a collapsed ratio was dropped in favor of this.
        omin, omax = chain_origin_range(((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap)
        df.at[idx, "origin_sens%"] = (f"{omin:.1f}   {omax:.1f}"
                                       if pd.notna(omin) and pd.notna(omax) else None)

        # Diagnostic only — overlap (always-invested) estimate over the same span.
        # Never feeds selection; ranking stays on chain_annual.
        step_v = row.get("step")
        step_v = int(step_v) if pd.notna(step_v) else None
        if step_v:
            lret, lannual, _sleeves, linv = laddered_portfolio(
                ((r[0], r[1], r[2]) for r in rows), hold, step_v, thr, floor, cap)
            # The overlap buys at every hop, so its dispersion is over the whole investable
            # population (~chain_n * hold/step investments) — independently computed, NOT
            # copied from the chain. ladder_n is that total investment count.
            lavg, lworst, lnloss, lcount = laddered_lot_stats(
                ((r[0], r[1], r[2]) for r in rows), thr, floor, cap)
            df.at[idx, "ladder_ret"]      = round(lret, 4)    if pd.notna(lret)    else None
            df.at[idx, "ladder_annual"]   = round(lannual, 4) if pd.notna(lannual) else None
            df.at[idx, "ladder_n"]        = lcount
            df.at[idx, "ladder_inv%"]     = round(linv * 100, 1) if pd.notna(linv) else None
            df.at[idx, "ladder_avg_gain"] = round(lavg, 4)   if pd.notna(lavg)   else None
            df.at[idx, "ladder_worst"]    = round(lworst, 4) if pd.notna(lworst) else None
            df.at[idx, "ladder_n_loss"]   = lnloss
        n_done += 1

    print(f"  chain reclamped to common span [{floor}, {cap}] for {n_done} run(s)"
          + (f"; {n_missing} run(s) lacked HopData (kept original)" if n_missing else ""))
    return df, floor, cap


# ---------------------------------------------------------------------------
# Selection
# ---------------------------------------------------------------------------

def _sort_cols(df: pd.DataFrame, primary: str, tiebreaker: str) -> list[str]:
    return [c for c in [primary, tiebreaker] if c in df.columns]


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def _cell_val(row: pd.Series, col: str):
    v = row.get(col)
    return None if pd.isna(v) else v


def _style_gain_cell(cell, val, col: str) -> None:
    cell.alignment = _CTR
    if _is_gain_col(col) and isinstance(val, (int, float)):
        cell.number_format = _PCT_FMT
        cell.fill = _GRN_FILL if val >= 0 else _RED_FILL


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def best_run(group: pd.DataFrame, primary: str, tiebreaker: str) -> pd.Series | None:
    """That strategy's best run for one criterion (primary desc, tiebreaker desc)."""
    cols  = _sort_cols(group, primary, tiebreaker)
    if not cols:
        return None
    valid = group.dropna(subset=cols[:1])
    if valid.empty:
        return None
    return valid.sort_values(cols, ascending=False).iloc[0]


_HDR_ROW = 3   # the strategy-header row; titles occupy rows 1-2 above it


def fill_best_sheet(ws, columns: list[dict], chained_rows: list[str],
                    floor: int | None = None, cap: int | None = None,
                    period: int | None = None,
                    sheet_title: str = "Best Strategy") -> None:
    """
    Write the strategy comparison into the GIVEN worksheet `ws`: two side-by-side tables
    sharing the same strategy columns — a left "Chain investment" table and a right
    "Overlap investment" table, row-aligned, with two title rows above both.

    The caller owns the workbook (this is sheet 1 of the combined best_strategy_<date>.xlsx,
    built by extension.run()); this function only populates and styles the sheet.

    columns: [{"strategy", "row"}], one per strategy (its best run by chain_annual).
    chained_rows: chained metric names in display order; each row's overlap twin is
                  resolved via _CHAIN_TO_LADDER (shared keys repeat in both tables).
    floor/cap/period: unused by the title rows now (kept for signature compatibility
                      with the caller in extension.py).
    """
    ws.title = sheet_title
    ns = len(columns)

    # ---- column geometry: [A]label [B]comment [chained strats] | [H]label [I]comment [overlap strats]
    A_LBL, A_CMT = 1, 2
    chain_c0 = 3                 # first chained strategy column (C)
    lad_lbl  = chain_c0 + ns     # overlap metric label    (H when ns=5)
    lad_cmt  = lad_lbl + 1       # overlap comment         (I)
    lad_c0   = lad_lbl + 2       # first overlap strategy column (J)

    # ---- title rows (1-2): a bold page title, then a plain left/right table label ----
    ws.cell(1, A_LBL, "Strategy comparison").font = _TITLE_FONT
    ws.cell(2, A_LBL, "Chain investment")
    ws.cell(2, lad_lbl, "overlap investment")

    # ---- header row ----
    for col, text in ((A_LBL, "Chain investment"), (A_CMT, "Comment"),
                      (lad_lbl, _LADDER_TITLE), (lad_cmt, "Comment")):
        h = ws.cell(_HDR_ROW, col, text); h.font, h.fill = _BOLD, _HDR_FILL
    for j, col in enumerate(columns):
        for base in (chain_c0, lad_c0):
            h = ws.cell(_HDR_ROW, base + j, col["strategy"])
            h.font, h.fill, h.alignment = _BOLD, _SECT_FILL, _CTR

    # ---- one row per chained metric, with its overlap twin on the right ----
    for i, metric in enumerate(chained_rows, start=_HDR_ROW + 1):
        lad_metric = _CHAIN_TO_LADDER.get(metric, metric)   # default: shared verbatim

        lbl = ws.cell(i, A_LBL, metric); lbl.font = _BOLD
        lbl.fill = _PARAM_FILL if metric in _PARAM_COLS else _HDR_FILL
        ws.cell(i, A_CMT, _COMMENTS_CHAIN.get(metric))
        if lad_metric is not None:              # None => dropped from the overlap table
            rlbl = ws.cell(i, lad_lbl, lad_metric); rlbl.font = _BOLD
            rlbl.fill = _PARAM_FILL if lad_metric in _PARAM_COLS else _HDR_FILL
            # ladder_* rows from _COMMENTS_LADDER; shared rows fall back to the chain text
            ws.cell(i, lad_cmt,
                    _COMMENTS_LADDER.get(lad_metric, _COMMENTS_CHAIN.get(lad_metric)))

        for j, col in enumerate(columns):
            row = col["row"]
            cv = _cell_val(row, metric) if row is not None else None
            cc = ws.cell(i, chain_c0 + j, cv); _style_gain_cell(cc, cv, metric)
            if metric == "chain_annual":        # the ranking criterion
                cc.font = _BOLD
            elif metric == "origin_sens%":       # "min   max" text is wide — shrink to fit
                cc.font = _SMALL
            if lad_metric is not None:
                lv = _cell_val(row, lad_metric) if row is not None else None
                lc = ws.cell(i, lad_c0 + j, lv); _style_gain_cell(lc, lv, lad_metric)
                if lad_metric == "ladder_annual":
                    lc.font = _BOLD

    # ---- widths / freeze ----
    ws.column_dimensions[get_column_letter(A_LBL)].width = 16
    ws.column_dimensions[get_column_letter(A_CMT)].width = 38
    ws.column_dimensions[get_column_letter(lad_lbl)].width = 16
    ws.column_dimensions[get_column_letter(lad_cmt)].width = 38
    for j in range(ns):
        ws.column_dimensions[get_column_letter(chain_c0 + j)].width = 16
        ws.column_dimensions[get_column_letter(lad_c0 + j)].width = 16
    ws.freeze_panes = f"{get_column_letter(chain_c0)}{_HDR_ROW + 1}"


# ---------------------------------------------------------------------------
# Ranking (shared by main() and extension.py)
# ---------------------------------------------------------------------------

def _order_key(name: str) -> tuple[int, str]:
    # Fixed column order from STRATEGY_ORDER (sweep_config.py) so the report doesn't
    # reshuffle as swept params change which strategy performs best. Unlisted
    # strategies sort after all listed ones, alphabetically among themselves.
    try:
        idx = STRATEGY_ORDER.index(name)
    except ValueError:
        idx = len(STRATEGY_ORDER)
    return (idx, name)


def select_best_runs(verbose: bool = False) -> tuple[list[dict], list[str]]:
    """Load every run, group by forward horizon, reclamp each group to ITS common span,
    and order strategies.

    Returns (blocks, all_cols):
      blocks   : one dict per horizon, ordered by period ascending (the smallest —
                 normally 20d — is the primary): {"period", "floor", "cap", "columns"},
                 where columns = [{"strategy", "row"}], one per strategy (its best run
                 by chain_annual, chain_ret as tiebreaker) in STRATEGY_ORDER order.
                 Chains of different hold lengths are never compared in one block.
      all_cols : every column present across the loaded runs (for metric-row order).

    Returns an empty blocks list (never raises) when there is nothing comparable.
    """
    df = load_all_runs()
    if verbose:
        print(f"Total runs across all strategies: {len(df)}")

    if "StrategyName" not in df.columns:
        if verbose:
            print("No StrategyName column — nothing to report.")
        return [], list(df.columns)

    if "period" in df.columns:
        per = pd.to_numeric(df["period"], errors="coerce")
        period_values: list[int | None] = [int(p) for p in sorted(per.dropna().unique())]
    else:
        per = None
        period_values = [None]

    blocks: list[dict] = []
    all_cols: list[str] = []
    for pv in period_values:
        sub = df if pv is None else df[per == pv].copy()
        if verbose:
            print(f"\n-- horizon {pv}d: {len(sub)} run(s) --" if pv is not None else "")
        sub, floor, cap = reclamp_chains(sub)
        all_cols += [c for c in sub.columns if c not in all_cols]

        groups = {str(name): g for name, g in sub.groupby("StrategyName", sort=False)}
        order = sorted(groups, key=_order_key)
        columns = [{"strategy": name,
                    "row": best_run(groups[name], "chain_annual", "chain_ret")}
                   for name in order]
        blocks.append({"period": pv, "floor": floor, "cap": cap, "columns": columns})

    return blocks, all_cols


def chained_rows_for(all_cols: list[str]) -> list[str]:
    """Display order of the LEFT (chained) table's metric rows: the curated _CHAINED_KEYS
    that are present, then any remaining columns, with _DROP_ROWS excluded throughout."""
    rows = [c for c in _CHAINED_KEYS if c in all_cols and c not in _DROP_ROWS]
    rows += [c for c in all_cols if c not in rows and c not in _DROP_ROWS]
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Output is the single combined workbook (comparison sheet + one extension sheet per
    # strategy). Delegate to the combined builder so running this script alone still emits
    # exactly that one file. Local import avoids the best_strategy <-> extension cycle.
    import extension
    extension.run()


if __name__ == "__main__":
    main()
