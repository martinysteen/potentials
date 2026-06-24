"""
Compare strategies side by side — one column per strategy, metrics down the rows.

Each strategy's column is its best run by chain_cagr (phase-averaged and re-clamped
to the span every strategy shares), with chain_ret as tiebreaker. All runs must share
the same forward horizon (the `period` param); a mismatch is reported and aborts.

Output:
  app/report/best_strategy.xlsx  — transposed, strategies as columns, sorted by chain_cagr

Usage (from app/code/):
  python best_strategy.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.chain import (realizable_chain, laddered_portfolio,
                          chain_lot_stats, laddered_lot_stats)
from shared.config import REPORT_ROOT
from shared.data_loader import daynum_to_date

# Decision metric is chain CAGR (phase-averaged, common-span); chain_ret breaks
# ties. Per-strategy columns are defined by _SUBCOLS near main().
OUTPUT_XLSX = REPORT_ROOT / "best_strategy.xlsx"

# ===========================================================================
# Human-facing text — EDIT FREELY. Everything a reader sees as prose in the
# workbook is gathered here: the two title rows, the right-table header, and
# every row's plain-language Comment. Nothing here affects any calculation —
# only the words on the page.
#
# _TITLE1/_TITLE2 are .format()-ed in write_xlsx with the common-span bounds, so
# keep the {placeholders} intact. Comments are split by table: _COMMENTS_CHAIN
# for the LEFT (chained) column B, _COMMENTS_LADDER for the RIGHT (ladder)
# column I. A metric with no entry shows a blank comment, so add/remove keys
# freely — the key must match the row's name (the metric/param identifier in
# column A / H). Shared rows (period, focusset_size, …) live only in
# _COMMENTS_CHAIN and are reused for column I, so edit them once.
# ===========================================================================
_TITLE1 = ("Comparison of strategies all recalculated for performance in period "
           "daynum {floor} to {cap} ({floor_date} to {cap_date})")
_TITLE2 = ("Using {period} days as investment horizon, this allows a chain of "
           "{n_runs} not-overlapping investment runs.")
_TITLE1_FALLBACK = "Comparison of strategies"    # shown when the common span is unknown
_TITLE2_FALLBACK = ""

_LADDER_TITLE = "Ladder investment"              # header of the right-hand table

_COMMENTS_CHAIN = {     # left table (column B) — plus the shared rows reused on the right
    "period":        "Trading days (per investment)",
    "chain_cagr":    "Annualized return (stacked gain%)",
    "chain_ret":     "Return of full chain (stacked gain%)",
    "chain_n":       "Number of lots in full chain",
    "avg_gain":      "Average gain% per chain lot",
    "Worst":         "Worst chain lot (gain%)",
    "N_loss":        "Negative lots in chain (of chain_n)",
    "focusset_size": "Number of stocks in each investment lot",
}
_COMMENTS_LADDER = {    # right table (column I) — the ladder_* rows only
    "ladder_cagr":   "Annualized return, laddered always-invested style (diagnostic)",
    "ladder_ret":    "Total return of laddered portfolio (stacked gain%)",
    "ladder_n":      "Total laddered investments (period/step sleeves, continuous)",
    "ladder_inv%":   "Share of tranches invested vs cash (%)",
    "ladder_avg_gain": "Average gain% per laddered investment",
    "ladder_worst":  "Worst laddered investment (gain%)",
    "ladder_n_loss": "Negative investments in ladder (of ladder_n)",
}

# The report is two side-by-side tables sharing the same strategy columns: a left
# "chained" table and a right "Ladder investment" table. _CHAINED_KEYS is the row
# order of the LEFT table (these first, then any remaining param cols appended). The
# right table mirrors it row-for-row, swapping chained metrics for their laddered
# twins via _CHAIN_TO_LADDER.
#
# StrategyName is absent: the strategy name is the column header, so a row would just
# repeat it. Floor/cap are absent: they are already stated in the title rows.
_CHAINED_KEYS = [
    "Run#", "period",
    "chain_cagr", "chain_ret", "chain_n",
    "avg_gain", "Worst", "N_loss",
    "N_hops", "N_hops_active",
    "focusset_size", "step", "No_go_GSPC_rsi",
    "source_file",
]
_GAIN_COLS = {"avg_gain", "Worst", "ladder_avg_gain", "ladder_worst"}

# Never a row of their own: StrategyName (it's the header); floor/cap (they're in the
# title); the ladder_* metrics (rendered in the right table by mapping their chained
# twin, not as standalone rows).
_DROP_ROWS = {"StrategyName", "chain_floor", "chain_cap",
              "ladder_cagr", "ladder_ret", "ladder_n", "ladder_inv%",
              "ladder_avg_gain", "ladder_worst", "ladder_n_loss"}

# Chained metric -> its counterpart in the right-hand laddered table. A key not listed
# is shared verbatim (identical value, shown in both tables for every strategy). None
# drops the row from the laddered side (only source_file does). avg_gain/Worst/N_loss
# are per-investment dispersion and so are NOT shared — each side gets its own set,
# computed over its own (chain vs ladder) investment population.
_CHAIN_TO_LADDER = {
    "chain_cagr":    "ladder_cagr",
    "chain_ret":     "ladder_ret",
    "chain_n":       "ladder_n",
    "avg_gain":      "ladder_avg_gain",
    "Worst":         "ladder_worst",
    "N_loss":        "ladder_n_loss",
    "N_hops_active": "ladder_inv%",
    "source_file":   None,
}
_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _fmt_date(daynum: int) -> str:
    """daynum -> 'D. Mon YYYY' (e.g. 1804 -> '8. Jan 2025'); ISO/fallback on parse miss."""
    iso = daynum_to_date(int(daynum))  # "YYYY-MM-DD" from Cal.csv
    try:
        y, m, d = iso.split("-")
        return f"{int(d)}. {_MONTHS[int(m) - 1]} {y}"
    except Exception:
        return iso


def _is_gain_col(col: str) -> bool:
    return (col in _GAIN_COLS or "gain" in col
            or col.startswith(("chain_ret", "chain_cagr", "ladder_ret", "ladder_cagr")))

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_BOLD       = Font(bold=True)
_SMALL      = Font(size=9)
_HDR_FILL   = PatternFill("solid", fgColor="BDD7EE")   # blue header row
_SECT_FILL  = PatternFill("solid", fgColor="D6DCE4")   # grey section header
_GRN_FILL   = PatternFill("solid", fgColor="C6EFCE")   # green
_RED_FILL   = PatternFill("solid", fgColor="FFC7CE")   # red
_BEST_FILL  = PatternFill("solid", fgColor="FFE599")   # amber for "Best overall" row
_PARAM_FILL = PatternFill("solid", fgColor="FFFF99")   # yellow for simulation parameter headers
_PCT_FMT    = '+0.00;-0.00;"-"'
_CTR        = Alignment(horizontal="center")

_PARAM_COLS = {"focusset_size", "step", "period", "No_go_GSPC_rsi",
               "p20d_win_min", "p50d_win_min", "q10_20_min", "q20_50_min"}


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def load_all_runs() -> pd.DataFrame:
    """Read every aggregated_summary.xlsx under REPORT_ROOT and combine."""
    frames: list[pd.DataFrame] = []
    for folder in sorted(REPORT_ROOT.iterdir()):
        if not folder.is_dir():
            continue
        agg = folder / "aggregated_summary.xlsx"
        if not agg.exists():
            print(f"  skip {folder.name}: no aggregated_summary.xlsx")
            continue
        try:
            df = pd.read_excel(agg, sheet_name="Aggregated Summary")
            frames.append(df)
            print(f"  loaded {folder.name}: {len(df)} run(s)")
        except Exception as exc:
            print(f"  skip {folder.name}: {exc}")
    if not frames:
        raise ValueError("No aggregated_summary.xlsx files found under " + str(REPORT_ROOT))
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Common-span chain reclamp
# ---------------------------------------------------------------------------
# chain_ret/cagr/n are written per run over that run's OWN span, so they are not
# comparable across strategies (a strategy that only covers recent daynums shows a
# bigger chain return than one spanning a longer, choppier history). Recompute them
# here over the span every compared strategy shares: [max(EndDaynum), min(StartDaynum)].

def _load_hopdata(strategy_name, source_file) -> list[tuple[int, float, float]] | None:
    """Read a run's HopData sheet -> [(daynum, gain, gspc_rsi_prev), ...]."""
    if not strategy_name or not source_file or pd.isna(source_file):
        return None
    path = REPORT_ROOT / str(strategy_name) / str(source_file)
    if not path.exists():
        return None
    try:
        hd = pd.read_excel(path, sheet_name="HopData")
    except Exception:
        return None
    out = []
    for _, r in hd.iterrows():
        out.append((
            int(r["daynum"]),
            float(r["gain"]) if pd.notna(r.get("gain")) else float("nan"),
            float(r["gspc_rsi_prev"]) if pd.notna(r.get("gspc_rsi_prev")) else float("nan"),
        ))
    return out


def reclamp_chains(df: pd.DataFrame) -> tuple[pd.DataFrame, int | None, int | None]:
    """Recompute chain_ret/cagr/n over the span shared by every compared run.

    Each run's hold horizon is its own `period`; the chain is phase-averaged.
    """
    needed = {"EndDaynum", "StartDaynum", "StrategyName", "source_file", "period"}
    if not needed.issubset(df.columns):
        print("  chain reclamp skipped: missing columns "
              f"{sorted(needed - set(df.columns))}")
        return df, None, None
    ends   = pd.to_numeric(df["EndDaynum"], errors="coerce").dropna()
    starts = pd.to_numeric(df["StartDaynum"], errors="coerce").dropna()
    if ends.empty or starts.empty:
        return df, None, None

    floor, cap = int(ends.max()), int(starts.min())
    df = df.copy()
    df["chain_floor"], df["chain_cap"] = floor, cap

    n_done = n_missing = 0
    for idx, row in df.iterrows():
        rows = _load_hopdata(row.get("StrategyName"), row.get("source_file"))
        if rows is None:
            n_missing += 1
            continue
        thr  = row.get("No_go_GSPC_rsi")
        thr  = None if pd.isna(thr) else float(thr)
        hold = int(row["period"])
        ret, cagr, ntr = realizable_chain(
            ((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap,
            phase_average=True)
        df.at[idx, "chain_ret"]  = round(ret, 4)  if pd.notna(ret)  else None
        df.at[idx, "chain_cagr"] = round(cagr, 4) if pd.notna(cagr) else None
        df.at[idx, "chain_n"]    = ntr

        # Per-lot dispersion of the CHAIN's ~chain_n non-overlapping lots, on the common
        # span — overwrites the run-Summary avg_gain/Worst/N_loss (those were over ALL of
        # the run's hops, over the run's own span; that is why Ranknow showed N_loss=42
        # next to chain_n=17). Now N_loss <= chain_n by construction.
        cavg, cworst, cnloss = chain_lot_stats(
            ((r[0], r[1], r[2]) for r in rows), hold, thr, floor, cap,
            phase_average=True)
        df.at[idx, "avg_gain"] = round(cavg, 4)   if pd.notna(cavg)   else None
        df.at[idx, "Worst"]    = round(cworst, 4) if pd.notna(cworst) else None
        df.at[idx, "N_loss"]   = cnloss

        # Diagnostic only — laddered (always-invested) estimate over the same span.
        # Never feeds selection; ranking stays on chain_cagr.
        step_v = row.get("step")
        step_v = int(step_v) if pd.notna(step_v) else None
        if step_v:
            lret, lcagr, _sleeves, linv = laddered_portfolio(
                ((r[0], r[1], r[2]) for r in rows), hold, step_v, thr, floor, cap)
            # The ladder buys at every hop, so its dispersion is over the whole investable
            # population (~chain_n * hold/step investments) — independently computed, NOT
            # copied from the chain. ladder_n is that total investment count.
            lavg, lworst, lnloss, lcount = laddered_lot_stats(
                ((r[0], r[1], r[2]) for r in rows), thr, floor, cap)
            df.at[idx, "ladder_ret"]      = round(lret, 4)   if pd.notna(lret)   else None
            df.at[idx, "ladder_cagr"]     = round(lcagr, 4)  if pd.notna(lcagr)  else None
            df.at[idx, "ladder_n"]        = lcount
            df.at[idx, "ladder_inv%"]     = round(linv * 100, 1) if pd.notna(linv) else None
            df.at[idx, "ladder_avg_gain"] = round(lavg, 4)   if pd.notna(lavg)   else None
            df.at[idx, "ladder_worst"]    = round(lworst, 4) if pd.notna(lworst) else None
            df.at[idx, "ladder_n_loss"]   = lnloss
        n_done += 1

    print(f"  chain reclamped to common span [{floor}, {cap}] for {n_done} run(s)"
          + (f"; {n_missing} run(s) lacked HopData (kept original)" if n_missing else ""))
    return df, floor, cap


# ---------------------------------------------------------------------------
# Selection
# ---------------------------------------------------------------------------

def _sort_cols(df: pd.DataFrame, primary: str, tiebreaker: str) -> list[str]:
    return [c for c in [primary, tiebreaker] if c in df.columns]


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def _cell_val(row: pd.Series, col: str):
    v = row.get(col)
    return None if pd.isna(v) else v


def _style_gain_cell(cell, val, col: str) -> None:
    cell.alignment = _CTR
    if _is_gain_col(col) and isinstance(val, (int, float)):
        cell.number_format = _PCT_FMT
        cell.fill = _GRN_FILL if val >= 0 else _RED_FILL


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def best_run(group: pd.DataFrame, primary: str, tiebreaker: str) -> pd.Series | None:
    """That strategy's best run for one criterion (primary desc, tiebreaker desc)."""
    cols  = _sort_cols(group, primary, tiebreaker)
    if not cols:
        return None
    valid = group.dropna(subset=cols[:1])
    if valid.empty:
        return None
    return valid.sort_values(cols, ascending=False).iloc[0]


_HDR_ROW = 3   # the strategy-header row; titles occupy rows 1-2 above it


def write_xlsx(columns: list[dict], chained_rows: list[str],
               floor: int | None = None, cap: int | None = None,
               period: int | None = None) -> None:
    """
    Two side-by-side tables sharing the same strategy columns: a left "chained" table
    and a right "Ladder investment" table, row-aligned. Two title rows above both.

    columns: [{"strategy", "row"}], one per strategy (its best run by chain_cagr).
    chained_rows: chained metric names in display order; each row's laddered twin is
                  resolved via _CHAIN_TO_LADDER (shared keys repeat in both tables).
    floor/cap/period: common-span bounds + horizon, used to phrase the title rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Best Strategy"
    ns = len(columns)

    # ---- title rows (1-2): plain-language summary for unprepared readers ----
    # (text templates live in the human-facing block at the top of the module)
    if floor is not None and cap is not None and period:
        title1 = _TITLE1.format(floor=floor, cap=cap,
                                floor_date=_fmt_date(floor), cap_date=_fmt_date(cap))
        title2 = _TITLE2.format(period=period, n_runs=(cap - floor) // period)
    else:
        title1, title2 = _TITLE1_FALLBACK, _TITLE2_FALLBACK
    ws.cell(1, 1, title1).font = _BOLD
    ws.cell(2, 1, title2).font = _BOLD

    # ---- column geometry: [A]label [B]comment [chained strats] | [H]label [I]comment [laddered strats]
    A_LBL, A_CMT = 1, 2
    chain_c0 = 3                 # first chained strategy column (C)
    lad_lbl  = chain_c0 + ns     # laddered metric label   (H when ns=5)
    lad_cmt  = lad_lbl + 1       # laddered comment        (I)
    lad_c0   = lad_lbl + 2       # first laddered strategy column (J)

    # ---- header row ----
    for col, text in ((A_LBL, "StrategyName"), (A_CMT, "Comment"),
                      (lad_lbl, _LADDER_TITLE), (lad_cmt, "Comment")):
        h = ws.cell(_HDR_ROW, col, text); h.font, h.fill = _BOLD, _HDR_FILL
    for j, col in enumerate(columns):
        for base in (chain_c0, lad_c0):
            h = ws.cell(_HDR_ROW, base + j, col["strategy"])
            h.font, h.fill, h.alignment = _BOLD, _SECT_FILL, _CTR

    # ---- one row per chained metric, with its laddered twin on the right ----
    for i, metric in enumerate(chained_rows, start=_HDR_ROW + 1):
        lad_metric = _CHAIN_TO_LADDER.get(metric, metric)   # default: shared verbatim

        lbl = ws.cell(i, A_LBL, metric); lbl.font = _BOLD
        lbl.fill = _PARAM_FILL if metric in _PARAM_COLS else _HDR_FILL
        ws.cell(i, A_CMT, _COMMENTS_CHAIN.get(metric))
        if lad_metric is not None:              # None => dropped from the laddered table
            rlbl = ws.cell(i, lad_lbl, lad_metric); rlbl.font = _BOLD
            rlbl.fill = _PARAM_FILL if lad_metric in _PARAM_COLS else _HDR_FILL
            # ladder_* rows from _COMMENTS_LADDER; shared rows fall back to the chain text
            ws.cell(i, lad_cmt,
                    _COMMENTS_LADDER.get(lad_metric, _COMMENTS_CHAIN.get(lad_metric)))

        for j, col in enumerate(columns):
            row = col["row"]
            cv = _cell_val(row, metric) if row is not None else None
            cc = ws.cell(i, chain_c0 + j, cv); _style_gain_cell(cc, cv, metric)
            if metric == "chain_cagr":          # the ranking criterion
                cc.font = _BOLD
            if lad_metric is not None:
                lv = _cell_val(row, lad_metric) if row is not None else None
                lc = ws.cell(i, lad_c0 + j, lv); _style_gain_cell(lc, lv, lad_metric)
                if lad_metric == "ladder_cagr":
                    lc.font = _BOLD

    # ---- widths / freeze ----
    ws.column_dimensions[get_column_letter(A_LBL)].width = 16
    ws.column_dimensions[get_column_letter(A_CMT)].width = 38
    ws.column_dimensions[get_column_letter(lad_lbl)].width = 16
    ws.column_dimensions[get_column_letter(lad_cmt)].width = 38
    for j in range(ns):
        ws.column_dimensions[get_column_letter(chain_c0 + j)].width = 16
        ws.column_dimensions[get_column_letter(lad_c0 + j)].width = 16
    ws.freeze_panes = f"{get_column_letter(chain_c0)}{_HDR_ROW + 1}"

    wb.save(OUTPUT_XLSX)
    print(f"Written: {OUTPUT_XLSX}")


# ---------------------------------------------------------------------------
# Ranking (shared by main() and extension_of_best_strategy.py)
# ---------------------------------------------------------------------------

def select_best_runs(verbose: bool = False) -> tuple[
        list[dict], list[str], int | None, int | None, int | None]:
    """Load every run, reclamp chains to the common span, and rank strategies.

    Returns (columns, all_cols, floor, cap, period):
      columns  : [{"strategy", "row"}], one per strategy (its best run by chain_cagr,
                 chain_ret as tiebreaker), strongest chain_cagr first.
      all_cols : every column present across the loaded runs (for metric-row order).
      floor/cap: common-span bounds. period: the shared forward horizon.

    Returns an empty columns list (never raises) when there is nothing comparable —
    no StrategyName column, or runs mixing forward horizons.
    """
    df = load_all_runs()
    if verbose:
        print(f"Total runs across all strategies: {len(df)}")
    df, floor, cap = reclamp_chains(df)
    if verbose:
        print()

    all_cols = list(df.columns)
    if "StrategyName" not in df.columns:
        if verbose:
            print("No StrategyName column — nothing to report.")
        return [], all_cols, floor, cap, None

    # Guard: every compared run must share the same forward horizon.
    period: int | None = None
    if "period" in df.columns:
        periods = sorted(pd.to_numeric(df["period"], errors="coerce").dropna().unique())
        if len(periods) > 1:
            print(f"ERROR: mixed periods across runs {periods} — re-run the sweep at a "
                  "single period before comparing. Aborting.")
            return [], all_cols, floor, cap, None
        if periods:
            period = int(periods[0])

    groups = {str(name): g for name, g in df.groupby("StrategyName", sort=False)}

    def _rank(name: str) -> float:
        r = best_run(groups[name], "chain_cagr", "chain_ret")
        v = r.get("chain_cagr") if r is not None else None
        return float(v) if v is not None and pd.notna(v) else float("-inf")
    order = sorted(groups, key=_rank, reverse=True)   # strongest chain_cagr leftmost

    columns = [{"strategy": name, "row": best_run(groups[name], "chain_cagr", "chain_ret")}
               for name in order]
    return columns, all_cols, floor, cap, period


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Loading aggregated summaries...")
    columns, all_cols, floor, cap, period = select_best_runs(verbose=True)
    if not columns:
        return

    chained_rows = [c for c in _CHAINED_KEYS if c in all_cols and c not in _DROP_ROWS]
    chained_rows += [c for c in all_cols if c not in chained_rows and c not in _DROP_ROWS]

    for col in columns:
        c = col["row"].get("chain_cagr") if col["row"] is not None else None
        print(f"  {col['strategy']:18} best chain_cagr={_fmt(c)}")
    print()

    write_xlsx(columns, chained_rows, floor, cap, period)
    print("Done.")


def _fmt(v) -> str:
    return "n/a" if v is None or pd.isna(v) else f"{float(v):+.1f}"


if __name__ == "__main__":
    main()
