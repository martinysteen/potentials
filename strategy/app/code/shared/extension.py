"""
Extension runner: measures partial realised gains for the ~20 recent days where
future_gain20d is not yet available.

For each strategy, iterates entry daynums from start_daynum+step to today (newest
first), computes gain as (exit_price - entry_price) / entry_price * 100 from PotDat.

Output: app/report/<strategy>/extension_YYYYMMDD.xlsx

Operational sheet layout (focusset_size = N):
  Row 1      : A1="Size/Step/No_RSI/P20/P50"  | daynum headers     (blue)
  Row 2      : A2=param values string          | date headers       (blue)
  Rows 3…N+2 : ticker rows
  Row N+3    : per-column "avg_gainXd" labels                       (light green)
  Row N+4    : avg_partial_gain values                              (orange/light-yellow/grey)
  Row N+5    : (reserved — 50d labels)
  Row N+6    : (reserved — 50d results)
  Row N+7+   : day-1 ref rows (^GSPC_rsi, ^STOXX_rsi, ^HSI_rsi, ^VIX)  (yellow)
  ...        : GICS / Sector2 / Zone occurrence counts

Usage (from app/code/):
  python run_extension.py
"""

from datetime import date
from typing import Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.config import REPORT_ROOT
from shared.data_loader import load_longi, load_potdat, load_stamdata, daynum_to_date


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_BOLD         = Font(bold=True)
_SMALL        = Font(size=9)
_HDR_FILL     = PatternFill("solid", fgColor="BDD7EE")   # blue — headers
_GRY_FILL     = PatternFill("solid", fgColor="EEEEEE")   # grey — suppressed / n/a
_REF_FILL     = PatternFill("solid", fgColor="FFF2CC")   # yellow — day-1 ref rows
_LBL_FILL     = PatternFill("solid", fgColor="E2EFDA")   # light green — label row
_EXT_POS_FILL = PatternFill("solid", fgColor="FFFF99")   # light yellow — positive gain
_EXT_NEG_FILL = PatternFill("solid", fgColor="FFC000")   # orange — negative gain

_ATTR_FILLS: dict[str, PatternFill] = {
    "GICS":    PatternFill("solid", fgColor="D9D2E9"),
    "Sector2": PatternFill("solid", fgColor="FCE5CD"),
    "Zone":    PatternFill("solid", fgColor="D0E0E3"),
}

_PCT_FMT = '+0.00;-0.00;"-"'
_CTR     = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def _find_gain20_cutoff(gain20_df: pd.DataFrame, min_valid: int = 10) -> int:
    """Return the newest daynum where future_gain20d has sufficient realized data."""
    for col in gain20_df.columns:
        if gain20_df[col].dropna().size >= min_valid:
            return int(col)
    raise ValueError("No valid daynum found in future_gain20d.csv")


def _compute_partial_gains(potdat: pd.DataFrame, tickers: list[str],
                            entry_daynum: int, exit_daynum: int) -> dict[str, float]:
    """(price[exit] - price[entry]) / price[entry] * 100. NaN if prices unavailable."""
    ec = str(entry_daynum)
    xc = str(exit_daynum)
    result: dict[str, float] = {}
    for t in tickers:
        try:
            pe = (float(potdat.at[t, ec])
                  if t in potdat.index and ec in potdat.columns and pd.notna(potdat.at[t, ec])
                  else float("nan"))
            px = (float(potdat.at[t, xc])
                  if t in potdat.index and xc in potdat.columns and pd.notna(potdat.at[t, xc])
                  else float("nan"))
            result[t] = ((px - pe) / pe * 100
                         if pd.notna(pe) and pd.notna(px) and pe != 0
                         else float("nan"))
        except (KeyError, ValueError):
            result[t] = float("nan")
    return result


def _hop_avg(gains: dict[str, float]) -> float:
    vals = [v for v in gains.values() if pd.notna(v)]
    return sum(vals) / len(vals) if vals else float("nan")


def _count_attr(hop_results: list[dict], stamdata: pd.DataFrame,
                attr_col: str) -> tuple[list[str], list[dict]]:
    all_vals: set[str] = set()
    hop_counts: list[dict] = []
    for h in hop_results:
        counts: dict[str, int] = {}
        for ticker in h.get("tickers", []):
            if ticker in stamdata.index:
                raw = stamdata.at[ticker, attr_col]
                if pd.notna(raw):
                    v = str(raw).strip()
                    if v:
                        counts[v] = counts.get(v, 0) + 1
                        all_vals.add(v)
        hop_counts.append(counts)
    totals = {v: sum(hc.get(v, 0) for hc in hop_counts) for v in all_vals}
    return sorted(all_vals, key=lambda v: -totals[v]), hop_counts


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _fill_operational(ws, hop_results: list[dict], params: dict) -> None:
    daynums   = [h["daynum"] for h in hop_results]
    n         = params.get("focusset_size",
                            max(len(h.get("tickers", [])) for h in hop_results))
    threshold = params.get("No_go_GSPC_rsi")

    # ---- A1: param label, A2: param values ----
    def _pv(key: str) -> str:
        v = params.get(key)
        return "-" if v is None else str(v)

    c = ws.cell(1, 1, "Size/Step/No_RSI/P20/P50"); c.font, c.fill = _BOLD, _HDR_FILL
    c = ws.cell(2, 1, "/".join([_pv("focusset_size"), _pv("step"),
                                 _pv("No_go_GSPC_rsi"), _pv("p20d_win_min"),
                                 _pv("p50d_win_min")]));  c.font, c.fill = _SMALL, _HDR_FILL

    # ---- daynum / date header columns ----
    for j, dn in enumerate(daynums, start=2):
        c = ws.cell(1, j, dn);                 c.font, c.fill, c.alignment = _BOLD, _HDR_FILL, _CTR
        c = ws.cell(2, j, daynum_to_date(dn)); c.font, c.fill, c.alignment = _SMALL, _HDR_FILL, _CTR

    # ---- ticker rows (3 … N+2) ----
    for i in range(n):
        ws.cell(i + 3, 1, "")
        for j, h in enumerate(hop_results, start=2):
            tickers = h.get("tickers", [])
            ws.cell(i + 3, j, tickers[i] if i < len(tickers) else "")

    # ---- per-column "avg_gainXd" label row (N+3) ----
    lbl_row = n + 3
    c = ws.cell(lbl_row, 1, "realized"); c.font, c.fill = _BOLD, _LBL_FILL
    for j, h in enumerate(hop_results, start=2):
        cell = ws.cell(lbl_row, j, f"avg_gain{h['days_realized']}d")
        cell.fill, cell.alignment, cell.font = _LBL_FILL, _CTR, _SMALL

    # ---- avg_partial_gain values row (N+4) — orange/light-yellow/grey ----
    gain_row = n + 4
    c = ws.cell(gain_row, 1, "avg_partial_gain"); c.font = _BOLD
    for j, h in enumerate(hop_results, start=2):
        val    = _hop_avg(h["gains_partial"])
        gspc_p = h.get("ref_values_prev", {}).get("^GSPC_rsi", float("nan"))
        no_go  = threshold is not None and not pd.isna(gspc_p) and gspc_p < threshold
        cell   = ws.cell(gain_row, j)
        cell.font, cell.number_format, cell.alignment = _BOLD, _PCT_FMT, _CTR
        if pd.notna(val):
            cell.value = None if no_go else round(val, 4)
            cell.fill  = _GRY_FILL if no_go else (_EXT_NEG_FILL if val < 0 else _EXT_POS_FILL)
        else:
            cell.value = None
            cell.fill  = _GRY_FILL

    # ---- rows N+5 and N+6 intentionally blank (reserved for 50d labels / results) ----

    # ---- day-1 ref rows (starting at N+7) ----
    gspc_row  = n + 7
    prev_keys = list(hop_results[0].get("ref_values_prev", {}).keys()) if hop_results else []
    for idx, key in enumerate(prev_keys):
        row = gspc_row + idx
        c = ws.cell(row, 1, f"{key} (day-1)"); c.font, c.fill = _BOLD, _REF_FILL
        for j, h in enumerate(hop_results, start=2):
            val  = h.get("ref_values_prev", {}).get(key, float("nan"))
            cell = ws.cell(row, j)
            cell.fill, cell.alignment = _REF_FILL, _CTR
            if pd.notna(val):
                cell.value         = round(val, 2)
                cell.number_format = "0.0" if key.endswith("_rsi") else "0.00"

    # ---- attribute frequency rows ----
    stamdata = load_stamdata()
    attr_row = gspc_row + len(prev_keys)
    for attr_col in ("GICS", "Sector2", "Zone"):
        fill = _ATTR_FILLS[attr_col]
        sorted_vals, hop_counts = _count_attr(hop_results, stamdata, attr_col)
        for i, val in enumerate(sorted_vals):
            lbl_cell = ws.cell(attr_row + i, 1, f"{attr_col}_{val}")
            lbl_cell.font, lbl_cell.fill = _BOLD, fill
            for j, hc in enumerate(hop_counts, start=2):
                count = hc.get(val, 0)
                cell  = ws.cell(attr_row + i, j)
                cell.fill, cell.alignment = fill, _CTR
                cell.value = count if count > 0 else None
        attr_row += len(sorted_vals)

    # ---- dimensions ----
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(daynums) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = "B3"


def _write_xlsx(strategy_name: str, params: dict, hop_results: list[dict]) -> None:
    folder = REPORT_ROOT / strategy_name
    folder.mkdir(parents=True, exist_ok=True)
    xlsx_path = folder / f"extension_{date.today().strftime('%Y%m%d')}.xlsx"
    wb    = Workbook()
    ws_op = wb.active
    ws_op.title = "Extension"
    _fill_operational(ws_op, hop_results, params)
    wb.save(xlsx_path)
    print(f"** Extension written: {xlsx_path} **")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_extension(
    strategy_name: str,
    params: dict,
    get_focusset_fn: Callable[[int], list[str]],
    get_ref_fn: Callable[[int], dict],
) -> None:
    """
    Run the extension loop and write extension_YYYYMMDD.xlsx.

    get_focusset_fn(daynum) → list[str]  — strategy selector with pre-bound DataFrames
    get_ref_fn(daynum)      → dict       — market context (same as get_reference_values)
    """
    gain20_df    = load_longi("future_gain20d.csv")
    potdat       = load_potdat()
    start_daynum = _find_gain20_cutoff(gain20_df)   # last valid backtest daynum
    exit_daynum  = int(potdat.columns[0])            # most recent price available
    step: int    = params.get("step", 1)
    pot_cols     = set(potdat.columns)

    print(f"--- {strategy_name} extension ---")
    print(f"Backtest cutoff: {start_daynum} ({daynum_to_date(start_daynum)})")
    print(f"Exit daynum    : {exit_daynum}  ({daynum_to_date(exit_daynum)})")
    print(f"Step           : {step}")

    # Entry daynums aligned with the strategy's step, newest-first.
    # Only include daynums that exist in PotDat (guards against any daynum gaps).
    ext_daynums: list[int] = []
    d = start_daynum + step
    while d <= exit_daynum:
        if str(d) in pot_cols:
            ext_daynums.append(d)
        d += step
    ext_daynums.reverse()

    if not ext_daynums:
        print("Extension period is empty — data fully up to date.\n")
        return
    print(f"Entries        : {len(ext_daynums)}"
          f" (daynums {ext_daynums[-1]}→{ext_daynums[0]})\n")

    hop_results: list[dict] = []
    for daynum in ext_daynums:
        tickers = get_focusset_fn(daynum)
        if not tickers:
            continue
        days          = exit_daynum - daynum
        gains_partial = _compute_partial_gains(potdat, tickers, daynum, exit_daynum)
        avg           = _hop_avg(gains_partial)
        print(f"  entry {daynum} ({daynum_to_date(daynum)})"
              f"  tickers={tickers}"
              f"  gain_{days}d avg={avg:+.2f}%")
        hop_results.append({
            "daynum":          daynum,
            "tickers":         tickers,
            "gains_partial":   gains_partial,
            "days_realized":   days,
            "ref_values_prev": get_ref_fn(daynum - 1),
        })

    print()
    if not hop_results:
        print(f"No valid hops — strategy selected nothing in the extension period.\n")
        return

    _write_xlsx(strategy_name, params, hop_results)
    print(f"Done: {len(hop_results)} extension hops.\n")
