"""
Write reports for strategy runs.

Per run:   app/report/<strategy>/run<N>_<date>.xlsx
           Sheet "Operational" — ticker rows + avg rows + ref rows + attribute count rows
           Sheet "Summary"     — metadata + avg gains

Master:    app/report/summary.csv  (one appended row per run, all strategies)
"""

import csv
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.chain import realizable_chain, chain_lot_stats, chain_origin_sensitivity
from shared.config import REPORT_ROOT, SUMMARY_CSV
from shared.data_loader import daynum_to_date, load_stamdata


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hop_avg_topn(hop: dict, gain_key: str, n: int) -> float:
    """Average gain for the top-n tickers (rank order) in one hop."""
    tickers = hop.get("tickers", [])[:n]
    gains   = hop.get(gain_key, {})
    vals    = [gains[t] for t in tickers if pd.notna(gains.get(t, float("nan")))]
    return sum(vals) / len(vals) if vals else float("nan")


def _grand_avg_topn(hop_results: list[dict], gain_key: str, n: int,
                    params: dict | None = None) -> float:
    """Average gain for the top-n tickers across active hops (respects No_go_GSPC_rsi)."""
    threshold = params.get("No_go_GSPC_rsi") if params else None
    vals: list[float] = []
    for h in hop_results:
        if threshold is not None:
            gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            if not pd.isna(gspc) and gspc < threshold:
                continue
        tickers = h.get("tickers", [])[:n]
        gains   = h.get(gain_key, {})
        vals.extend(gains[t] for t in tickers if pd.notna(gains.get(t, float("nan"))))
    return sum(vals) / len(vals) if vals else float("nan")


def _avg_rows(n: int) -> list[tuple[str, str, int]]:
    """The single avg-gain row for the active focusset size — (label, gain_key, top_n).

    Horizon-agnostic: the forward horizon is the `period` param; gains live under the
    single key "gains" in each hop. Names carry no 20d/50d suffix so one run = one column.
    """
    return [("avg_gain", "gains", n)]


# Gains for the chosen horizon live under this single hop key.
_GAIN_KEY = "gains"


def _chain_metric_labels() -> list[str]:
    """Realizable-chain summary labels, in write order."""
    return ["chain_ret", "chain_annual", "chain_n"]


def _chain_metrics(hop_results: list[dict], gain_key: str, n: int,
                   hold: int, params: dict) -> tuple[float, float, int]:
    """
    Realizable non-overlapping ADDITIVE chain over this run's full span.

    Thin wrapper over shared.chain.realizable_chain (the single source of the chain
    math). best_strategy.py re-runs the same function with a common floor/cap so
    cross-strategy chain returns are comparable; see shared/chain.py.

    Returns (total_return_pct, annual_pct, n_trades) — additive sum of lot gains and
    its simple per-year average (not a compound CAGR).
    """
    threshold = params.get("No_go_GSPC_rsi")
    rows = ((h["daynum"], _hop_avg_topn(h, gain_key, n),
             h.get("ref_values", {}).get("^GSPC_rsi", float("nan")))
            for h in hop_results)
    return realizable_chain(rows, hold, threshold, phase_average=True)


def _count_attr(hop_results: list[dict], stamdata: pd.DataFrame,
                attr_col: str) -> tuple[list[str], list[dict]]:
    """
    For each hop count occurrences of each unique value of attr_col in the focusset.
    Returns (sorted_unique_values, list_of_count_dicts) — values sorted by total desc.
    """
    all_vals: set[str] = set()
    hop_counts: list[dict] = []

    for h in hop_results:
        counts: dict[str, int] = {}
        for ticker in h.get("tickers", []):
            if ticker in stamdata.index:
                raw = stamdata.at[ticker, attr_col]
                if pd.notna(raw):
                    v = str(raw).strip()
                    if v:
                        counts[v] = counts.get(v, 0) + 1
                        all_vals.add(v)
        hop_counts.append(counts)

    totals = {v: sum(hc.get(v, 0) for hc in hop_counts) for v in all_vals}
    sorted_vals = sorted(all_vals, key=lambda v: -totals[v])
    return sorted_vals, hop_counts


def _next_run_num(folder: Path) -> int:
    nums = []
    for p in folder.glob("run*.xlsx"):
        try:
            nums.append(int(p.name[3:].split("_")[0]))
        except ValueError:
            pass
    return max(nums, default=0) + 1


def _count_active_hops(hop_results: list[dict], params: dict) -> int:
    """Count hops actually invested: a non-empty focusset surviving No_go_GSPC_rsi.

    No-pick hops (empty tickers — days the strategy sat out, now recorded so they are
    visible in HopData/Operational) are NOT active, so N_hops_active is the number of
    days with a real investment (the chain-side mirror of the ladder's invested count).
    NaN RSI counts as passing the No_go gate.
    """
    threshold = params.get("No_go_GSPC_rsi")
    count = 0
    for h in hop_results:
        if not h.get("tickers"):
            continue
        gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
        if threshold is None or pd.isna(gspc) or gspc >= threshold:
            count += 1
    return count


def _usable_daynums(hop_results: list[dict], gain_key: str, params: dict) -> list[int]:
    """Daynums of hops that actually invest: a real (non-NaN) top-N gain surviving No_go.

    This is the strategy's *usable* span. The cross / win-prob strategies have no real
    gain before the ML warm-up (~daynum 1797), so their usable range starts later than
    the full evaluated range — even though the empty warm-up hops are still recorded.
    """
    threshold = params.get("No_go_GSPC_rsi")
    n = params.get("focusset_size", 10)
    dns: list[int] = []
    for h in hop_results:
        if threshold is not None:
            gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            if not pd.isna(gspc) and gspc < threshold:
                continue
        if pd.notna(_hop_avg_topn(h, gain_key, n)):
            dns.append(int(h["daynum"]))
    return dns


def _chain_hop_rows(hop_results: list[dict], n: int) -> list[tuple[int, float, float]]:
    """(daynum, top-N avg gain, gspc_rsi) per hop — the chain's raw inputs (one span)."""
    return [(h["daynum"], _hop_avg_topn(h, _GAIN_KEY, n),
             h.get("ref_values", {}).get("^GSPC_rsi", float("nan")))
            for h in hop_results]


def _chain_dispersion(hop_results: list[dict], n: int, hold: int,
                      params: dict) -> tuple[float, int, float]:
    """Chained WORST-CASE dispersion over this run's own span (no floor/cap).

    Returns (worst_lot, n_loss, origin_sens_pct) using the SAME rule best_strategy.py
    applies on the common span — worst = lowest lot over all start origins, n_loss = most
    losers in any one origin's chain (worst < 0 <=> n_loss >= 1), sens = annual spread
    across origins. So the run-file Summary and the comparison sheet reconcile (they differ
    only by span, exactly as chain_ret does).
    """
    thr  = params.get("No_go_GSPC_rsi")
    rows = _chain_hop_rows(hop_results, n)
    _avg, worst, n_loss = chain_lot_stats(rows, hold, thr, phase_average=True)
    sens = chain_origin_sensitivity(rows, hold, thr)
    return worst, n_loss, sens


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_BOLD     = Font(bold=True)
_SMALL    = Font(size=9)
_HDR_FILL = PatternFill("solid", fgColor="BDD7EE")  # light blue header
_GRN_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red
_GRY_FILL = PatternFill("solid", fgColor="EEEEEE")  # gray (n/a)
_SEP_FILL = PatternFill("solid", fgColor="D9E1F2")  # separator 20d→50d
_REF_FILL   = PatternFill("solid", fgColor="FFF2CC")  # yellow for ref rows
_INPUT_FILL = PatternFill("solid", fgColor="FFE599")  # amber for editable input cells
_SURV_FILL  = PatternFill("solid", fgColor="DDEBF7")  # pale blue for N_survivors row

_ATTR_FILLS: dict[str, PatternFill] = {
    "GICS":    PatternFill("solid", fgColor="D9D2E9"),  # light purple
    "Sector2": PatternFill("solid", fgColor="FCE5CD"),  # light peach
    "Zone":    PatternFill("solid", fgColor="D0E0E3"),  # light teal
}

_PCT_FMT = '+0.00;-0.00;"-"'
_CTR     = Alignment(horizontal="center")


def _gain_fill(val: float) -> PatternFill:
    if pd.isna(val):
        return _GRY_FILL
    return _GRN_FILL if val >= 0 else _RED_FILL


# ---------------------------------------------------------------------------
# Operational sheet
# ---------------------------------------------------------------------------

def _fill_operational(ws, hop_results: list[dict], params: dict) -> None:
    """
    Row 1   : (blank)  | daynum1 | daynum2 | ...
    Row 2   : (blank)  | date1   | date2   | ...
    Rows 3…n+2: (blank) | ticker_rank1 … ticker_rankN
    +2 rows : avg_gain20d / avg_gain50d
    +4 rows : market context (^GSPC_rsi, ^STOXX_rsi, ^HSI_rsi, ^VIX)
    +?? rows: GICS count rows (sorted by frequency)
    +?? rows: Sector2 count rows
    +?? rows: Zone count rows
    """
    daynums   = [h["daynum"] for h in hop_results]
    n         = params.get("focusset_size", max(len(h.get("tickers", [])) for h in hop_results))
    n_tickers = n
    threshold = params.get("No_go_GSPC_rsi")

    # Optional N_survivors row, inserted directly below the ticker rows.
    has_surv = any("n_survivors" in h for h in hop_results)
    surv_off = 1 if has_surv else 0

    # Pre-compute the row where ^GSPC_rsi will land, for use in avg formulas.
    rows_list    = _avg_rows(n)
    ref_base     = n_tickers + 3 + surv_off + len(rows_list)
    ref_keys     = list(hop_results[0].get("ref_values", {}).keys()) if hop_results else []
    gspc_rsi_row = next((ref_base + i for i, k in enumerate(ref_keys) if "GSPC_rsi" in k), None)

    # ---- header rows ----
    # A1/A2 hold the No_go label and editable threshold (safe for any focusset size).
    if threshold is not None:
        c = ws.cell(1, 1, "No_go_GSPC_rsi"); c.font = _BOLD
        c = ws.cell(2, 1, threshold);        c.font, c.fill, c.number_format = _BOLD, _INPUT_FILL, "0"
    else:
        ws.cell(1, 1).fill = _HDR_FILL
        ws.cell(2, 1).fill = _HDR_FILL
    for j, dn in enumerate(daynums, start=2):
        c = ws.cell(1, j, dn);                 c.font, c.fill, c.alignment = _BOLD, _HDR_FILL, _CTR
        c = ws.cell(2, j, daynum_to_date(dn)); c.font, c.fill, c.alignment = _SMALL, _HDR_FILL, _CTR

    # ---- ticker rows ----
    for i in range(n_tickers):
        row = i + 3
        ws.cell(row, 1, "")
        for j, h in enumerate(hop_results, start=2):
            tickers = h.get("tickers", [])
            ws.cell(row, j, tickers[i] if i < len(tickers) else "")

    # ---- N_survivors row (optional — only when hops carry the count) ----
    if has_surv:
        srow = n_tickers + 3
        c = ws.cell(srow, 1, "N_survivors"); c.font, c.fill = _BOLD, _SURV_FILL
        for j, h in enumerate(hop_results, start=2):
            val  = h.get("n_survivors")
            cell = ws.cell(srow, j)
            cell.font, cell.fill, cell.alignment = _BOLD, _SURV_FILL, _CTR
            cell.value = int(val) if val is not None else None

    # ---- avg rows ----
    for idx, (label, gain_key, top_n) in enumerate(rows_list):
        row      = n_tickers + 3 + surv_off + idx
        sep      = (gain_key == "gains_50d")
        lbl_cell = ws.cell(row, 1, label)
        lbl_cell.font = _BOLD
        if sep:
            lbl_cell.fill = _SEP_FILL
        for j, h in enumerate(hop_results, start=2):
            val    = _hop_avg_topn(h, gain_key, top_n)
            gspc_p = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
            no_go  = threshold is not None and not pd.isna(gspc_p) and gspc_p < threshold
            cell   = ws.cell(row, j)
            cell.font, cell.number_format, cell.alignment = _BOLD, _PCT_FMT, _CTR

            if pd.notna(val) and gspc_rsi_row is not None:
                # Interactive formula: recalculates when user edits A4
                col_ltr    = get_column_letter(j)
                cell.value = f'=IF({col_ltr}{gspc_rsi_row}<$A$2,"",{round(val, 4)})'
                cell.fill  = _GRY_FILL if no_go else _gain_fill(val)
            elif pd.notna(val):
                cell.value = None if no_go else round(val, 4)
                cell.fill  = _GRY_FILL if no_go else _gain_fill(val)
            else:
                cell.value = None
                cell.fill  = PatternFill() if sep else _GRY_FILL

    # ---- ref rows ----
    def _write_ref_rows(start_row: int, hop_key: str, label_suffix: str = "") -> int:
        keys = list(hop_results[0].get(hop_key, {}).keys()) if hop_results else []
        for idx, key in enumerate(keys):
            lbl = ws.cell(start_row + idx, 1, f"{key}{label_suffix}")
            lbl.font, lbl.fill = _BOLD, _REF_FILL
            for j, h in enumerate(hop_results, start=2):
                val  = h.get(hop_key, {}).get(key, float("nan"))
                cell = ws.cell(start_row + idx, j)
                cell.fill, cell.alignment = _REF_FILL, _CTR
                if pd.notna(val):
                    cell.value         = round(val, 2)
                    cell.number_format = "0.0" if key.endswith("_rsi") else "0.00"
                else:
                    cell.value = None
        return len(keys)

    base   = n_tickers + 3 + surv_off + len(rows_list)
    n_ref = _write_ref_rows(base, "ref_values")

    # ---- attribute frequency rows ----
    stamdata = load_stamdata()
    attr_row = base + n_ref

    for attr_col in ("GICS", "Sector2", "Zone"):
        fill = _ATTR_FILLS[attr_col]
        sorted_vals, hop_counts = _count_attr(hop_results, stamdata, attr_col)

        for i, val in enumerate(sorted_vals):
            row      = attr_row + i
            label    = f"{attr_col}_{val}"
            lbl_cell = ws.cell(row, 1, label)
            lbl_cell.font, lbl_cell.fill = _BOLD, fill
            for j, hc in enumerate(hop_counts, start=2):
                count = hc.get(val, 0)
                cell  = ws.cell(row, j)
                cell.fill, cell.alignment = fill, _CTR
                cell.value = count if count > 0 else None

        attr_row += len(sorted_vals)

    # ---- dimensions ----
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(daynums) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 14
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _fill_summary(ws, strategy_name: str, run_num: int, params: dict,
                  hop_results: list[dict]) -> None:
    daynums = [h["daynum"] for h in hop_results]
    n       = params.get("focusset_size", 10)

    # StartDaynum/EndDaynum are the strategy's USABLE span, chronological (Start = oldest).
    usable   = _usable_daynums(hop_results, _GAIN_KEY, params)
    start_dn = min(usable) if usable else (min(daynums) if daynums else "")
    end_dn   = max(usable) if usable else (max(daynums) if daynums else "")

    rows: list[tuple] = [
        ("StrategyName",  strategy_name),
        ("Run#",          run_num),
        ("StartDaynum",   start_dn),
        ("N_hops",        len(hop_results)),
        ("N_hops_active", _count_active_hops(hop_results, params)),
        ("EndDaynum",     end_dn),
    ]
    for k, v in params.items():
        rows.append((k, v))

    for label, gain_key, top_n in _avg_rows(n):
        val = _grand_avg_topn(hop_results, gain_key, top_n, params)
        rows.append((label, round(val, 4) if pd.notna(val) else None))

    # Realizable non-overlapping additive chain for the active horizon (= period).
    hold = int(params.get("period", 22))
    ret, annual, ntr = _chain_metrics(hop_results, _GAIN_KEY, n, hold, params)
    rows.append(("chain_ret",    round(ret, 4)    if pd.notna(ret)    else None))
    rows.append(("chain_annual", round(annual, 4) if pd.notna(annual) else None))
    rows.append(("chain_n",      ntr))

    worst, n_loss, sens = _chain_dispersion(hop_results, n, hold, params)
    rows.append(("origin_sens%", round(sens, 1) if pd.notna(sens) else None))
    rows.append(("N_loss", n_loss))
    rows.append(("Worst", round(worst, 4) if pd.notna(worst) else None))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    gain_labels = {r[0] for r in _avg_rows(n)} | {"chain_ret", "chain_annual"}
    for i, (k, v) in enumerate(rows, start=1):
        kc = ws.cell(i, 1, k);  kc.font = _BOLD
        vc = ws.cell(i, 2, v)
        if k in gain_labels:
            vc.number_format = _PCT_FMT
            if v is not None:
                vc.fill = _gain_fill(v)


# ---------------------------------------------------------------------------
# HopData sheet — machine-readable per-hop chain inputs
# ---------------------------------------------------------------------------

def _fill_hopdata(ws, hop_results: list[dict], params: dict) -> None:
    """
    Raw per-hop values needed to recompute the realizable chain later over an
    arbitrary daynum window (best_strategy.py clamps to a common floor/cap).

    Stored as plain numbers (not Excel formulas) so they read back reliably.
    Columns: daynum | gain | gspc_rsi  (gain = top-N avg for the active period).
    """
    n = params.get("focusset_size", 10)
    ws.append(["daynum", "gain", "gspc_rsi"])
    for h in hop_results:
        g    = _hop_avg_topn(h, _GAIN_KEY, n)
        gspc = h.get("ref_values", {}).get("^GSPC_rsi", float("nan"))
        ws.append([
            int(h["daynum"]),
            None if pd.isna(g)    else round(float(g), 6),
            None if pd.isna(gspc) else round(float(gspc), 4),
        ])


# ---------------------------------------------------------------------------
# Master summary CSV
# ---------------------------------------------------------------------------

def _append_summary_csv(strategy_name: str, run_num: int, params: dict,
                        hop_results: list[dict]) -> None:
    daynums = [h["daynum"] for h in hop_results]
    n       = params.get("focusset_size", 10)

    def _fmt(val: float) -> str:
        return "" if pd.isna(val) else f"{val:.4f}".replace(".", ",")

    param_cols   = list(params.keys())
    avg_labels   = [r[0] for r in _avg_rows(n)]
    chain_labels = _chain_metric_labels()
    extra_cols   = ["origin_sens%", "N_loss", "Worst"]
    all_cols     = (["StrategyName", "Run#", "StartDaynum", "N_hops", "N_hops_active", "EndDaynum"]
                    + param_cols + avg_labels + chain_labels + extra_cols)

    SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)
    write_header = not SUMMARY_CSV.exists()

    with SUMMARY_CSV.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        if write_header:
            w.writerow(all_cols)
        avg_vals   = [_fmt(_grand_avg_topn(hop_results, gk, tn, params)) for _, gk, tn in _avg_rows(n)]
        hold = int(params.get("period", 22))
        ret, annual, ntr = _chain_metrics(hop_results, _GAIN_KEY, n, hold, params)
        chain_vals = [_fmt(ret), _fmt(annual), str(ntr)]
        worst, n_loss, sens = _chain_dispersion(hop_results, n, hold, params)
        extra_vals = [
            "" if pd.isna(sens) else f"{sens:.1f}".replace(".", ","),
            n_loss,
            _fmt(worst),
        ]
        usable   = _usable_daynums(hop_results, _GAIN_KEY, params)
        start_dn = min(usable) if usable else (min(daynums) if daynums else "")
        end_dn   = max(usable) if usable else (max(daynums) if daynums else "")
        w.writerow(
            [strategy_name, run_num, start_dn, len(hop_results),
             _count_active_hops(hop_results, params), end_dn]
            + [params.get(k, "") for k in param_cols]
            + avg_vals + chain_vals + extra_vals
        )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def save_report(strategy_name: str, params: dict, hop_results: list[dict],
                run_num: int | None = None) -> None:
    """
    Write one Excel file and append one row to the master summary CSV.

    hop_results items contain:
        daynum, tickers (rank-ordered), gains_20d, gains_50d,
        ref_values (market context at daynum)
    """
    folder = REPORT_ROOT / strategy_name
    folder.mkdir(parents=True, exist_ok=True)

    if run_num is None:
        run_num = _next_run_num(folder)

    today     = date.today().strftime("%Y%m%d")
    xlsx_path = folder / f"run{run_num}_{today}.xlsx"

    wb    = Workbook()
    ws_op = wb.active
    ws_op.title = "Operational"
    _fill_operational(ws_op, hop_results, params)

    ws_sum = wb.create_sheet("Summary")
    _fill_summary(ws_sum, strategy_name, run_num, params, hop_results)

    ws_hop = wb.create_sheet("HopData")
    _fill_hopdata(ws_hop, hop_results, params)

    wb.save(xlsx_path)
    _append_summary_csv(strategy_name, run_num, params, hop_results)

    print(f"** Report written: {xlsx_path} **")
