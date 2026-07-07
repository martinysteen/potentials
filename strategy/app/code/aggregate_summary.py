"""
Aggregate the Summary sheets from all run*.xlsx files for one or more strategies.

Usage (from app/code/):
    python aggregate_summary.py                  # all strategies in app/report/
    python aggregate_summary.py "Best ranknow"   # one specific strategy

Output: app/report/<strategy>/aggregated_summary.xlsx
        One row per run, columns = all Summary keys found across runs.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from shared.config import REPORT_ROOT

# ---------------------------------------------------------------------------
# Styles (self-contained — no dependency on report.py)
# ---------------------------------------------------------------------------
_BOLD      = Font(bold=True)
_SMALL     = Font(size=9)
_HDR_FILL  = PatternFill("solid", fgColor="BDD7EE")
_GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
_GRY_FILL  = PatternFill("solid", fgColor="EEEEEE")
_PARAM_FILL = PatternFill("solid", fgColor="FFFF99")  # yellow for simulation parameter headers
_PCT_FMT   = '+0.00;-0.00;"-"'
_CTR       = Alignment(horizontal="center")

_PARAM_COLS = {"focusset_size", "step", "period", "No_go_GSPC_rsi",
               "corner_bins", "vola_keep_frac", "q10_20_min", "q20_50_min"}


# ---------------------------------------------------------------------------
# Core aggregation
# ---------------------------------------------------------------------------

def aggregate_strategy(strategy_name: str) -> Path | None:
    folder    = REPORT_ROOT / strategy_name
    run_files = sorted(folder.glob("run*.xlsx"), key=lambda p: p.name)

    if not run_files:
        print(f"  No run*.xlsx files found in {folder}")
        return None

    records: list[dict] = []
    for path in run_files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"  skip {path.name}: {exc}")
            continue

        if "Summary" not in wb.sheetnames:
            wb.close()
            continue

        ws     = wb["Summary"]
        record: dict[str, object] = {"_file": path.name}
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                record[str(row[0])] = row[1]
        # Always derive Run# from the filename so renames stay consistent.
        try:
            record["Run#"] = int(path.name[3:].split("_")[0])
        except (ValueError, IndexError):
            pass
        records.append(record)
        wb.close()

    if not records:
        print(f"  No Summary sheets found in {folder}")
        return None

    # Sort: focusset_size ASC, step ASC, No_go_GSPC_rsi ASC
    records.sort(key=lambda r: (
        r.get("focusset_size") or 0,
        r.get("step") or 0,
        r.get("No_go_GSPC_rsi") or 0,
    ))

    # Ordered column union across all records (preserve first-encounter order)
    seen: dict[str, None] = {}
    for rec in records:
        for k in rec:
            if k != "_file":
                seen[k] = None
    cols = list(seen.keys())

    # Retired metrics — drop legacy columns still present in older run files.
    cols = [c for c in cols if not c.startswith(("acc_gain", "top"))]

    # Ensure N_hops_active is always adjacent to N_hops regardless of which run introduced it
    if "N_hops_active" in cols and "N_hops" in cols:
        cols.remove("N_hops_active")
        cols.insert(cols.index("N_hops") + 1, "N_hops_active")

    # Keep the dispersion columns grouped to the right of the avg_gain columns
    _trailing = ["origin_sens%", "N_loss", "Worst"]
    for _col in _trailing:
        if _col in cols:
            cols.remove(_col)
            cols.append(_col)

    # ---------------------------------------------------------------------------
    # Write aggregated Excel
    # ---------------------------------------------------------------------------
    out_path = folder / "aggregated_summary.xlsx"
    if out_path.exists():
        out_path.unlink()
    wb_out   = Workbook()
    ws_out   = wb_out.active
    ws_out.title = "Aggregated Summary"

    all_cols = ["source_file"] + cols

    # Header row
    for j, col in enumerate(all_cols, start=1):
        c = ws_out.cell(1, j, col)
        c.font      = _BOLD
        c.fill      = _PARAM_FILL if col in _PARAM_COLS else _HDR_FILL
        c.alignment = _CTR

    # Data rows — one per run
    for i, rec in enumerate(records, start=2):
        ws_out.cell(i, 1, rec.get("_file", ""))
        for j, col in enumerate(cols, start=2):
            val  = rec.get(col)
            cell = ws_out.cell(i, j, val)
            cell.alignment = _CTR
            if ("gain" in col or col.startswith(("chain_ret", "chain_annual"))) and isinstance(val, (int, float)):
                cell.number_format = _PCT_FMT
                cell.fill = _GRN_FILL if val >= 0 else _RED_FILL
            elif col == "Worst" and isinstance(val, (int, float)):
                cell.number_format = _PCT_FMT
                cell.fill = _GRN_FILL if val >= 0 else _RED_FILL
            elif col == "N_loss" and isinstance(val, (int, float)) and val > 0:
                cell.fill = _RED_FILL

    # Column widths
    ws_out.column_dimensions["A"].width = 28   # source filename
    for j, col in enumerate(cols, start=2):
        ws_out.column_dimensions[get_column_letter(j)].width = max(len(col) + 2, 10)

    ws_out.freeze_panes = "B2"   # freeze header row and source-file column
    wb_out.save(out_path)
    print(f"  {len(records)} runs → {out_path.name}")
    return out_path


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) > 1:
        strategies = sys.argv[1:]
    else:
        strategies = [d.name for d in sorted(REPORT_ROOT.iterdir()) if d.is_dir()]

    for name in strategies:
        print(f"Aggregating: {name}")
        aggregate_strategy(name)
    print("Done.")


if __name__ == "__main__":
    main()
